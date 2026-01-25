from __future__ import annotations

from contextlib import asynccontextmanager
import csv
import io
import secrets
import string
from typing import Sequence
from datetime import datetime, timezone
import asyncio
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import Depends, FastAPI, HTTPException, Query, Header, Request
from fastapi.responses import StreamingResponse, HTMLResponse, RedirectResponse, JSONResponse, Response
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.requests import Request
# –í—Ä–µ–º–µ–Ω–Ω–æ –æ—Ç–∫–ª—é—á–µ–Ω–æ –∏–∑-–∑–∞ –ø—Ä–æ–±–ª–µ–º —Å Pydantic
# from slowapi import Limiter, _rate_limit_exceeded_handler
# from slowapi.util import get_remote_address
# from slowapi.errors import RateLimitExceeded
import secrets
import hashlib
import hmac
import time
from datetime import datetime, timezone
from sqlalchemy import select, func, text, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from core.config import get_settings
from core.db.session import engine, get_session, SessionLocal, recreate_engine

logger = logging.getLogger(__name__)
from core.db.models import (
    Base,
    User,
    Subscription,
    SubscriptionPlan,
    Payment,
    SubscriptionStatus,
    PaymentStatus,
    BalanceTransaction,
    AuditLog,
    AuditLogAction,
    AdminOverride,
    Ticket,
    TicketMessage,
    TicketStatus,
    MessageDirection,
    SystemSetting,
    ReferralReward,
    PromoCode,
    PromoCodeUsage,
    Server,
    ServerStatus,
    Backup,
    VpnCredential,
    IpLog,
    UserBan,
    SubscriptionNotification,
)
from core.xray import generate_vless_config, generate_uuid
from core.schemas import (
    PaymentCreateIn,
    PaymentWebhookIn,
    UserOut,
    UserUpsertIn,
    SubscriptionStatusOut,
    AdminCreditIn,
    AdminSetActiveIn,
    ReferralInfoOut,
    AuditLogOut,
    PromoCodeValidateIn,
    PromoCodeApplyIn,
    SubscriptionPurchaseIn,
    SubscriptionTrialIn,
    ServerCreateIn,
    ServerUpdateIn,
    ServerOut,
)

settings = get_settings()


def _check_port_sync(host: str, port: int, timeout: int = 10) -> bool:
    """–°–∏–Ω—Ö—Ä–æ–Ω–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏ –ø–æ—Ä—Ç–∞"""
    import socket
    sock = None
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        # –ù–µ –∏—Å–ø–æ–ª—å–∑—É–µ–º SO_REUSEADDR –¥–ª—è –∫–ª–∏–µ–Ω—Ç—Å–∫–∏—Ö —Å–æ–∫–µ—Ç–æ–≤, —ç—Ç–æ –º–æ–∂–µ—Ç –≤—ã–∑—ã–≤–∞—Ç—å –ø—Ä–æ–±–ª–µ–º—ã
        result = sock.connect_ex((host, port))
        return result == 0
    except socket.timeout:
        logger.debug(f"–¢–∞–π–º–∞—É—Ç –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –ø–æ—Ä—Ç–∞ {host}:{port}")
        return False
    except Exception as e:
        logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –ø–æ—Ä—Ç–∞ {host}:{port}: {e}")
        return False
    finally:
        if sock:
            try:
                sock.close()
            except Exception:
                pass


async def _test_connection_speed(server: Server) -> float | None:
    """
    –¢–µ—Å—Ç–∏—Ä—É–µ—Ç —Å–∫–æ—Ä–æ—Å—Ç—å —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è —Å —Å–µ—Ä–≤–µ—Ä–æ–º —á–µ—Ä–µ–∑ —Ä–µ–∞–ª—å–Ω—É—é –ø–µ—Ä–µ–¥–∞—á—É –¥–∞–Ω–Ω—ã—Ö
    
    –ú–µ—Ç–æ–¥—ã –ø—Ä–æ–≤–µ—Ä–∫–∏ (–≤ –ø–æ—Ä—è–¥–∫–µ –ø—Ä–∏–æ—Ä–∏—Ç–µ—Ç–∞):
    1. iperf3 - —Ç–æ—á–Ω–æ–µ –∏–∑–º–µ—Ä–µ–Ω–∏–µ –ø—Ä–æ–ø—É—Å–∫–Ω–æ–π —Å–ø–æ—Å–æ–±–Ω–æ—Å—Ç–∏ (–µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–µ–Ω)
    2. TCP socket —Å —Ä–µ–∞–ª—å–Ω–æ–π –ø–µ—Ä–µ–¥–∞—á–µ–π –¥–∞–Ω–Ω—ã—Ö - –∞–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π –º–µ—Ç–æ–¥
    """
    import subprocess
    import socket
    import time
    import asyncio
    
    host = server.host
    
    # –ú–µ—Ç–æ–¥ 1: –ò—Å–ø–æ–ª—å–∑—É–µ–º iperf3 –¥–ª—è —Ç–æ—á–Ω–æ–≥–æ –∏–∑–º–µ—Ä–µ–Ω–∏—è —Å–∫–æ—Ä–æ—Å—Ç–∏
    try:
        # iperf3 –æ–±—ã—á–Ω–æ —Ä–∞–±–æ—Ç–∞–µ—Ç –Ω–∞ –ø–æ—Ä—Ç—É 5201
        iperf3_port = 5201
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –¥–æ—Å—Ç—É–ø–µ–Ω –ª–∏ iperf3 server –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ
        # –°–Ω–∞—á–∞–ª–∞ –ø—Ä–æ–≤–µ—Ä—è–µ–º –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å –ø–æ—Ä—Ç–∞
        loop = asyncio.get_event_loop()
        port_check = await loop.run_in_executor(
            None,
            lambda: _check_port_sync(host, iperf3_port, timeout=3)
        )
        
        if port_check:
            # –ü—ã—Ç–∞–µ–º—Å—è –∑–∞–ø—É—Å—Ç–∏—Ç—å iperf3 client –¥–ª—è –∏–∑–º–µ—Ä–µ–Ω–∏—è —Å–∫–æ—Ä–æ—Å—Ç–∏
            # iperf3 -c host -p port -t 5 -f m (5 —Å–µ–∫—É–Ω–¥ —Ç–µ—Å—Ç–∞, —Ä–µ–∑—É–ª—å—Ç–∞—Ç –≤ –ú–±–∏—Ç/—Å)
            iperf3_cmd = [
                'iperf3',
                '-c', host,
                '-p', str(iperf3_port),
                '-t', '5',  # 5 —Å–µ–∫—É–Ω–¥ —Ç–µ—Å—Ç–∞
                '-f', 'm',  # –§–æ—Ä–º–∞—Ç –≤—ã–≤–æ–¥–∞: –ú–±–∏—Ç/—Å
                '--json'  # JSON —Ñ–æ—Ä–º–∞—Ç –¥–ª—è –ø–∞—Ä—Å–∏–Ω–≥–∞
            ]
            
            try:
                result = await asyncio.wait_for(
                    loop.run_in_executor(
                        None,
                        lambda: subprocess.run(iperf3_cmd, capture_output=True, text=True, timeout=10)
                    ),
                    timeout=10
                )
                
                if result.returncode == 0 and result.stdout:
                    # –ü–∞—Ä—Å–∏–º JSON –≤—ã–≤–æ–¥ iperf3
                    import json
                    try:
                        data = json.loads(result.stdout)
                        # iperf3 –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Å–∫–æ—Ä–æ—Å—Ç—å –≤ end.sum_sent.bits_per_second –∏–ª–∏ end.sum_received.bits_per_second
                        if 'end' in data and 'sum_sent' in data['end']:
                            bits_per_second = data['end']['sum_sent'].get('bits_per_second', 0)
                            speed_mbps = bits_per_second / 1_000_000  # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ –ú–±–∏—Ç/—Å
                            logger.info(f"‚úÖ –°–∫–æ—Ä–æ—Å—Ç—å {server.name} —á–µ—Ä–µ–∑ iperf3: {speed_mbps:.2f} –ú–±–∏—Ç/—Å")
                            return speed_mbps
                        elif 'end' in data and 'sum_received' in data['end']:
                            bits_per_second = data['end']['sum_received'].get('bits_per_second', 0)
                            speed_mbps = bits_per_second / 1_000_000
                            logger.info(f"‚úÖ –°–∫–æ—Ä–æ—Å—Ç—å {server.name} —á–µ—Ä–µ–∑ iperf3: {speed_mbps:.2f} –ú–±–∏—Ç/—Å")
                            return speed_mbps
                    except (json.JSONDecodeError, KeyError) as e:
                        logger.debug(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–∞—Ä—Å–∏—Ç—å –≤—ã–≤–æ–¥ iperf3 –¥–ª—è {server.name}: {e}")
            except (subprocess.TimeoutExpired, asyncio.TimeoutError):
                logger.debug(f"–¢–∞–π–º–∞—É—Ç –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ iperf3 –¥–ª—è {server.name}")
            except FileNotFoundError:
                logger.debug(f"iperf3 –Ω–µ —É—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω, –∏—Å–ø–æ–ª—å–∑—É–µ–º –∞–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π –º–µ—Ç–æ–¥")
            except Exception as e:
                logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∑–∞–ø—É—Å–∫–µ iperf3 –¥–ª—è {server.name}: {e}")
    except Exception as e:
        logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ iperf3 –¥–ª—è {server.name}: {e}")
    
    # –ú–µ—Ç–æ–¥ 2: TCP socket —Å —Ä–µ–∞–ª—å–Ω–æ–π –ø–µ—Ä–µ–¥–∞—á–µ–π –¥–∞–Ω–Ω—ã—Ö
    try:
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ø–æ—Ä—Ç —Å–µ—Ä–≤–µ—Ä–∞ –∏–ª–∏ —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã–π –ø–æ—Ä—Ç –¥–ª—è —Ç–µ—Å—Ç–∞
        test_port = server.xray_port or 443
        
        def test_tcp_speed():
            """–¢–µ—Å—Ç —Å–∫–æ—Ä–æ—Å—Ç–∏ —á–µ—Ä–µ–∑ TCP —Å —Ä–µ–∞–ª—å–Ω–æ–π –ø–µ—Ä–µ–¥–∞—á–µ–π –¥–∞–Ω–Ω—ã—Ö"""
            try:
                # –†–∞–∑–º–µ—Ä —Ç–µ—Å—Ç–æ–≤—ã—Ö –¥–∞–Ω–Ω—ã—Ö: 1 –ú–ë
                test_size_bytes = 1024 * 1024
                test_data = b'0' * test_size_bytes
                
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(10)
                
                # –ü–æ–¥–∫–ª—é—á–∞–µ–º—Å—è –∫ —Å–µ—Ä–≤–µ—Ä—É
                connect_start = time.time()
                result = sock.connect_ex((host, test_port))
                if result != 0:
                    sock.close()
                    return None
                
                connect_time = time.time() - connect_start
                
                # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –∏ –∏–∑–º–µ—Ä—è–µ–º –≤—Ä–µ–º—è
                send_start = time.time()
                total_sent = 0
                chunk_size = 8192  # 8 –ö–ë —á–∞–Ω–∫–∏
                
                try:
                    for i in range(0, len(test_data), chunk_size):
                        chunk = test_data[i:i + chunk_size]
                        sent = sock.send(chunk)
                        if sent == 0:
                            break
                        total_sent += sent
                    
                    send_time = time.time() - send_start
                    sock.close()
                    
                    if send_time > 0 and total_sent > 0:
                        # –í—ã—á–∏—Å–ª—è–µ–º —Å–∫–æ—Ä–æ—Å—Ç—å: (–±–∞–π—Ç—ã * 8 –±–∏—Ç) / –≤—Ä–µ–º—è –≤ —Å–µ–∫—É–Ω–¥–∞—Ö / 1_000_000 = –ú–±–∏—Ç/—Å
                        speed_mbps = (total_sent * 8) / send_time / 1_000_000
                        logger.info(f"‚úÖ –°–∫–æ—Ä–æ—Å—Ç—å {server.name} —á–µ—Ä–µ–∑ TCP: {speed_mbps:.2f} –ú–±–∏—Ç/—Å (–æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ {total_sent} –±–∞–π—Ç –∑–∞ {send_time:.2f}—Å)")
                        return speed_mbps
                except socket.timeout:
                    sock.close()
                    return None
                except Exception:
                    sock.close()
                    return None
                    
            except Exception as e:
                logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ TCP —Ç–µ—Å—Ç–µ —Å–∫–æ—Ä–æ—Å—Ç–∏ –¥–ª—è {server.name}: {e}")
                return None
        
        loop = asyncio.get_event_loop()
        speed = await loop.run_in_executor(None, test_tcp_speed)
        return speed
        
    except Exception as e:
        logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ç–µ—Å—Ç–µ —Å–∫–æ—Ä–æ—Å—Ç–∏ –¥–ª—è {server.name}: {e}")
        return None


async def _check_server_status(server: Server) -> dict:
    """
    –ü—Ä–æ—Å—Ç–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ —Å–æ—Å—Ç–æ—è–Ω–∏—è —Å–µ—Ä–≤–µ—Ä–∞: –µ—Å–ª–∏ —Å–µ—Ä–≤–µ—Ä –ø–∏–Ω–≥—É–µ—Ç—Å—è - –æ–Ω –æ–Ω–ª–∞–π–Ω
    
    –ü—Ä–æ–≤–µ—Ä—è–µ—Ç –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å —Å–µ—Ä–≤–µ—Ä–∞ —á–µ—Ä–µ–∑ ping (ICMP)
    """
    import subprocess
    import platform
    import time
    
    host = server.host
    
    logger.debug(f"–ü—Ä–æ–≤–µ—Ä–∫–∞ —Å–µ—Ä–≤–µ—Ä–∞ {server.name} ({host})")
    
    try:
        start_time = time.time()
        
        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–æ–º–∞–Ω–¥—É ping –≤ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏ –æ—Ç –û–°
        if platform.system().lower() == 'windows':
            ping_cmd = ['ping', '-n', '1', '-w', '5000', host]
        else:
            ping_cmd = ['ping', '-c', '1', '-W', '5', host]
        
        # –í—ã–ø–æ–ª–Ω—è–µ–º ping
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            lambda: subprocess.run(ping_cmd, capture_output=True, timeout=10)
        )
        
        response_time_ms = int((time.time() - start_time) * 1000)
        is_online = result.returncode == 0
        
        if is_online:
            logger.info(f"‚úÖ –°–µ—Ä–≤–µ—Ä {server.name}: –æ–Ω–ª–∞–π–Ω (ping —É—Å–ø–µ—à–µ–Ω), –≤—Ä–µ–º—è={response_time_ms}ms")
            # –ï—Å–ª–∏ —Å–µ—Ä–≤–µ—Ä –æ–Ω–ª–∞–π–Ω, –∏–∑–º–µ—Ä—è–µ–º —Å–∫–æ—Ä–æ—Å—Ç—å —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è
            connection_speed_mbps = await _test_connection_speed(server)
        else:
            logger.warning(f"‚ùå –°–µ—Ä–≤–µ—Ä {server.name}: –æ—Ñ—Ñ–ª–∞–π–Ω (ping –Ω–µ –ø—Ä–æ—à–µ–ª), –≤—Ä–µ–º—è={response_time_ms}ms")
            connection_speed_mbps = None
        
        return {
            "is_online": is_online,
            "response_time_ms": response_time_ms,
            "connection_speed_mbps": connection_speed_mbps,
            "error_message": None if is_online else f"–°–µ—Ä–≤–µ—Ä {host} –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω (ping failed)",
        }
    except subprocess.TimeoutExpired:
        logger.warning(f"‚ùå –°–µ—Ä–≤–µ—Ä {server.name}: —Ç–∞–π–º–∞—É—Ç –ø—Ä–∏ ping")
        return {
            "is_online": False,
            "response_time_ms": 10000,
            "connection_speed_mbps": None,
            "error_message": f"–¢–∞–π–º–∞—É—Ç –ø—Ä–∏ ping {host}",
        }
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {e}")
        return {
            "is_online": False,
            "response_time_ms": None,
            "connection_speed_mbps": None,
            "error_message": str(e),
        }


async def _close_old_pending_payments():
    """–ó–∞–∫—Ä—ã–≤–∞–µ—Ç –ø–ª–∞—Ç–µ–∂–∏ —Å–æ —Å—Ç–∞—Ç—É—Å–æ–º pending, –∫–æ—Ç–æ—Ä—ã–µ —Å–æ–∑–¥–∞–Ω—ã –±–æ–ª—å—à–µ —á–∞—Å–∞ –Ω–∞–∑–∞–¥"""
    from core.db.session import SessionLocal
    from datetime import timedelta, timezone
    import logging
    
    async with SessionLocal() as session:
        try:
            # –ù–∞—Ö–æ–¥–∏–º –≤—Å–µ pending –ø–ª–∞—Ç–µ–∂–∏ —Å—Ç–∞—Ä—à–µ 1 —á–∞—Å–∞
            now_utc = datetime.now(timezone.utc)
            one_hour_ago = now_utc - timedelta(hours=1)
            logging.info(f"Checking for pending payments older than {one_hour_ago} (UTC)")
            
            old_pending = await session.scalars(
                select(Payment)
                .where(
                    Payment.status == PaymentStatus.pending,
                    Payment.created_at < one_hour_ago
                )
            )
            
            payments_list = old_pending.all()
            logging.info(f"Found {len(payments_list)} old pending payments to close")
            
            closed_count = 0
            for payment in payments_list:
                old_status = payment.status
                payment.status = PaymentStatus.failed
                
                # –õ–æ–≥–∏—Ä—É–µ–º –∑–∞–∫—Ä—ã—Ç–∏–µ
                user = await session.scalar(select(User).where(User.id == payment.user_id))
                amount_rub = payment.amount_cents / 100
                age_hours = (now_utc - payment.created_at).total_seconds() / 3600
                logging.info(f"Closing payment #{payment.id}: created_at={payment.created_at}, age={age_hours:.2f} hours")
                
                session.add(
                    AuditLog(
                        action=AuditLogAction.payment_status_changed,
                        user_tg_id=user.tg_id if user else None,
                        admin_tg_id=None,
                        details=f"–ü–ª–∞—Ç–µ–∂ #{payment.id} –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –∑–∞–∫—Ä—ã—Ç (pending > 1 —á–∞—Å–∞). –°—Ç–∞—Ç—É—Å: {old_status.value} -> failed. –ü—Ä–æ–≤–∞–π–¥–µ—Ä: {payment.provider}, —Å—É–º–º–∞: {amount_rub:.2f} RUB ({payment.currency})",
                    )
                )
                closed_count += 1
            
            if closed_count > 0:
                await session.commit()
                logging.info(f"‚úÖ Successfully closed {closed_count} old pending payments")
            else:
                logging.info("No old pending payments to close")
        except Exception as e:
            logging.error(f"Error in _close_old_pending_payments: {e}", exc_info=True)
            await session.rollback()


async def _check_servers_health():
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ —Å–æ—Å—Ç–æ—è–Ω–∏—è —Å–µ—Ä–≤–µ—Ä–æ–≤"""
    from core.db.session import SessionLocal
    
    async with SessionLocal() as session:
        servers = await session.scalars(select(Server).where(Server.is_enabled == True))
        
        for server in servers.all():
            try:
                import httpx
                import time
                
                start_time = time.time()
                is_online = False
                response_time_ms = None
                error_message = None
                
                # –ü—Ä–æ—Å—Ç–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç–∏
                try:
                    # –ü—ã—Ç–∞–µ–º—Å—è –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ —Å–µ—Ä–≤–µ—Ä—É
                    host_parts = server.host.split(":")
                    host = host_parts[0]
                    port = int(host_parts[1]) if len(host_parts) > 1 else 80
                    
                    # –ü—Ä–æ–≤–µ—Ä–∫–∞ —á–µ—Ä–µ–∑ socket
                    import socket
                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(5)
                    result = sock.connect_ex((host, port))
                    sock.close()
                    is_online = result == 0
                    
                    elapsed = (time.time() - start_time) * 1000
                    response_time_ms = int(elapsed)
                except Exception as e:
                    error_message = str(e)[:200]
                    is_online = False
                
                # –ü–æ–ª—É—á–∞–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∞–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–π
                active_connections = await session.scalar(
                    select(func.count())
                    .select_from(VpnCredential)
                    .where(VpnCredential.server_id == server.id, VpnCredential.active == True)
                ) or 0
                
                # –°–æ–∑–¥–∞–µ–º –∑–∞–ø–∏—Å—å –æ —Å—Ç–∞—Ç—É—Å–µ
                status = ServerStatus(
                    server_id=server.id,
                    is_online=is_online,
                    response_time_ms=response_time_ms,
                    active_connections=active_connections,
                    error_message=error_message,
                )
                session.add(status)
                await session.commit()
                
            except Exception as e:
                import logging
                logging.error(f"Error checking server {server.id}: {e}")
                continue


async def _create_database_backup(created_by_tg_id: int | None = None) -> Backup | None:
    """–°–æ–∑–¥–∞–Ω–∏–µ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö"""
    import os
    import subprocess
    from pathlib import Path
    
    try:
        # –°–æ–∑–¥–∞–µ–º –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—é –¥–ª—è –±—ç–∫–∞–ø–æ–≤, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        backup_dir = Path("/app/backups")
        backup_dir.mkdir(exist_ok=True)
        
        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –∏–º—è —Ñ–∞–π–ª–∞ —Å –¥–∞—Ç–æ–π –∏ –≤—Ä–µ–º–µ–Ω–µ–º
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"backup_{timestamp}.dump"
        backup_path = backup_dir / backup_filename
        
        # –ü–æ–ª—É—á–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ –ë–î
        db_url = settings.db_url
        # –ü–∞—Ä—Å–∏–º URL: postgresql+asyncpg://user:password@db:5432/vpn
        if "postgresql" in db_url:
            # –ò–∑–≤–ª–µ–∫–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –∏–∑ URL
            import re
            match = re.match(r"postgresql\+asyncpg://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)", db_url)
            if match:
                db_user, db_password, db_host, db_port, db_name = match.groups()
                
                # –ò—Å–ø–æ–ª—å–∑—É–µ–º pg_dump –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –±—ç–∫–∞–ø–∞
                env = os.environ.copy()
                env["PGPASSWORD"] = db_password
                
                # –°–æ–∑–¥–∞–µ–º –±—ç–∫–∞–ø —á–µ—Ä–µ–∑ pg_dump –≤ custom format
                # Custom format –±–æ–ª–µ–µ –Ω–∞–¥–µ–∂–µ–Ω –¥–ª—è –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö
                cmd = [
                    "pg_dump",
                    "-h", db_host,
                    "-p", db_port,
                    "-U", db_user,
                    "-d", db_name,
                    "-F", "c",  # Custom format (—Å–∂–∞—Ç—ã–π, –±–æ–ª–µ–µ –Ω–∞–¥–µ–∂–Ω—ã–π)
                    "--no-owner",  # –ù–µ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—Ç—å –≤–ª–∞–¥–µ–ª—å—Ü–∞ –æ–±—ä–µ–∫—Ç–æ–≤
                    "--no-privileges",  # –ù–µ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—Ç—å –ø—Ä–∏–≤–∏–ª–µ–≥–∏–∏
                    "--no-comments",  # –ù–µ –≤–∫–ª—é—á–∞—Ç—å –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏ (—É–º–µ–Ω—å—à–∞–µ—Ç —Ä–∞–∑–º–µ—Ä)
                    "-f", str(backup_path),
                ]
                
                result = subprocess.run(
                    cmd,
                    env=env,
                    capture_output=True,
                    text=True,
                    timeout=300  # 5 –º–∏–Ω—É—Ç —Ç–∞–π–º–∞—É—Ç
                )
                
                if result.returncode != 0:
                    raise Exception(f"pg_dump failed: {result.stderr}")
                
                # –ü–æ–ª—É—á–∞–µ–º —Ä–∞–∑–º–µ—Ä —Ñ–∞–π–ª–∞
                file_size = backup_path.stat().st_size if backup_path.exists() else 0
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –±—ç–∫–∞–ø–µ –≤ –ë–î
                async with SessionLocal() as session:
                    backup = Backup(
                        backup_type="database",
                        file_path=str(backup_path),
                        file_size_bytes=file_size,
                        status="completed",
                        created_by_tg_id=created_by_tg_id,
                    )
                    session.add(backup)
                    await session.commit()
                    await session.refresh(backup)
                    
                    # –õ–æ–≥–∏—Ä—É–µ–º —É—Å–ø–µ—à–Ω–æ–µ —Å–æ–∑–¥–∞–Ω–∏–µ –±—ç–∫–∞–ø–∞
                    if created_by_tg_id:
                        session.add(
                            AuditLog(
                                action=AuditLogAction.backup_action,
                                admin_tg_id=created_by_tg_id,
                                details=f"–°–æ–∑–¥–∞–Ω –±—ç–∫–∞–ø #{backup.id} (—Ä–∞–∑–º–µ—Ä: {file_size / (1024*1024):.2f} MB)",
                            )
                        )
                        await session.commit()
                    
                    # –£–¥–∞–ª—è–µ–º —Å—Ç–∞—Ä—ã–µ –±—ç–∫–∞–ø—ã (–æ—Å—Ç–∞–≤–ª—è–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–µ 10)
                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ–≤—É—é —Å–µ—Å—Å–∏—é, —á—Ç–æ–±—ã –≥–∞—Ä–∞–Ω—Ç–∏—Ä–æ–≤–∞—Ç—å, —á—Ç–æ –Ω–æ–≤—ã–π –±—ç–∫–∞–ø —É–∂–µ –∑–∞–∫–æ–º–º–∏—á–µ–Ω
                    import logging
                    import sys
                    print("=== STARTING BACKUP CLEANUP ===", file=sys.stderr, flush=True)
                    logging.info("=== STARTING BACKUP CLEANUP ===")
                    
                    async with SessionLocal() as cleanup_session:
                        print("Cleanup session created", file=sys.stderr, flush=True)
                        logging.info("Starting cleanup of old backups...")
                        
                        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –±—ç–∫–∞–ø—ã, –æ—Ç—Å–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –ø–æ –¥–∞—Ç–µ (–Ω–æ–≤—ã–µ –ø–µ—Ä–≤—ã–µ)
                        all_backups_stmt = (
                            select(Backup)
                            .where(Backup.backup_type == "database", Backup.status == "completed")
                            .order_by(Backup.created_at.desc())
                        )
                        all_backups_result = await cleanup_session.scalars(all_backups_stmt)
                        backups_list = list(all_backups_result.all())
                        
                        print(f"Found {len(backups_list)} completed database backups", file=sys.stderr, flush=True)
                        logging.info(f"Found {len(backups_list)} completed database backups")
                        
                        # –ï—Å–ª–∏ –±—ç–∫–∞–ø–æ–≤ –±–æ–ª—å—à–µ 10, —É–¥–∞–ª—è–µ–º –≤—Å–µ –∫—Ä–æ–º–µ –ø–µ—Ä–≤—ã—Ö 10
                        if len(backups_list) > 10:
                            backups_to_delete = backups_list[10:]  # –í—Å–µ –∫—Ä–æ–º–µ –ø–µ—Ä–≤—ã—Ö 10
                            print(f"Need to delete {len(backups_to_delete)} old backups (keeping 10 newest)", file=sys.stderr, flush=True)
                            logging.info(f"Need to delete {len(backups_to_delete)} old backups (keeping 10 newest)")
                            
                            deleted_count = 0
                            for old_backup in backups_to_delete:
                                try:
                                    old_path = Path(old_backup.file_path)
                                    if old_path.exists():
                                        old_path.unlink()
                                        print(f"Deleted backup file: {old_path}", file=sys.stderr, flush=True)
                                        logging.info(f"Deleted backup file: {old_path}")
                                    await cleanup_session.delete(old_backup)
                                    deleted_count += 1
                                    print(f"Deleted backup record: ID={old_backup.id}, created_at={old_backup.created_at}", file=sys.stderr, flush=True)
                                    logging.info(f"Deleted backup record: ID={old_backup.id}, created_at={old_backup.created_at}")
                                except Exception as e:
                                    print(f"ERROR: Could not delete old backup {old_backup.id}: {e}", file=sys.stderr, flush=True)
                                    logging.error(f"Could not delete old backup {old_backup.id}: {e}", exc_info=True)
                            
                            if deleted_count > 0:
                                await cleanup_session.commit()
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –±—ç–∫–∞–ø—ã –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–æ —É–¥–∞–ª–µ–Ω—ã
                                verify_stmt = (
                                    select(Backup)
                                    .where(Backup.backup_type == "database", Backup.status == "completed")
                                    .order_by(Backup.created_at.desc())
                                )
                                verify_result = await cleanup_session.scalars(verify_stmt)
                                remaining_backups = list(verify_result.all())
                                print(f"Verification: {len(remaining_backups)} backups remaining after deletion", file=sys.stderr, flush=True)
                                logging.info(f"Verification: {len(remaining_backups)} backups remaining after deletion")
                                
                                print(f"Successfully deleted {deleted_count} old backups", file=sys.stderr, flush=True)
                                logging.info(f"Successfully deleted {deleted_count} old backups")
                            else:
                                print("WARNING: No backups were deleted", file=sys.stderr, flush=True)
                                logging.warning("No backups were deleted")
                        else:
                            print(f"Only {len(backups_list)} backups found, no cleanup needed (keeping all)", file=sys.stderr, flush=True)
                            logging.info(f"Only {len(backups_list)} backups found, no cleanup needed (keeping all)")
                    
                    print("=== BACKUP CLEANUP COMPLETED ===", file=sys.stderr, flush=True)
                    logging.info("=== BACKUP CLEANUP COMPLETED ===")
                    
                    return backup
        else:
            raise Exception("Only PostgreSQL backups are supported")
            
    except Exception as e:
        import logging
        logging.error(f"Error creating backup: {e}")
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± –æ—à–∏–±–∫–µ
        try:
            async with SessionLocal() as session:
                backup = Backup(
                    backup_type="database",
                    file_path="",
                    file_size_bytes=0,
                    status="failed",
                    error_message=str(e)[:500],
                    created_by_tg_id=created_by_tg_id,
                )
                session.add(backup)
                await session.commit()
        except Exception:
            pass
        
        return None


@asynccontextmanager
async def lifespan(app: FastAPI):
    # –ê–≤—Ç–æ-—Å–æ–∑–¥–∞–Ω–∏–µ —Ç–∞–±–ª–∏—Ü –≤ dev. –î–ª—è –ø—Ä–æ–¥–∞–∫—à–µ–Ω–∞ ‚Äî –º–∏–≥—Ä–∞—Ü–∏–∏ alembic.
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É has_active_subscription, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='has_active_subscription'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE users ADD COLUMN has_active_subscription BOOLEAN NOT NULL DEFAULT FALSE"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_users_has_active_subscription ON users(has_active_subscription)"))
                import logging
                logging.info("Added has_active_subscription column to users table")
        except Exception as e:
            import logging
            logging.warning(f"Could not add has_active_subscription column (may already exist): {e}")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É subscription_ends_at, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='subscription_ends_at'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE users ADD COLUMN subscription_ends_at TIMESTAMP WITH TIME ZONE"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_users_subscription_ends_at ON users(subscription_ends_at)"))
                import logging
                logging.info("Added subscription_ends_at column to users table")
        except Exception as e:
            import logging
            logging.warning(f"Could not add subscription_ends_at column (may already exist): {e}")
        
        # –°–æ–∑–¥–∞–µ–º —Ç–∞–±–ª–∏—Ü—É subscription_notifications, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT table_name FROM information_schema.tables WHERE table_name='subscription_notifications'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("""
                    CREATE TABLE subscription_notifications (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                        subscription_id INTEGER REFERENCES subscriptions(id) ON DELETE CASCADE,
                        notification_type VARCHAR(50) NOT NULL,
                        sent_at TIMESTAMP WITH TIME ZONE NOT NULL DEFAULT NOW()
                    )
                """))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_subscription_notifications_user_id ON subscription_notifications(user_id)"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_subscription_notifications_subscription_id ON subscription_notifications(subscription_id)"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_subscription_notifications_sent_at ON subscription_notifications(sent_at)"))
                import logging
                logging.info("Created subscription_notifications table")
        except Exception as e:
            import logging
            logging.warning(f"Could not create subscription_notifications table (may already exist): {e}")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É auto_renew_subscription, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='auto_renew_subscription'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE users ADD COLUMN auto_renew_subscription BOOLEAN NOT NULL DEFAULT TRUE"))
                import logging
                logging.info("Added auto_renew_subscription column to users table")
            else:
                # –ö–æ–ª–æ–Ω–∫–∞ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç, –æ–±–Ω–æ–≤–ª—è–µ–º NULL –∑–Ω–∞—á–µ–Ω–∏—è –Ω–∞ True
                await conn.execute(text("UPDATE users SET auto_renew_subscription = TRUE WHERE auto_renew_subscription IS NULL"))
                import logging
                logging.info("Updated NULL values in auto_renew_subscription column to TRUE")
        except Exception as e:
            import logging
            logging.warning(f"Could not add/update auto_renew_subscription column (may already exist): {e}")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É selected_server_id, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='selected_server_id'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE users ADD COLUMN selected_server_id INTEGER"))
                await conn.execute(text("ALTER TABLE users ADD CONSTRAINT fk_users_selected_server_id FOREIGN KEY (selected_server_id) REFERENCES servers(id) ON DELETE SET NULL"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_users_selected_server_id ON users(selected_server_id)"))
                import logging
                logging.info("Added selected_server_id column to users table")
        except Exception as e:
            import logging
            logging.warning(f"Could not add selected_server_id column (may already exist): {e}")
        
        # ALTER TYPE ... ADD VALUE –Ω–µ–ª—å–∑—è –≤—ã–ø–æ–ª–Ω—è—Ç—å –≤–Ω—É—Ç—Ä–∏ —Ç—Ä–∞–Ω–∑–∞–∫—Ü–∏–∏ –≤ PostgreSQL
        # Enum –∑–Ω–∞—á–µ–Ω–∏—è –æ–ø—Ä–µ–¥–µ–ª–µ–Ω—ã –≤ models.py –∏ —Å–æ–∑–¥–∞—é—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è —Ç–∞–±–ª–∏—Ü—ã servers, –µ—Å–ª–∏ –∏—Ö –Ω–µ—Ç
        server_columns = [
            ("xray_port", "INTEGER", "DEFAULT 443"),
            ("xray_uuid", "VARCHAR(36)", ""),
            ("xray_flow", "VARCHAR(16)", ""),
            ("xray_network", "VARCHAR(16)", "DEFAULT 'tcp'"),
            ("xray_security", "VARCHAR(16)", "DEFAULT 'tls'"),
            ("xray_sni", "VARCHAR(255)", ""),
            ("xray_reality_public_key", "VARCHAR(255)", ""),
            ("xray_reality_short_id", "VARCHAR(16)", ""),
            ("xray_path", "VARCHAR(255)", ""),
            ("xray_host", "VARCHAR(255)", ""),
            ("x3ui_api_url", "VARCHAR(255)", ""),
            ("x3ui_username", "VARCHAR(64)", ""),
            ("x3ui_password", "VARCHAR(255)", ""),
            ("x3ui_inbound_id", "INTEGER", ""),
        ]
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ —Ç–∞–±–ª–∏—Ü–∞ servers
        try:
            table_exists = await conn.execute(
                text("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name='servers')")
            )
            if not table_exists.scalar():
                import logging
                logging.warning("Table 'servers' does not exist, skipping column migration")
            else:
                # –¢–∞–±–ª–∏—Ü–∞ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç, –¥–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏
                for column_name, column_type, default in server_columns:
                    try:
                        result = await conn.execute(
                            text(f"SELECT column_name FROM information_schema.columns WHERE table_name='servers' AND column_name='{column_name}'")
                        )
                        exists = result.scalar()
                        if not exists:
                            default_clause = f" {default}" if default else ""
                            await conn.execute(text(f"ALTER TABLE servers ADD COLUMN {column_name} {column_type}{default_clause}"))
                            import logging
                            logging.info(f"Added {column_name} column to servers table")
                    except Exception as e:
                        import logging
                        logging.error(f"Error adding {column_name} column: {e}", exc_info=True)
        except Exception as e:
            import logging
            logging.error(f"Error checking servers table: {e}", exc_info=True)
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É connection_speed_mbps –≤ server_status, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='server_status' AND column_name='connection_speed_mbps'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE server_status ADD COLUMN connection_speed_mbps NUMERIC(10, 2)"))
                import logging
                logging.info("Added connection_speed_mbps column to server_status table")
        except Exception as e:
            import logging
            logging.warning(f"Could not add connection_speed_mbps column (may already exist): {e}")
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∫–æ–ª–æ–Ω–∫—É user_uuid –≤ vpn_credentials, –µ—Å–ª–∏ –µ—ë –Ω–µ—Ç
        try:
            result = await conn.execute(
                text("SELECT column_name FROM information_schema.columns WHERE table_name='vpn_credentials' AND column_name='user_uuid'")
            )
            exists = result.scalar()
            if not exists:
                await conn.execute(text("ALTER TABLE vpn_credentials ADD COLUMN user_uuid VARCHAR(36)"))
                await conn.execute(text("CREATE INDEX IF NOT EXISTS idx_vpn_credentials_user_uuid ON vpn_credentials(user_uuid)"))
                import logging
                logging.info("Added user_uuid column to vpn_credentials table")
        except Exception as e:
            import logging
            logging.warning(f"Could not add user_uuid column (may already exist): {e}")
        
    # –¢–∞–±–ª–∏—Ü—ã ip_logs –∏ user_bans —Å–æ–∑–¥–∞—é—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —á–µ—Ä–µ–∑ Base.metadata.create_all
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º —Ñ–æ–Ω–æ–≤—É—é –∑–∞–¥–∞—á—É –¥–ª—è –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞ —Å–µ—Ä–≤–µ—Ä–æ–≤
    async def monitor_servers():
        while True:
            try:
                await asyncio.sleep(60)  # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–∞–∂–¥—É—é –º–∏–Ω—É—Ç—É
                await _check_servers_health()
            except asyncio.CancelledError:
                break
            except Exception as e:
                import logging
                logging.error(f"Error in server monitoring: {e}")
    
    monitor_task = asyncio.create_task(monitor_servers())
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º —Ñ–æ–Ω–æ–≤—É—é –∑–∞–¥–∞—á—É –¥–ª—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏—Ö –±—ç–∫–∞–ø–æ–≤
    async def auto_backup():
        # –î–µ–ª–∞–µ–º –ø–µ—Ä–≤—ã–π –±–µ–∫–∞–ø —Å—Ä–∞–∑—É –ø—Ä–∏ —Å—Ç–∞—Ä—Ç–µ
        try:
            await _create_database_backup(created_by_tg_id=None)
            import logging
            logging.info("Initial backup created on startup")
        except Exception as e:
            import logging
            logging.error(f"Error creating initial backup: {e}")
        
        # –ó–∞—Ç–µ–º –¥–µ–ª–∞–µ–º –±–µ–∫–∞–ø—ã —Ä–∞–∑ –≤ —Å—É—Ç–∫–∏
        while True:
            try:
                await asyncio.sleep(86400)  # –†–∞–∑ –≤ —Å—É—Ç–∫–∏ (24 —á–∞—Å–∞)
                await _create_database_backup(created_by_tg_id=None)
                import logging
                logging.info("Scheduled daily backup created")
            except asyncio.CancelledError:
                break
            except Exception as e:
                import logging
                logging.error(f"Error in auto backup: {e}")
    
    backup_task = asyncio.create_task(auto_backup())
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º —Ñ–æ–Ω–æ–≤—É—é –∑–∞–¥–∞—á—É –¥–ª—è –∑–∞–∫—Ä—ã—Ç–∏—è —Å—Ç–∞—Ä—ã—Ö pending –ø–ª–∞—Ç–µ–∂–µ–π
    async def close_old_pending_payments():
        import logging
        # –ó–∞–ø—É—Å–∫–∞–µ–º —Å—Ä–∞–∑—É –ø—Ä–∏ —Å—Ç–∞—Ä—Ç–µ
        logging.info("Starting payment cleanup task - checking for old pending payments...")
        await _close_old_pending_payments()
        
        while True:
            try:
                await asyncio.sleep(3600)  # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–∞–∂–¥—ã–π —á–∞—Å
                logging.info("Running scheduled payment cleanup...")
                await _close_old_pending_payments()
            except asyncio.CancelledError:
                break
            except Exception as e:
                logging.error(f"Error closing old pending payments: {e}", exc_info=True)
    
    payments_cleanup_task = asyncio.create_task(close_old_pending_payments())
    
    # –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ –∏—Å—Ç–µ—á–µ–Ω–∏—è –ø–æ–¥–ø–∏—Å–æ–∫ –∏ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π
    async def check_expired_subscriptions():
        """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç –∏—Å—Ç–µ—á–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–æ–∫, –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è –∏ —É–¥–∞–ª—è–µ—Ç –∫–ª–∏–µ–Ω—Ç–æ–≤ –∏–∑ 3x-UI"""
        from core.db.session import SessionLocal
        import logging
        from zoneinfo import ZoneInfo
        
        while True:
            try:
                await asyncio.sleep(3600)  # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–∞–∂–¥—ã–π —á–∞—Å
                logging.info("Running scheduled subscription status check...")
                
                async with SessionLocal() as session:
                    try:
                        now = datetime.now(timezone.utc)
                        three_days = timedelta(days=3)
                        one_day = timedelta(days=1)
                        
                        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π —Å –∞–∫—Ç–∏–≤–Ω—ã–º–∏ –ø–æ–¥–ø–∏—Å–∫–∞–º–∏
                        users_with_subs = await session.scalars(
                            select(User)
                            .where(User.has_active_subscription == True)
                            .where(User.subscription_ends_at.isnot(None))
                            .options(selectinload(User.credentials).selectinload(VpnCredential.server))
                        )
                        
                        updated_count = 0
                        notifications_sent = 0
                        clients_deleted = 0
                        
                        for user in users_with_subs.all():
                            if not user.subscription_ends_at:
                                continue
                            
                            time_until_expiry = user.subscription_ends_at - now
                            
                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω—É–∂–Ω–æ –ª–∏ –æ—Ç–ø—Ä–∞–≤–∏—Ç—å —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                            notification_to_send = None
                            if time_until_expiry <= timedelta(0):
                                # –ü–æ–¥–ø–∏—Å–∫–∞ –∏—Å—Ç–µ–∫–ª–∞
                                notification_to_send = "expired"
                            elif timedelta(0) < time_until_expiry <= one_day:
                                # –ú–µ–Ω–µ–µ 1 –¥–Ω—è –¥–æ –∏—Å—Ç–µ—á–µ–Ω–∏—è
                                notification_to_send = "1_day"
                            elif one_day < time_until_expiry <= three_days:
                                # –û—Ç 1 –¥–æ 3 –¥–Ω–µ–π –¥–æ –∏—Å—Ç–µ—á–µ–Ω–∏—è
                                notification_to_send = "3_days"
                            
                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –æ—Ç–ø—Ä–∞–≤–ª—è–ª–∏ –ª–∏ —É–∂–µ —Ç–∞–∫–æ–µ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                            if notification_to_send:
                                active_sub = await session.scalar(
                                    select(Subscription)
                                    .where(Subscription.user_id == user.id)
                                    .where(Subscription.status == SubscriptionStatus.active)
                                    .order_by(Subscription.ends_at.desc().nullslast())
                                    .limit(1)
                                )
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –æ—Ç–ø—Ä–∞–≤–ª—è–ª–∏ –ª–∏ —É–∂–µ —ç—Ç–æ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –¥–ª—è —ç—Ç–æ–π –ø–æ–¥–ø–∏—Å–∫–∏
                                existing_notification = await session.scalar(
                                    select(SubscriptionNotification)
                                    .where(SubscriptionNotification.user_id == user.id)
                                    .where(SubscriptionNotification.notification_type == notification_to_send)
                                    .where(
                                        (SubscriptionNotification.subscription_id == active_sub.id) 
                                        if active_sub else True
                                    )
                                    .order_by(SubscriptionNotification.sent_at.desc())
                                    .limit(1)
                                )
                                
                                # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –Ω–µ –æ—Ç–ø—Ä–∞–≤–ª—è–ª–∏ –Ω–µ–¥–∞–≤–Ω–æ (–¥–ª—è expired - –≤—Å–µ–≥–¥–∞)
                                if not existing_notification or notification_to_send == "expired":
                                    # –§–æ—Ä–º–∏—Ä—É–µ–º —Ç–µ–∫—Å—Ç —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è
                                    ends_at_moscow = user.subscription_ends_at.astimezone(ZoneInfo("Europe/Moscow"))
                                    ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M")
                                    
                                    if notification_to_send == "expired":
                                        notification_text = (
                                            f"‚è∞ <b>–ü–æ–¥–ø–∏—Å–∫–∞ –∏—Å—Ç–µ–∫–ª–∞</b>\n\n"
                                            f"–í–∞—à–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –∑–∞–∫–æ–Ω—á–∏–ª–∞—Å—å {ends_str} –ú–°–ö.\n\n"
                                            f"–î–ª—è –ø—Ä–æ–¥–æ–ª–∂–µ–Ω–∏—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è VPN –ø—Ä–∏–æ–±—Ä–µ—Ç–∏—Ç–µ –Ω–æ–≤—É—é –ø–æ–¥–ø–∏—Å–∫—É –≤ —Ä–∞–∑–¥–µ–ª–µ 'üì¶ –¢–∞—Ä–∏—Ñ—ã'."
                                        )
                                    elif notification_to_send == "1_day":
                                        notification_text = (
                                            f"‚è∞ <b>–ü–æ–¥–ø–∏—Å–∫–∞ —Å–∫–æ—Ä–æ –∏—Å—Ç–µ—á–µ—Ç</b>\n\n"
                                            f"–í–∞—à–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –∑–∞–∫–æ–Ω—á–∏—Ç—Å—è —á–µ—Ä–µ–∑ <b>1 –¥–µ–Ω—å</b> ({ends_str} –ú–°–ö).\n\n"
                                            f"–ù–µ –∑–∞–±—É–¥—å—Ç–µ –ø—Ä–æ–¥–ª–∏—Ç—å –ø–æ–¥–ø–∏—Å–∫—É, —á—Ç–æ–±—ã –Ω–µ –ø–æ—Ç–µ—Ä—è—Ç—å –¥–æ—Å—Ç—É–ø –∫ VPN."
                                        )
                                    elif notification_to_send == "3_days":
                                        notification_text = (
                                            f"‚è∞ <b>–ù–∞–ø–æ–º–∏–Ω–∞–Ω–∏–µ –æ –ø–æ–¥–ø–∏—Å–∫–µ</b>\n\n"
                                            f"–í–∞—à–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –∑–∞–∫–æ–Ω—á–∏—Ç—Å—è —á–µ—Ä–µ–∑ <b>3 –¥–Ω—è</b> ({ends_str} –ú–°–ö).\n\n"
                                            f"–†–µ–∫–æ–º–µ–Ω–¥—É–µ–º –ø—Ä–æ–¥–ª–∏—Ç—å –ø–æ–¥–ø–∏—Å–∫—É –∑–∞—Ä–∞–Ω–µ–µ."
                                        )
                                    
                                    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                                    try:
                                        asyncio.create_task(_send_user_notification(user.tg_id, notification_text))
                                        
                                        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∞–∫—Ç –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è
                                        notification_record = SubscriptionNotification(
                                            user_id=user.id,
                                            subscription_id=active_sub.id if active_sub else None,
                                            notification_type=notification_to_send,
                                        )
                                        session.add(notification_record)
                                        notifications_sent += 1
                                    except Exception as e:
                                        logging.error(f"Error sending notification to user {user.tg_id}: {e}")
                            
                            # –ï—Å–ª–∏ –ø–æ–¥–ø–∏—Å–∫–∞ –∏—Å—Ç–µ–∫–ª–∞, –ø—Ä–æ–≤–µ—Ä—è–µ–º –∞–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ
                            if time_until_expiry <= timedelta(0):
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –≤–∫–ª—é—á–µ–Ω–æ –ª–∏ –∞–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ
                                if user.auto_renew_subscription:
                                    # –ü—ã—Ç–∞–µ–º—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –ø—Ä–æ–¥–ª–∏—Ç—å –ø–æ–¥–ø–∏—Å–∫—É
                                    active_sub = await session.scalar(
                                        select(Subscription)
                                        .where(Subscription.user_id == user.id)
                                        .where(Subscription.status == SubscriptionStatus.active)
                                        .order_by(Subscription.ends_at.desc().nullslast())
                                        .limit(1)
                                    )
                                    
                                    if active_sub:
                                        # –ò—â–µ–º —Ç–∞—Ä–∏—Ñ –¥–ª—è –ø—Ä–æ–¥–ª–µ–Ω–∏—è
                                        # –°–Ω–∞—á–∞–ª–∞ –ø—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ —Ç–∞—Ä–∏—Ñ –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é (–¥–ª—è –æ–±—Ä–∞—Ç–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏)
                                        original_plan = None
                                        if active_sub.plan_name:
                                            # –ü—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ —Ç–∞—Ä–∏—Ñ –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é
                                            plans_by_name = await session.scalars(
                                                select(SubscriptionPlan)
                                                .where(SubscriptionPlan.name == active_sub.plan_name)
                                                .where(SubscriptionPlan.is_active == True)
                                            )
                                            original_plan = plans_by_name.first()
                                        
                                        # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏ –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é, –∏—â–µ–º —Å–∞–º—ã–π –ø–æ—Ö–æ–∂–∏–π –ø–æ —Ü–µ–Ω–µ
                                        if not original_plan and active_sub.price_cents:
                                            plans_by_price = await session.scalars(
                                                select(SubscriptionPlan)
                                                .where(SubscriptionPlan.price_cents == active_sub.price_cents)
                                                .where(SubscriptionPlan.is_active == True)
                                            )
                                            original_plan = plans_by_price.first()
                                        
                                        # –ï—Å–ª–∏ –Ω–µ —Ö–≤–∞—Ç–∞–µ—Ç —Å—Ä–µ–¥—Å—Ç–≤ –Ω–∞ —Ç–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ, –∏—â–µ–º –±–æ–ª–µ–µ –¥–µ—à–µ–≤—ã–π
                                        plan = None
                                        if original_plan and user.balance >= original_plan.price_cents:
                                            # –•–≤–∞—Ç–∞–µ—Ç –Ω–∞ —Ç–µ–∫—É—â–∏–π —Ç–∞—Ä–∏—Ñ
                                            plan = original_plan
                                        else:
                                            # –ò—â–µ–º —Å–∞–º—ã–π –¥–µ—à–µ–≤—ã–π —Ç–∞—Ä–∏—Ñ, –Ω–∞ –∫–æ—Ç–æ—Ä—ã–π —Ö–≤–∞—Ç–∞–µ—Ç —Å—Ä–µ–¥—Å—Ç–≤
                                            all_plans = await session.scalars(
                                                select(SubscriptionPlan)
                                                .where(SubscriptionPlan.is_active == True)
                                                .order_by(SubscriptionPlan.price_cents.asc())
                                            )
                                            
                                            for candidate_plan in all_plans.all():
                                                if user.balance >= candidate_plan.price_cents:
                                                    plan = candidate_plan
                                                    break
                                        
                                        if plan:
                                            # –ü—Ä–æ–¥–ª–µ–≤–∞–µ–º –ø–æ–¥–ø–∏—Å–∫—É –Ω–∞ –Ω–∞–π–¥–µ–Ω–Ω—ã–π —Ç–∞—Ä–∏—Ñ
                                            new_ends_at = now + timedelta(days=plan.days)
                                            active_sub.ends_at = new_ends_at
                                            active_sub.price_cents = plan.price_cents
                                            active_sub.plan_name = plan.name  # –û–±–Ω–æ–≤–ª—è–µ–º –Ω–∞–∑–≤–∞–Ω–∏–µ —Ç–∞—Ä–∏—Ñ–∞
                                            
                                            # –°–ø–∏—Å—ã–≤–∞–µ–º –±–∞–ª–∞–Ω—Å
                                            user.balance -= plan.price_cents
                                            
                                            # –õ–æ–≥–∏—Ä—É–µ–º
                                            session.add(
                                                BalanceTransaction(
                                                    user_id=user.id,
                                                    amount_cents=-plan.price_cents,
                                                    type=BalanceTransactionType.subscription_purchase,
                                                    details=f"–ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∫–∏ '{plan.name}' –Ω–∞ {plan.days} –¥–Ω–µ–π. –ë–∞–ª–∞–Ω—Å: {user.balance / 100:.2f} RUB",
                                                )
                                            )
                                            
                                            # –§–æ—Ä–º–∏—Ä—É–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –¥–ª—è –ª–æ–≥–∞
                                            log_message = f"–ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∫–∏ '{plan.name}' –Ω–∞ {plan.days} –¥–Ω–µ–π"
                                            if original_plan and plan.id != original_plan.id:
                                                log_message += f" (–ø–µ—Ä–µ–∫–ª—é—á–µ–Ω–æ —Å '{original_plan.name}' –∏–∑-–∑–∞ –Ω–µ–¥–æ—Å—Ç–∞—Ç–∫–∞ —Å—Ä–µ–¥—Å—Ç–≤)"
                                            
                                            session.add(
                                                AuditLog(
                                                    action=AuditLogAction.subscription_extended,
                                                    user_tg_id=user.tg_id,
                                                    details=f"{log_message}. –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {new_ends_at.strftime('%d.%m.%Y %H:%M')} (UTC). –ë–∞–ª–∞–Ω—Å: {user.balance / 100:.2f} RUB",
                                                )
                                            )
                                            
                                            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                                            try:
                                                from zoneinfo import ZoneInfo
                                                ends_at_moscow = new_ends_at.astimezone(ZoneInfo("Europe/Moscow"))
                                                ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M")
                                                
                                                notification_text = (
                                                    f"‚úÖ <b>–ü–æ–¥–ø–∏—Å–∫–∞ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –ø—Ä–æ–¥–ª–µ–Ω–∞</b>\n\n"
                                                    f"üì¶ –¢–∞—Ä–∏—Ñ: <b>{plan.name}</b>\n"
                                                    f"üí∞ –°—Ç–æ–∏–º–æ—Å—Ç—å: {plan.price_cents / 100:.2f} RUB\n"
                                                )
                                                
                                                # –ï—Å–ª–∏ —Ç–∞—Ä–∏—Ñ –∏–∑–º–µ–Ω–∏–ª—Å—è, –¥–æ–±–∞–≤–ª—è–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ–± —ç—Ç–æ–º
                                                if original_plan and plan.id != original_plan.id:
                                                    notification_text += (
                                                        f"‚ö†Ô∏è <b>–í–Ω–∏–º–∞–Ω–∏–µ:</b> –¢–∞—Ä–∏—Ñ –∏–∑–º–µ–Ω–µ–Ω —Å '{original_plan.name}' –Ω–∞ '{plan.name}' "
                                                        f"–∏–∑-–∑–∞ –Ω–µ–¥–æ—Å—Ç–∞—Ç–∫–∞ —Å—Ä–µ–¥—Å—Ç–≤ –Ω–∞ –ø—Ä–µ–¥—ã–¥—É—â–∏–π —Ç–∞—Ä–∏—Ñ.\n\n"
                                                    )
                                                
                                                notification_text += (
                                                    f"üìÖ –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_str} –ú–°–ö\n"
                                                    f"üíµ –û—Å—Ç–∞—Ç–æ–∫ –±–∞–ª–∞–Ω—Å–∞: {user.balance / 100:.2f} RUB"
                                                )
                                                
                                                asyncio.create_task(_send_user_notification(user.tg_id, notification_text))
                                            except Exception as e:
                                                logging.error(f"Error sending auto-renewal notification to user {user.tg_id}: {e}")
                                            
                                            logging.info(f"–ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∫–∏ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: –ø—Ä–æ–¥–ª–µ–Ω–æ –Ω–∞ {plan.days} –¥–Ω–µ–π (—Ç–∞—Ä–∏—Ñ: {plan.name})")
                                            updated_count += 1
                                        else:
                                            # –ù–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Å—Ä–µ–¥—Å—Ç–≤ –Ω–∏ –Ω–∞ –æ–¥–∏–Ω —Ç–∞—Ä–∏—Ñ - —É–¥–∞–ª—è–µ–º –∫–ª–∏–µ–Ω—Ç–∞
                                            logging.info(f"–ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –Ω–µ–≤–æ–∑–º–æ–∂–Ω–æ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: –Ω–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ —Å—Ä–µ–¥—Å—Ç–≤ –Ω–∏ –Ω–∞ –æ–¥–∏–Ω —Ç–∞—Ä–∏—Ñ (–±–∞–ª–∞–Ω—Å: {user.balance / 100:.2f} RUB)")
                                            
                                            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –æ –Ω–µ–¥–æ—Å—Ç–∞—Ç–∫–µ —Å—Ä–µ–¥—Å—Ç–≤
                                            try:
                                                from zoneinfo import ZoneInfo
                                                ends_at_moscow = user.subscription_ends_at.astimezone(ZoneInfo("Europe/Moscow")) if user.subscription_ends_at else None
                                                ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M") if ends_at_moscow else "‚Äî"
                                                
                                                # –ü–æ–ª—É—á–∞–µ–º —Å–∞–º—ã–π –¥–µ—à–µ–≤—ã–π —Ç–∞—Ä–∏—Ñ –¥–ª—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏
                                                cheapest_plan = await session.scalar(
                                                    select(SubscriptionPlan)
                                                    .where(SubscriptionPlan.is_active == True)
                                                    .order_by(SubscriptionPlan.price_cents.asc())
                                                    .limit(1)
                                                )
                                                
                                                notification_text = (
                                                    f"‚ùå <b>–ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –Ω–µ —É–¥–∞–ª–æ—Å—å</b>\n\n"
                                                    f"–í–∞—à–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –∏—Å—Ç–µ–∫–ª–∞ {ends_str} –ú–°–ö.\n\n"
                                                    f"üíµ –¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>\n"
                                                )
                                                
                                                if cheapest_plan:
                                                    notification_text += (
                                                        f"üí∞ –ú–∏–Ω–∏–º–∞–ª—å–Ω—ã–π —Ç–∞—Ä–∏—Ñ: <b>{cheapest_plan.name}</b> ‚Äî {cheapest_plan.price_cents / 100:.2f} RUB\n\n"
                                                    )
                                                
                                                notification_text += (
                                                    "–î–ª—è –ø—Ä–æ–¥–æ–ª–∂–µ–Ω–∏—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏—è VPN –ø–æ–ø–æ–ª–Ω–∏—Ç–µ –±–∞–ª–∞–Ω—Å –∏ –ø—Ä–∏–æ–±—Ä–µ—Ç–∏—Ç–µ –ø–æ–¥–ø–∏—Å–∫—É –≤ —Ä–∞–∑–¥–µ–ª–µ 'üì¶ –¢–∞—Ä–∏—Ñ—ã'."
                                                )
                                                
                                                asyncio.create_task(_send_user_notification(user.tg_id, notification_text))
                                            except Exception as e:
                                                logging.error(f"Error sending auto-renewal failure notification to user {user.tg_id}: {e}")
                                            
                                            await _handle_subscription_expiry(user, session)
                                            updated_count += 1
                                    else:
                                        # –ù–µ—Ç –∞–∫—Ç–∏–≤–Ω–æ–π –ø–æ–¥–ø–∏—Å–∫–∏ - —É–¥–∞–ª—è–µ–º –∫–ª–∏–µ–Ω—Ç–∞
                                        await _handle_subscription_expiry(user, session)
                                        updated_count += 1
                                else:
                                    # –ê–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –æ—Ç–∫–ª—é—á–µ–Ω–æ - —É–¥–∞–ª—è–µ–º –∫–ª–∏–µ–Ω—Ç–∞
                                    await _handle_subscription_expiry(user, session)
                                    updated_count += 1
                        
                        # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π
                        all_users = await session.scalars(select(User))
                        for user in all_users.all():
                            old_status = user.has_active_subscription
                            await _update_user_subscription_status(user.id, session)
                            await session.flush()
                            if old_status != user.has_active_subscription and old_status:
                                updated_count += 1
                        
                        if updated_count > 0 or notifications_sent > 0 or clients_deleted > 0:
                            await session.commit()
                            logging.info(f"Updated subscription status for {updated_count} users, sent {notifications_sent} notifications, deleted {clients_deleted} clients from 3x-UI")
                    except Exception as e:
                        logging.error(f"Error checking expired subscriptions: {e}", exc_info=True)
                        await session.rollback()
            except asyncio.CancelledError:
                break
            except Exception as e:
                logging.error(f"Error in subscription status check task: {e}", exc_info=True)
    
    subscription_check_task = asyncio.create_task(check_expired_subscriptions())
    
    # –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏ —Å–æ—Å—Ç–æ—è–Ω–∏—è —Å–µ—Ä–≤–µ—Ä–æ–≤
    async def check_servers_status():
        """–ü—Ä–æ–≤–µ—Ä—è–µ—Ç —Å–æ—Å—Ç–æ—è–Ω–∏–µ —Å–µ—Ä–≤–µ—Ä–æ–≤ (–¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å –ø–æ—Ä—Ç–æ–≤)"""
        from core.db.session import SessionLocal
        import logging
        import time
        
        while True:
            try:
                await asyncio.sleep(60)  # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∫–∞–∂–¥—ã–µ 60 —Å–µ–∫—É–Ω–¥
                logging.info("Running scheduled server status check...")
                
                async with SessionLocal() as session:
                    try:
                        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –∞–∫—Ç–∏–≤–Ω—ã–µ —Å–µ—Ä–≤–µ—Ä—ã
                        servers = await session.scalars(
                            select(Server)
                            .where(Server.is_enabled == True)
                        )
                        servers_list = servers.all()
                        
                        checked_count = 0
                        for server in servers_list:
                            try:
                                port = server.xray_port or 443
                                host = server.host
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –∫–æ–≥–¥–∞ –±—ã–ª–∞ –ø–æ—Å–ª–µ–¥–Ω—è—è –ø—Ä–æ–≤–µ—Ä–∫–∞ —ç—Ç–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
                                last_status = await session.scalar(
                                    select(ServerStatus)
                                    .where(ServerStatus.server_id == server.id)
                                    .order_by(ServerStatus.checked_at.desc())
                                )
                                
                                # –ï—Å–ª–∏ –ø–æ—Å–ª–µ–¥–Ω—è—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –±—ã–ª–∞ –º–µ–Ω–µ–µ 20 —Å–µ–∫—É–Ω–¥ –Ω–∞–∑–∞–¥, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º
                                if last_status and last_status.checked_at:
                                    from datetime import datetime, timezone
                                    time_since_check = (datetime.now(timezone.utc) - last_status.checked_at).total_seconds()
                                    if time_since_check < 20:
                                        logger.debug(f"–ü—Ä–æ–ø—É—Å–∫–∞–µ–º –ø—Ä–æ–≤–µ—Ä–∫—É —Å–µ—Ä–≤–µ—Ä–∞ {server.name}, –ø–æ—Å–ª–µ–¥–Ω—è—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –±—ã–ª–∞ {time_since_check:.1f} —Å–µ–∫—É–Ω–¥ –Ω–∞–∑–∞–¥")
                                        continue
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å –ø–æ—Ä—Ç–∞
                                status_result = await _check_server_status(server)
                                is_online = status_result["is_online"]
                                response_time_ms = status_result["response_time_ms"]
                                error_message = status_result["error_message"]
                                
                                # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å—Ç–∞—Ç—É—Å
                                status = ServerStatus(
                                    server_id=server.id,
                                    is_online=is_online,
                                    response_time_ms=response_time_ms,
                                    connection_speed_mbps=status_result.get("connection_speed_mbps"),
                                    error_message=error_message,
                                )
                                session.add(status)
                                checked_count += 1
                                
                                # –ó–∞–¥–µ—Ä–∂–∫–∞ –º–µ–∂–¥—É –ø—Ä–æ–≤–µ—Ä–∫–∞–º–∏ —Å–µ—Ä–≤–µ—Ä–æ–≤ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è rate limiting
                                await asyncio.sleep(1.0)  # –£–≤–µ–ª–∏—á–µ–Ω–æ –¥–æ 1 —Å–µ–∫—É–Ω–¥—ã
                                
                            except Exception as e:
                                logging.error(f"Error checking server {server.id} ({server.name}): {e}")
                                # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å—Ç–∞—Ç—É—Å —Å –æ—à–∏–±–∫–æ–π
                                status = ServerStatus(
                                    server_id=server.id,
                                    is_online=False,
                                    error_message=f"Check error: {str(e)}",
                                )
                                session.add(status)
                        
                        if checked_count > 0:
                            await session.commit()
                            logging.info(f"Checked status for {checked_count} servers")
                            
                            # –£–¥–∞–ª—è–µ–º —Å—Ç–∞—Ä—ã–µ –∑–∞–ø–∏—Å–∏ (–æ—Å—Ç–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –ø–æ—Å–ª–µ–¥–Ω–∏–µ 100 –Ω–∞ —Å–µ—Ä–≤–µ—Ä)
                            for server in servers_list:
                                old_statuses = await session.scalars(
                                    select(ServerStatus)
                                    .where(ServerStatus.server_id == server.id)
                                    .order_by(ServerStatus.checked_at.desc())
                                    .offset(100)
                                )
                                for old_status in old_statuses.all():
                                    await session.delete(old_status)
                            await session.commit()
                            
                    except Exception as e:
                        logging.error(f"Error checking servers status: {e}", exc_info=True)
                        await session.rollback()
            except asyncio.CancelledError:
                break
            except Exception as e:
                logging.error(f"Error in server status check task: {e}", exc_info=True)
    
    server_check_task = asyncio.create_task(check_servers_status())
    
    # –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –¥–ª—è –º–æ–Ω–∏—Ç–æ—Ä–∏–Ω–≥–∞ IP –∏ –∞–≤—Ç–æ–±–∞–Ω–∞
    async def monitor_client_ips():
        """–ú–æ–Ω–∏—Ç–æ—Ä–∏—Ç IP –∞–¥—Ä–µ—Å–∞ –∫–ª–∏–µ–Ω—Ç–æ–≤ –∏ –±–∞–Ω–∏—Ç –ø—Ä–∏ –ø—Ä–µ–≤—ã—à–µ–Ω–∏–∏ –ª–∏–º–∏—Ç–∞"""
        from core.db.session import SessionLocal
        from core.db.models import Server, VpnCredential, User, IpLog, UserBan, SystemSetting
        from core.x3ui_api import X3UIAPI
        import logging
        
        # –ñ–¥–µ–º 2 –º–∏–Ω—É—Ç—ã –ø–µ—Ä–µ–¥ –ø–µ—Ä–≤—ã–º –∑–∞–ø—É—Å–∫–æ–º
        await asyncio.sleep(120)
        
        while True:
            try:
                async with SessionLocal() as session:
                    try:
                        # –ü–æ–ª—É—á–∞–µ–º –Ω–∞—Å—Ç—Ä–æ–π–∫–∏
                        ip_limit_setting = await session.scalar(
                            select(SystemSetting).where(SystemSetting.key == "vpn_limit_ip")
                        )
                        autoban_enabled_setting = await session.scalar(
                            select(SystemSetting).where(SystemSetting.key == "autoban_enabled")
                        )
                        autoban_duration_setting = await session.scalar(
                            select(SystemSetting).where(SystemSetting.key == "autoban_duration_hours")
                        )
                        
                        ip_limit = 1
                        if ip_limit_setting:
                            try:
                                ip_limit = int(ip_limit_setting.value)
                            except (ValueError, TypeError):
                                ip_limit = 1
                        
                        autoban_enabled = True
                        if autoban_enabled_setting:
                            autoban_enabled = autoban_enabled_setting.value.lower() in ("true", "1", "yes")
                        
                        autoban_duration_hours = 24
                        if autoban_duration_setting:
                            try:
                                autoban_duration_hours = int(autoban_duration_setting.value)
                            except (ValueError, TypeError):
                                autoban_duration_hours = 24
                        
                        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –∞–∫—Ç–∏–≤–Ω—ã–µ —Å–µ—Ä–≤–µ—Ä—ã —Å 3x-UI API
                        servers = await session.scalars(
                            select(Server).where(
                                Server.is_enabled == True,
                                Server.x3ui_api_url.isnot(None),
                                Server.x3ui_username.isnot(None),
                                Server.x3ui_password.isnot(None)
                            )
                        )
                        
                        for server in servers.all():
                            try:
                                x3ui = X3UIAPI(
                                    api_url=server.x3ui_api_url,
                                    username=server.x3ui_username,
                                    password=server.x3ui_password,
                                )
                                
                                try:
                                    # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –∞–∫—Ç–∏–≤–Ω—ã–µ credentials –¥–ª—è —ç—Ç–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
                                    credentials = await session.scalars(
                                        select(VpnCredential)
                                        .where(VpnCredential.server_id == server.id)
                                        .where(VpnCredential.active == True)
                                        .options(selectinload(VpnCredential.user))
                                    )
                                    
                                    for cred in credentials.all():
                                        if not cred.user:
                                            continue
                                        
                                        # –§–æ—Ä–º–∏—Ä—É–µ–º email –∫–ª–∏–µ–Ω—Ç–∞ —Å tg_id
                                        client_email = f"tg_{cred.user.tg_id}_server_{server.id}@fiorevpn"
                                        
                                        # –ü–æ–ª—É—á–∞–µ–º IP –∞–¥—Ä–µ—Å–∞ –∫–ª–∏–µ–Ω—Ç–∞
                                        ips = await x3ui.get_client_ips(client_email)
                                        
                                        if not ips:
                                            continue
                                        
                                        now = datetime.utcnow()
                                        
                                        # –õ–æ–≥–∏—Ä—É–µ–º IP –∞–¥—Ä–µ—Å–∞
                                        for ip in ips:
                                            if not ip or ip == "No IP Record":
                                                continue
                                            
                                            # –ò—â–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –∑–∞–ø–∏—Å—å
                                            existing_log = await session.scalar(
                                                select(IpLog).where(
                                                    IpLog.user_id == cred.user_id,
                                                    IpLog.server_id == server.id,
                                                    IpLog.ip_address == ip
                                                )
                                            )
                                            
                                            if existing_log:
                                                existing_log.last_seen = now
                                                existing_log.connection_count += 1
                                            else:
                                                session.add(IpLog(
                                                    user_id=cred.user_id,
                                                    server_id=server.id,
                                                    ip_address=ip,
                                                    first_seen=now,
                                                    last_seen=now,
                                                    connection_count=1
                                                ))
                                        
                                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø—Ä–µ–≤—ã—à–µ–Ω–∏–µ –ª–∏–º–∏—Ç–∞ IP
                                        if autoban_enabled and len(ips) > ip_limit:
                                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∑–∞–±–∞–Ω–µ–Ω –ª–∏ —É–∂–µ
                                            existing_ban = await session.scalar(
                                                select(UserBan).where(
                                                    UserBan.user_id == cred.user_id,
                                                    UserBan.is_active == True
                                                )
                                            )
                                            
                                            if not existing_ban:
                                                # –°–æ–∑–¥–∞–µ–º –±–∞–Ω
                                                ban = UserBan(
                                                    user_id=cred.user_id,
                                                    reason="ip_limit_exceeded",
                                                    details=f"–û–±–Ω–∞—Ä—É–∂–µ–Ω–æ {len(ips)} IP –∞–¥—Ä–µ—Å–æ–≤ (–ª–∏–º–∏—Ç: {ip_limit}). IP: {', '.join(ips)}",
                                                    is_active=True,
                                                    auto_ban=True,
                                                    banned_until=now + timedelta(hours=autoban_duration_hours)
                                                )
                                                session.add(ban)
                                                
                                                # –û—Ç–∫–ª—é—á–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –≤ 3x-UI
                                                if cred.user_uuid and server.x3ui_inbound_id:
                                                    await x3ui.disable_client(server.x3ui_inbound_id, cred.user_uuid)
                                                
                                                # –£–≤–µ–¥–æ–º–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
                                                notification_text = (
                                                    "‚ö†Ô∏è <b>–í–∞—à –∞–∫–∫–∞—É–Ω—Ç –≤—Ä–µ–º–µ–Ω–Ω–æ –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
                                                    f"–ü—Ä–∏—á–∏–Ω–∞: –ø—Ä–µ–≤—ã—à–µ–Ω –ª–∏–º–∏—Ç –æ–¥–Ω–æ–≤—Ä–µ–º–µ–Ω–Ω—ã—Ö –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–π ({len(ips)} –∏–∑ {ip_limit})\n"
                                                    f"–ë–ª–æ–∫–∏—Ä–æ–≤–∫–∞ —Å–Ω–∏–º–µ—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —á–µ—Ä–µ–∑ {autoban_duration_hours} —á.\n\n"
                                                    "–ï—Å–ª–∏ –≤—ã —Å—á–∏—Ç–∞–µ—Ç–µ —ç—Ç–æ –æ—à–∏–±–∫–æ–π, –æ–±—Ä–∞—Ç–∏—Ç–µ—Å—å –≤ –ø–æ–¥–¥–µ—Ä–∂–∫—É."
                                                )
                                                asyncio.create_task(_send_user_notification(cred.user.tg_id, notification_text))
                                                
                                                logging.warning(
                                                    f"–ê–≤—Ç–æ–±–∞–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {cred.user.tg_id}: "
                                                    f"–ø—Ä–µ–≤—ã—à–µ–Ω –ª–∏–º–∏—Ç IP ({len(ips)} > {ip_limit})"
                                                )
                                        
                                        await session.commit()
                                        
                                finally:
                                    await x3ui.close()
                                    
                            except Exception as e:
                                logging.error(f"Error monitoring IPs for server {server.name}: {e}")
                                continue
                                
                    except Exception as e:
                        logging.error(f"Error in IP monitoring task: {e}", exc_info=True)
                        await session.rollback()
                        
            except asyncio.CancelledError:
                break
            except Exception as e:
                logging.error(f"Error in IP monitoring task: {e}", exc_info=True)
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—ã–µ 5 –º–∏–Ω—É—Ç
            await asyncio.sleep(300)
    
    ip_monitor_task = asyncio.create_task(monitor_client_ips())
    
    # –§–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –¥–ª—è —Å–Ω—è—Ç–∏—è –∏—Å—Ç–µ–∫—à–∏—Ö –±–∞–Ω–æ–≤
    async def unban_expired_users():
        """–°–Ω–∏–º–∞–µ—Ç –±–∞–Ω—ã —Å –∏—Å—Ç–µ–∫—à–∏–º —Å—Ä–æ–∫–æ–º"""
        from core.db.session import SessionLocal
        from core.db.models import UserBan, VpnCredential, Server
        from core.x3ui_api import X3UIAPI
        import logging
        
        # –ñ–¥–µ–º 3 –º–∏–Ω—É—Ç—ã –ø–µ—Ä–µ–¥ –ø–µ—Ä–≤—ã–º –∑–∞–ø—É—Å–∫–æ–º
        await asyncio.sleep(180)
        
        while True:
            try:
                async with SessionLocal() as session:
                    try:
                        now = datetime.utcnow()
                        
                        # –ù–∞—Ö–æ–¥–∏–º –∏—Å—Ç–µ–∫—à–∏–µ –±–∞–Ω—ã
                        expired_bans = await session.scalars(
                            select(UserBan).where(
                                UserBan.is_active == True,
                                UserBan.banned_until.isnot(None),
                                UserBan.banned_until < now
                            )
                        )
                        
                        for ban in expired_bans.all():
                            ban.is_active = False
                            ban.unbanned_at = now
                            
                            # –í–∫–ª—é—á–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –æ–±—Ä–∞—Ç–Ω–æ –≤ 3x-UI
                            credentials = await session.scalars(
                                select(VpnCredential)
                                .where(VpnCredential.user_id == ban.user_id)
                                .where(VpnCredential.active == True)
                                .options(selectinload(VpnCredential.server))
                            )
                            
                            for cred in credentials.all():
                                if not cred.server or not cred.user_uuid:
                                    continue
                                    
                                server = cred.server
                                if server.x3ui_api_url and server.x3ui_username and server.x3ui_password and server.x3ui_inbound_id:
                                    try:
                                        x3ui = X3UIAPI(
                                            api_url=server.x3ui_api_url,
                                            username=server.x3ui_username,
                                            password=server.x3ui_password,
                                        )
                                        try:
                                            await x3ui.enable_client(server.x3ui_inbound_id, cred.user_uuid)
                                        finally:
                                            await x3ui.close()
                                    except Exception as e:
                                        logging.error(f"Error enabling client after unban: {e}")
                            
                            logging.info(f"–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Å–Ω—è—Ç –±–∞–Ω —Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {ban.user_id}")
                        
                        await session.commit()
                        
                    except Exception as e:
                        logging.error(f"Error in unban task: {e}", exc_info=True)
                        await session.rollback()
                        
            except asyncio.CancelledError:
                break
            except Exception as e:
                logging.error(f"Error in unban task: {e}", exc_info=True)
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—ã–µ 10 –º–∏–Ω—É—Ç
            await asyncio.sleep(600)
    
    unban_task = asyncio.create_task(unban_expired_users())
    
    yield
    
    monitor_task.cancel()
    backup_task.cancel()
    payments_cleanup_task.cancel()
    subscription_check_task.cancel()
    server_check_task.cancel()
    ip_monitor_task.cancel()
    unban_task.cancel()
    try:
        await monitor_task
        await backup_task
        await payments_cleanup_task
        await subscription_check_task
        await server_check_task
        await ip_monitor_task
        await unban_task
    except asyncio.CancelledError:
        pass


app = FastAPI(title="fioreVPN Core API", version="0.1.0", lifespan=lifespan)

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è rate limiter (–≤—Ä–µ–º–µ–Ω–Ω–æ –æ—Ç–∫–ª—é—á–µ–Ω –∏–∑-–∑–∞ –ø—Ä–æ–±–ª–µ–º —Å Pydantic)
# limiter = Limiter(key_func=get_remote_address)
# app.state.limiter = limiter
# app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
limiter = None  # –í—Ä–µ–º–µ–Ω–Ω–æ –æ—Ç–∫–ª—é—á–µ–Ω

# –î–æ–±–∞–≤–ª—è–µ–º middleware –¥–ª—è —Å–µ—Å—Å–∏–π
# –ò—Å–ø–æ–ª—å–∑—É–µ–º SECRET_KEY –∏–∑ env, –∏–ª–∏ BOT_TOKEN, –∏–ª–∏ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º —Å–ª—É—á–∞–π–Ω—ã–π
import os
secret_key = os.getenv("SECRET_KEY", "").strip()
if not secret_key:
    bot_token = os.getenv("BOT_TOKEN", "").strip()
    if bot_token:
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º BOT_TOKEN –∫–∞–∫ –æ—Å–Ω–æ–≤—É –¥–ª—è secret_key
        import hashlib
        secret_key = hashlib.sha256(bot_token.encode()).digest()
    else:
        # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º —Å–ª—É—á–∞–π–Ω—ã–π –∫–ª—é—á
        secret_key = secrets.token_urlsafe(32)
app.add_middleware(SessionMiddleware, secret_key=secret_key)

# –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏—è —à–∞–±–ª–æ–Ω–æ–≤ –¥–ª—è –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞
import os
if os.path.exists("core/templates"):
    templates = Jinja2Templates(directory="core/templates")
else:
    templates = None

# –ü–æ–¥–∫–ª—é—á–µ–Ω–∏–µ —Å—Ç–∞—Ç–∏—á–µ—Å–∫–∏—Ö —Ñ–∞–π–ª–æ–≤
if os.path.exists("core/static"):
    app.mount("/static", StaticFiles(directory="core/static"), name="static")




# –û–±—Ä–∞–±–æ—Ç—á–∏–∫–∏ –¥–ª—è —Å—Ç–∞–Ω–¥–∞—Ä—Ç–Ω—ã—Ö –∑–∞–ø—Ä–æ—Å–æ–≤ (—á—Ç–æ–±—ã –Ω–µ –±—ã–ª–æ 404 –≤ –ª–æ–≥–∞—Ö)
@app.get("/")
@app.head("/")
async def root():
    """–ö–æ—Ä–Ω–µ–≤–æ–π –ø—É—Ç—å - —Ä–µ–¥–∏—Ä–µ–∫—Ç –Ω–∞ –∞–¥–º–∏–Ω–∫—É"""
    return RedirectResponse(url="/admin/login", status_code=302)


@app.get("/privacy", response_class=HTMLResponse)
@app.get("/privacy-policy", response_class=HTMLResponse)
async def privacy_policy(request: Request, session: AsyncSession = Depends(get_session)):
    """–ü–æ–ª–∏—Ç–∏–∫–∞ –∫–æ–Ω—Ñ–∏–¥–µ–Ω—Ü–∏–∞–ª—å–Ω–æ—Å—Ç–∏"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ –ø–æ–ª–∏—Ç–∏–∫–∏ –∏ –¥–∞—Ç—É –ø–æ—Å–ª–µ–¥–Ω–µ–≥–æ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –∏–∑ –Ω–∞—Å—Ç—Ä–æ–µ–∫
    privacy_content_setting = await session.scalar(
        select(SystemSetting).where(SystemSetting.key == "privacy_policy_content")
    )
    last_updated_setting = await session.scalar(
        select(SystemSetting).where(SystemSetting.key == "privacy_policy_updated_at")
    )
    
    # –ï—Å–ª–∏ –µ—Å—Ç—å –∫–∞—Å—Ç–æ–º–Ω–æ–µ —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ, –∏—Å–ø–æ–ª—å–∑—É–µ–º –µ–≥–æ, –∏–Ω–∞—á–µ –¥–µ—Ñ–æ–ª—Ç–Ω—ã–π —à–∞–±–ª–æ–Ω
    if privacy_content_setting and privacy_content_setting.value:
        # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –∫–∞—Å—Ç–æ–º–Ω–æ–µ —Å–æ–¥–µ—Ä–∂–∏–º–æ–µ –∫–∞–∫ HTML, –∑–∞–º–µ–Ω—è—è –ø–ª–µ–π—Å—Ö–æ–ª–¥–µ—Ä –¥–∞—Ç—ã
        last_updated = last_updated_setting.value if last_updated_setting else "–ù–µ —É–∫–∞–∑–∞–Ω–æ"
        html_content = privacy_content_setting.value.replace("{{ last_updated }}", last_updated)
        return HTMLResponse(content=html_content)
    
    # –î–µ—Ñ–æ–ª—Ç–Ω—ã–π —à–∞–±–ª–æ–Ω
    last_updated = last_updated_setting.value if last_updated_setting else "–ù–µ —É–∫–∞–∑–∞–Ω–æ"
    return templates.TemplateResponse(
        "privacy.html",
        {
            "request": request,
            "last_updated": last_updated,
        },
    )


@app.get("/robots.txt")
async def robots_txt():
    """Robots.txt –¥–ª—è –ø–æ–∏—Å–∫–æ–≤—ã—Ö –±–æ—Ç–æ–≤"""
    return Response(
        content="User-agent: *\nDisallow: /\n",
        media_type="text/plain"
    )


@app.get("/favicon.ico")
async def favicon_ico():
    """Favicon - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –ª–æ–≥–æ—Ç–∏–ø –µ—Å–ª–∏ –µ—Å—Ç—å"""
    import os
    favicon_path = "core/static/images/logo.png"
    if os.path.exists(favicon_path):
        from fastapi.responses import FileResponse
        return FileResponse(favicon_path, media_type="image/png")
    return Response(status_code=204)


@app.get("/favicon.png")
async def favicon_png():
    """Favicon PNG - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º 204 No Content"""
    return Response(status_code=204)


def _require_admin(x_admin_token: str | None = Header(default=None)) -> None:
    if settings.admin_token and x_admin_token != settings.admin_token:
        raise HTTPException(status_code=403, detail="admin_forbidden")


def _require_admin_or_web(request: Request, x_admin_token: str | None = Header(default=None, alias="X-Admin-Token")) -> dict:
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ —Ç–æ–∫–µ–Ω (–¥–ª—è –±–æ—Ç–∞) –∏–ª–∏ –≤–µ–±-—Å–µ—Å—Å–∏—é (–¥–ª—è –∞–¥–º–∏–Ω–∫–∏)"""
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –ø–µ—Ä–µ–¥–∞–Ω –≤ –∑–∞–≥–æ–ª–æ–≤–∫–µ
    if x_admin_token:
        # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–∞—Å—Ç—Ä–æ–µ–Ω –≤ settings, –ø—Ä–æ–≤–µ—Ä—è–µ–º –µ–≥–æ
        if settings.admin_token:
            if x_admin_token == settings.admin_token:
                return {"tg_id": None, "username": "bot", "first_name": "Bot"}
            # –¢–æ–∫–µ–Ω –ø–µ—Ä–µ–¥–∞–Ω, –Ω–æ –Ω–µ–≤–µ—Ä–Ω—ã–π
            raise HTTPException(status_code=403, detail="invalid_admin_token")
        # –¢–æ–∫–µ–Ω –ø–µ—Ä–µ–¥–∞–Ω, –Ω–æ –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω –≤ settings - —Ä–∞–∑—Ä–µ—à–∞–µ–º (–¥–ª—è –æ–±—Ä–∞—Ç–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏)
        return {"tg_id": None, "username": "bot", "first_name": "Bot"}
    
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω –≤ settings - —Ä–∞–∑—Ä–µ—à–∞–µ–º –¥–æ—Å—Ç—É–ø –±–µ–∑ –ø—Ä–æ–≤–µ—Ä–∫–∏ (–¥–ª—è –æ–±—Ä–∞—Ç–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏)
    # –≠—Ç–æ –ø–æ–∑–≤–æ–ª—è–µ—Ç —Ä–∞–±–æ—Ç–∞—Ç—å –±–µ–∑ —Ç–æ–∫–µ–Ω–∞, –µ—Å–ª–∏ –æ–Ω –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω
    if not settings.admin_token:
        return {"tg_id": None, "username": "bot", "first_name": "Bot"}
    
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–∞—Å—Ç—Ä–æ–µ–Ω, –Ω–æ –Ω–µ –ø–µ—Ä–µ–¥–∞–Ω - –ø—Ä–æ–≤–µ—Ä—è–µ–º –≤–µ–±-—Å–µ—Å—Å–∏—é
    return _require_web_admin(request)


def _require_web_admin(request: Request) -> dict:
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ –≤–µ–±-–∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ —Å–µ—Å—Å–∏—é"""
    session_data = request.session.get("admin_user")
    if not session_data:
        raise HTTPException(status_code=403, detail="not_authenticated")
    return session_data


def _get_csrf_token(request: Request) -> str:
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


async def _require_csrf(request: Request) -> None:
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–µ –≤ –∑–∞–≥–æ–ª–æ–≤–∫–µ, –ø—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –∏–∑ —Ñ–æ—Ä–º—ã
    if not provided:
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        import logging
        logging.warning(f"CSRF validation failed: expected={expected}, provided={provided}, path={request.url.path}")
        raise HTTPException(status_code=403, detail="csrf_forbidden")


def _get_effective_role(tg_id: int, admin_ids: set[int], overrides_map: dict[int, str]) -> str:
    """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç —Ä–æ–ª—å: superadmin|admin|moderator|user"""
    if tg_id in admin_ids:
        return "superadmin"
    ov = overrides_map.get(tg_id)
    if ov in {"admin", "moderator", "user"}:
        return ov
    return "user"


def _role_rank(role: str) -> int:
    """–ß–µ–º –≤—ã—à–µ –∑–Ω–∞—á–µ–Ω–∏–µ, —Ç–µ–º —Å—Ç–∞—Ä—à–µ —Ä–æ–ª—å"""
    ranks = {
        "user": 0,
        "moderator": 1,
        "admin": 2,
        "superadmin": 3,
    }
    return ranks.get(role, 0)


async def _fetch_avatar_url(tg_id: int, bot_token: str) -> str | None:
    """–ü–æ–ª—É—á–∞–µ–º —Å—Å—ã–ª–∫—É –Ω–∞ –∞–≤–∞—Ç–∞—Ä —á–µ—Ä–µ–∑ Telegram getUserProfilePhotos -> getFile"""
    if not bot_token:
        return None
    import httpx
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            r = await client.get(
                f"https://api.telegram.org/bot{bot_token}/getUserProfilePhotos",
                params={"user_id": tg_id, "limit": 1},
            )
            data = r.json()
            if not data.get("ok"):
                return None
            photos = data.get("result", {}).get("photos") or []
            if not photos or not photos[0]:
                return None
            file_id = photos[0][0].get("file_id")
            if not file_id:
                return None
            r2 = await client.get(f"https://api.telegram.org/bot{bot_token}/getFile", params={"file_id": file_id})
            d2 = r2.json()
            if not d2.get("ok"):
                return None
            file_path = d2.get("result", {}).get("file_path")
            if not file_path:
                return None
            return f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
    except Exception:
        return None


@app.post("/tickets/create")
async def create_ticket(payload: dict, session: AsyncSession = Depends(get_session)):
    tg_id = int(payload.get("tg_id") or 0)
    topic = (payload.get("topic") or "").strip()
    if not tg_id or not topic:
        raise HTTPException(status_code=400, detail="invalid_payload")

    now = datetime.utcnow()
    ticket = Ticket(
        user_tg_id=tg_id,
        topic=topic,
        status=TicketStatus.new,
        created_at=now,
        updated_at=now,
    )
    session.add(ticket)
    await session.flush()

    session.add(
        TicketMessage(
            ticket_id=ticket.id,
            user_tg_id=tg_id,
            direction=MessageDirection.system,
            admin_tg_id=None,
            text=f"[–¢–µ–º–∞] {topic}",
        )
    )
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            details=f"Ticket #{ticket.id} created. Topic: {topic}",
        )
    )
    await session.commit()

    # –ü–ª–∞–Ω–∏—Ä—É–µ–º –∞–≤—Ç–æ–∑–∞–∫—Ä—ã—Ç–∏–µ, –µ—Å–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–ø–∏—à–µ—Ç –≤ –ø–æ–¥–¥–µ—Ä–∂–∫—É –≤ —Ç–µ—á–µ–Ω–∏–µ 5 –º–∏–Ω—É—Ç
    async def autoclose(ticket_id: int):
        await asyncio.sleep(300)
        async with SessionLocal() as s:
            t = await s.scalar(select(Ticket).where(Ticket.id == ticket_id))
            if not t or t.status != TicketStatus.open:
                return
            # –µ—Å—Ç—å –ª–∏ –≤—Ö–æ–¥—è—â–∏–µ —Å–æ–æ–±—â–µ–Ω–∏—è (–ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å) –ø–æ–∑–∂–µ —Å–æ–∑–¥–∞–Ω–∏—è
            has_incoming = await s.scalar(
                select(func.count())
                .select_from(TicketMessage)
                .where(
                    TicketMessage.ticket_id == ticket_id,
                    TicketMessage.direction == MessageDirection.incoming,
                    TicketMessage.created_at > t.created_at,
                )
            )
            if has_incoming and has_incoming > 0:
                return
            t.status = TicketStatus.closed
            t.closed_at = datetime.utcnow()
            t.updated_at = t.closed_at
            s.add(
                AuditLog(
                    action=AuditLogAction.admin_action,
                    user_tg_id=t.user_tg_id,
                    details=f"Ticket #{t.id} auto-closed (no user response in 5 minutes)",
                )
            )
            await s.commit()

    asyncio.create_task(autoclose(ticket.id))
    return {"ticket_id": ticket.id}


def _get_csrf_token(request: Request) -> str:
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


async def _require_csrf(request: Request) -> None:
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–µ –≤ –∑–∞–≥–æ–ª–æ–≤–∫–µ, –ø—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –∏–∑ —Ñ–æ—Ä–º—ã
    if not provided:
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        import logging
        logging.warning(f"CSRF validation failed: expected={expected}, provided={provided}, path={request.url.path}")
        raise HTTPException(status_code=403, detail="csrf_forbidden")


def _get_csrf_token(request: Request) -> str:
    token = request.session.get("csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        request.session["csrf_token"] = token
    return token


async def _require_csrf(request: Request) -> None:
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    
    # –ï—Å–ª–∏ —Ç–æ–∫–µ–Ω –Ω–µ –≤ –∑–∞–≥–æ–ª–æ–≤–∫–µ, –ø—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –∏–∑ —Ñ–æ—Ä–º—ã
    if not provided:
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        import logging
        logging.warning(f"CSRF validation failed: expected={expected}, provided={provided}, path={request.url.path}")
        raise HTTPException(status_code=403, detail="csrf_forbidden")


@app.exception_handler(StarletteHTTPException)
async def starlette_http_exception_handler(request: Request, exc: StarletteHTTPException):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ HTTP –∏—Å–∫–ª—é—á–µ–Ω–∏–π –æ—Ç Starlette (–≤–∫–ª—é—á–∞—è 404 –∏ 405)"""
    import logging
    # –î–ª—è 405 (Method Not Allowed) - —Ç–∏—Ö–æ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –æ—Ç–≤–µ—Ç –±–µ–∑ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –æ—à–∏–±–∫–∏
    # –≠—Ç–æ –æ–±—ã—á–Ω–æ —Å–∫–∞–Ω–∏—Ä–æ–≤–∞–Ω–∏–µ –ø–æ—Ä—Ç–æ–≤ –∏–ª–∏ –∞—Ç–∞–∫–∏ –∏–∑–≤–Ω–µ
    if exc.status_code == 405:
        return JSONResponse(
            status_code=405,
            content={"detail": "Method Not Allowed"}
        )
    
    # –î–ª—è 404 –æ—à–∏–±–æ–∫
    if exc.status_code == 404:
        path = request.url.path
        logging.info(f"404 handler: path={path}")
        
        # –î–ª—è API endpoints –≤–æ–∑–≤—Ä–∞—â–∞–µ–º JSON (—Ç–æ–ª—å–∫–æ —á–∏—Å—Ç—ã–µ API –ø—É—Ç–∏, –±–µ–∑ /admin/web)
        is_api_path = (
            path.startswith("/api/") or 
            path.startswith("/subscriptions/") or 
            path.startswith("/payments/") or 
            (path.startswith("/users/") and not path.startswith("/admin/web")) or 
            path.startswith("/promo-codes/") or
            (path.startswith("/tickets/") and not path.startswith("/admin/web")) or
            path.startswith("/health") or
            path.startswith("/support/")
        )
        
        logging.info(f"404 handler: is_api_path={is_api_path}, templates={templates is not None}")
        
        if is_api_path:
            return JSONResponse(
                status_code=404,
                content={"detail": "Not found"}
            )
        
        # –î–ª—è –≤—Å–µ—Ö –æ—Å—Ç–∞–ª—å–Ω—ã—Ö –ø—É—Ç–µ–π (–≤–∫–ª—é—á–∞—è /admin/web/*) –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –∫—Ä–∞—Å–∏–≤—É—é 404
        if templates:
            logging.info(f"404 handler: returning 404.html template")
            return templates.TemplateResponse(
                "404.html",
                {"request": request},
                status_code=404
            )
        # –ï—Å–ª–∏ —à–∞–±–ª–æ–Ω—ã –Ω–µ –∑–∞–≥—Ä—É–∂–µ–Ω—ã, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –ø—Ä–æ—Å—Ç–æ–π —Ç–µ–∫—Å—Ç
        logging.info(f"404 handler: templates not loaded, returning simple HTML")
        return HTMLResponse(
            content="<h1>404 - –°—Ç—Ä–∞–Ω–∏—Ü–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞</h1><p><a href='/admin/login'>–ü–µ—Ä–µ–π—Ç–∏ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å</a></p>",
            status_code=404
        )
    
    # –î–ª—è –¥—Ä—É–≥–∏—Ö HTTP –æ—à–∏–±–æ–∫ –æ—Ç Starlette –ø—Ä–æ–±—Ä–∞—Å—ã–≤–∞–µ–º –¥–∞–ª—å—à–µ (–∫—Ä–æ–º–µ 405, –∫–æ—Ç–æ—Ä—ã–π —É–∂–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω)
    if exc.status_code != 405:
        raise exc


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ HTTP –∏—Å–∫–ª—é—á–µ–Ω–∏–π –æ—Ç FastAPI"""
    if exc.status_code == 403:
        # –î–ª—è –≤–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞ - —Ä–µ–¥–∏—Ä–µ–∫—Ç –Ω–∞ –ª–æ–≥–∏–Ω
        if request.url.path.startswith("/admin/web"):
            return RedirectResponse(url="/admin/login", status_code=303)
        # –î–ª—è API - –≤–æ–∑–≤—Ä–∞—â–∞–µ–º JSON
        return JSONResponse(
            status_code=403,
            content={"detail": exc.detail}
        )
    
    # –î–ª—è 405 (Method Not Allowed) - —Ç–∏—Ö–æ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –æ—Ç–≤–µ—Ç –±–µ–∑ –ª–æ–≥–∏—Ä–æ–≤–∞–Ω–∏—è –æ—à–∏–±–∫–∏
    if exc.status_code == 405:
        # –≠—Ç–æ –æ–±—ã—á–Ω–æ —Å–∫–∞–Ω–∏—Ä–æ–≤–∞–Ω–∏–µ –ø–æ—Ä—Ç–æ–≤ –∏–ª–∏ –∞—Ç–∞–∫–∏, –Ω–µ –ª–æ–≥–∏—Ä—É–µ–º –∫–∞–∫ –æ—à–∏–±–∫—É
        return JSONResponse(
            status_code=405,
            content={"detail": "Method Not Allowed"}
        )
    
    # –î–ª—è –≤—Å–µ—Ö –æ—Å—Ç–∞–ª—å–Ω—ã—Ö HTTPException –≤–æ–∑–≤—Ä–∞—â–∞–µ–º JSON –¥–ª—è API endpoints
    if request.url.path.startswith("/api/") or request.url.path.startswith("/subscriptions/") or request.url.path.startswith("/payments/") or request.url.path.startswith("/users/"):
        return JSONResponse(
            status_code=exc.status_code,
            content={"detail": exc.detail}
        )
    # –î–ª—è –≤–µ–±-—Å—Ç—Ä–∞–Ω–∏—Ü –ø—Ä–æ–±—Ä–∞—Å—ã–≤–∞–µ–º –∏—Å–∫–ª—é—á–µ–Ω–∏–µ –¥–∞–ª—å—à–µ
    raise exc


def _gen_ref_code() -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(8))


async def _ensure_ref_code_unique(session: AsyncSession) -> str:
    for _ in range(20):
        code = _gen_ref_code()
        exists = await session.scalar(select(User.id).where(User.referral_code == code))
        if not exists:
            return code
    raise RuntimeError("failed_to_generate_ref_code")


async def _get_referral_reward_referrer_amount(session: AsyncSession) -> int:
    """–ü–æ–ª—É—á–∞–µ—Ç —Å—É–º–º—É —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω–æ–π –Ω–∞–≥—Ä–∞–¥—ã –¥–ª—è –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–µ–≥–æ –∏–∑ –Ω–∞—Å—Ç—Ä–æ–µ–∫ (–≤ –∫–æ–ø–µ–π–∫–∞—Ö)"""
    setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == "referral_reward_referrer_cents"))
    if setting:
        try:
            return int(setting.value)
        except (ValueError, TypeError):
            pass
    # –ó–Ω–∞—á–µ–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é: 10000 –∫–æ–ø–µ–µ–∫ = 100 RUB
    return 10000


async def _get_referral_reward_referred_amount(session: AsyncSession) -> int:
    """–ü–æ–ª—É—á–∞–µ—Ç —Å—É–º–º—É —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω–æ–π –Ω–∞–≥—Ä–∞–¥—ã –¥–ª—è –ø—Ä–∏–≥–ª–∞—à–µ–Ω–Ω–æ–≥–æ –∏–∑ –Ω–∞—Å—Ç—Ä–æ–µ–∫ (–≤ –∫–æ–ø–µ–π–∫–∞—Ö)"""
    setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == "referral_reward_referred_cents"))
    if setting:
        try:
            return int(setting.value)
        except (ValueError, TypeError):
            pass
    # –ó–Ω–∞—á–µ–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é: 10000 –∫–æ–ø–µ–µ–∫ = 100 RUB
    return 10000


async def _update_user_subscription_status(user_id: int, session: AsyncSession) -> None:
    """–û–±–Ω–æ–≤–ª—è–µ—Ç –ø–æ–ª—è has_active_subscription –∏ subscription_ends_at —É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –Ω–∞ –æ—Å–Ω–æ–≤–µ –∞–∫—Ç–∏–≤–Ω—ã—Ö –ø–æ–¥–ø–∏—Å–æ–∫"""
    from datetime import datetime, timezone
    
    user = await session.scalar(select(User).where(User.id == user_id))
    if not user:
        return
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –∞–∫—Ç–∏–≤–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞, –∫–æ—Ç–æ—Ä–∞—è –µ—â–µ –Ω–µ –∏—Å—Ç–µ–∫–ª–∞
    now = datetime.now(timezone.utc)
    active_sub = await session.scalar(
        select(Subscription)
        .where(Subscription.user_id == user_id)
        .where(Subscription.status == SubscriptionStatus.active)
        .where((Subscription.ends_at.is_(None)) | (Subscription.ends_at > now))
        .order_by(Subscription.ends_at.desc().nullslast())
        .limit(1)
    )
    
    # –û–±–Ω–æ–≤–ª—è–µ–º –ø–æ–ª—è
    if active_sub:
        user.has_active_subscription = True
        user.subscription_ends_at = active_sub.ends_at
    else:
        user.has_active_subscription = False
        user.subscription_ends_at = None
        # –û—á–∏—â–∞–µ–º –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Å–µ—Ä–≤–µ—Ä, –µ—Å–ª–∏ –ø–æ–¥–ø–∏—Å–∫–∞ –Ω–µ –∞–∫—Ç–∏–≤–Ω–∞
        user.selected_server_id = None


async def _validate_promo_code(code: str, user_id: int, amount_cents: int, session: AsyncSession, check_percent_usage: bool = False) -> tuple[bool, str, int]:
    """
    –ü—Ä–æ–≤–µ—Ä—è–µ—Ç –ø—Ä–æ–º–æ–∫–æ–¥ –∏ –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç (is_valid, error_message, discount_cents)
    check_percent_usage: –µ—Å–ª–∏ True, –ø—Ä–æ–≤–µ—Ä—è–µ—Ç, –Ω–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∂–µ –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Å–∫–∏–¥–∫—É
    """
    promo = await session.scalar(select(PromoCode).where(PromoCode.code == code.upper().strip()))
    if not promo:
        return False, "–ü—Ä–æ–º–æ–∫–æ–¥ –Ω–µ –Ω–∞–π–¥–µ–Ω", 0
    if not promo.is_active:
        return False, "–ü—Ä–æ–º–æ–∫–æ–¥ –Ω–µ–∞–∫—Ç–∏–≤–µ–Ω", 0
    
    now = datetime.utcnow()
    if promo.valid_from and now < promo.valid_from:
        return False, "–ü—Ä–æ–º–æ–∫–æ–¥ –µ—â–µ –Ω–µ –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª–µ–Ω", 0
    if promo.valid_until and now > promo.valid_until:
        return False, "–ü—Ä–æ–º–æ–∫–æ–¥ –∏—Å—Ç–µ–∫", 0
    
    if promo.max_uses and promo.used_count >= promo.max_uses:
        return False, "–ü—Ä–æ–º–æ–∫–æ–¥ –∏—Å—á–µ—Ä–ø–∞–Ω", 0
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∂–µ —ç—Ç–æ—Ç –ø—Ä–æ–º–æ–∫–æ–¥
    existing_usage = await session.scalar(
        select(PromoCodeUsage)
        .where(PromoCodeUsage.promo_code_id == promo.id)
        .where(PromoCodeUsage.user_id == user_id)
    )
    if existing_usage:
        return False, "–í—ã —É–∂–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª–∏ —ç—Ç–æ—Ç –ø—Ä–æ–º–æ–∫–æ–¥", 0
    
    # –ï—Å–ª–∏ —ç—Ç–æ –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Å–∫–∏–¥–∫—É (–ø—Ä–æ—Ü–µ–Ω—Ç), –ø—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∂–µ –¥—Ä—É–≥–æ–π –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Å–∫–∏–¥–∫—É
    if check_percent_usage and promo.discount_percent:
        other_percent_usage = await session.scalar(
            select(PromoCodeUsage)
            .join(PromoCode)
            .where(PromoCode.discount_percent.isnot(None))
            .where(PromoCodeUsage.user_id == user_id)
            .where(PromoCodeUsage.promo_code_id != promo.id)
        )
        if other_percent_usage:
            return False, "–í—ã —É–∂–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª–∏ –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Å–∫–∏–¥–∫—É. –ù–µ–ª—å–∑—è –ø—Ä–∏–º–µ–Ω–∏—Ç—å –Ω–µ—Å–∫–æ–ª—å–∫–æ –ø—Ä–æ–º–æ–∫–æ–¥–æ–≤ –Ω–∞ —Å–∫–∏–¥–∫—É", 0
    
    # –í—ã—á–∏—Å–ª—è–µ–º —Å–∫–∏–¥–∫—É
    discount_cents = 0
    if promo.discount_percent:
        discount_cents = int(amount_cents * promo.discount_percent / 100)
    elif promo.discount_amount_cents:
        discount_cents = promo.discount_amount_cents
        # –î–ª—è —Ñ–∏–∫—Å —Å—É–º–º—ã –Ω–µ –æ–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º —Å—É–º–º–æ–π –ø–ª–∞—Ç–µ–∂–∞ - —ç—Ç–æ –±–æ–Ω—É—Å –Ω–∞ –±–∞–ª–∞–Ω—Å
    
    return True, "", discount_cents


@app.post("/promo-codes/validate")
async def validate_promo_code(
    payload: PromoCodeValidateIn,
    session: AsyncSession = Depends(get_session),
):
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ –ø—Ä–æ–º–æ–∫–æ–¥–∞"""
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    is_valid, error_msg, discount_cents = await _validate_promo_code(
        payload.code, user.id, payload.amount_cents, session
    )
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–æ–º–æ–∫–æ–¥–µ –¥–ª—è –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è —Ç–∏–ø–∞
    promo = None
    promo_type = None
    discount_percent = None
    discount_amount_cents = None
    if is_valid:
        promo = await session.scalar(select(PromoCode).where(PromoCode.code == payload.code.upper().strip()))
        if promo:
            if promo.discount_percent:
                promo_type = "percent"
                discount_percent = promo.discount_percent
            elif promo.discount_amount_cents:
                promo_type = "fixed"
                discount_amount_cents = promo.discount_amount_cents
    
    return {
        "valid": is_valid,
        "error": error_msg if not is_valid else None,
        "discount_cents": discount_cents,
        "promo_type": promo_type,
        "discount_percent": discount_percent,
        "discount_amount_cents": discount_amount_cents,
    }


@app.post("/promo-codes/apply")
async def apply_promo_code(
    payload: PromoCodeApplyIn,
    session: AsyncSession = Depends(get_session),
    admin_user: dict | None = Depends(_require_admin_or_web),
):
    """–ü—Ä–∏–º–µ–Ω–µ–Ω–∏–µ –ø—Ä–æ–º–æ–∫–æ–¥–∞"""
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    # –ù–∞—Ö–æ–¥–∏–º –ø—Ä–æ–º–æ–∫–æ–¥
    promo = await session.scalar(select(PromoCode).where(PromoCode.code == payload.code.upper().strip()))
    if not promo:
        return {"success": False, "error": "–ü—Ä–æ–º–æ–∫–æ–¥ –Ω–µ –Ω–∞–π–¥–µ–Ω"}
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø—Ä–æ–º–æ–∫–æ–¥ (–¥–ª—è –ø—Ä–æ–º–æ–∫–æ–¥–æ–≤ –Ω–∞ —Å–∫–∏–¥–∫—É –ø—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–ª –ª–∏ —É–∂–µ –¥—Ä—É–≥–æ–π –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Å–∫–∏–¥–∫—É)
    is_valid, error_msg, discount_cents = await _validate_promo_code(
        payload.code, user.id, payload.amount_cents, session, check_percent_usage=bool(promo.discount_percent)
    )
    
    if not is_valid:
        return {"success": False, "error": error_msg}
    
    # –ï—Å–ª–∏ —ç—Ç–æ –ø—Ä–æ–º–æ–∫–æ–¥ –Ω–∞ —Ñ–∏–∫—Å —Å—É–º–º—É, –Ω–∞—á–∏—Å–ª—è–µ–º –±–∞–ª–∞–Ω—Å —Å—Ä–∞–∑—É
    if promo.discount_amount_cents and not promo.discount_percent:
        user.balance += promo.discount_amount_cents
        session.add(
            BalanceTransaction(
                user_id=user.id,
                admin_tg_id=None,
                amount=promo.discount_amount_cents,
                reason=f"–ü—Ä–æ–º–æ–∫–æ–¥ {promo.code}",
            )
        )
        # –õ–æ–≥–∏—Ä—É–µ–º –≤ –∞–¥–º–∏–Ω–∫–µ
        admin_tg_id = admin_user.get("tg_id") if admin_user else None
        session.add(
            AuditLog(
                action=AuditLogAction.admin_action,
                user_tg_id=user.tg_id,
                admin_tg_id=admin_tg_id,
                details=f"–ü—Ä–∏–º–µ–Ω–µ–Ω –ø—Ä–æ–º–æ–∫–æ–¥ {promo.code} (—Ñ–∏–∫—Å —Å—É–º–º–∞: {promo.discount_amount_cents / 100:.2f} RUB). –ë–∞–ª–∞–Ω—Å –ø–æ–ø–æ–ª–Ω–µ–Ω.",
            )
        )
    
    # –°–æ–∑–¥–∞–µ–º –∑–∞–ø–∏—Å—å –æ–± –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–∏
    usage = PromoCodeUsage(
        promo_code_id=promo.id,
        user_id=user.id,
        discount_amount_cents=discount_cents,
    )
    session.add(usage)
    
    # –£–≤–µ–ª–∏—á–∏–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–π
    promo.used_count += 1
    
    await session.commit()
    
    return {
        "success": True,
        "discount_cents": discount_cents,
        "promo_type": "fixed" if promo.discount_amount_cents and not promo.discount_percent else "percent" if promo.discount_percent else None,
    }


@app.get("/health")
async def health_check(session: AsyncSession = Depends(get_session)):
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ —Ä–∞–±–æ—Ç–æ—Å–ø–æ—Å–æ–±–Ω–æ—Å—Ç–∏ API –∏ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ –ë–î"""
    try:
        # –ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ –ë–î
        await session.execute(select(1))
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "database": "disconnected", "error": str(e)}


@app.get("/users")
async def list_users(
    session: AsyncSession = Depends(get_session),
    limit: int = Query(20, ge=1, le=100),
    offset: int = Query(0, ge=0),
) -> list[UserOut]:
    stmt = select(User).options(selectinload(User.referred_by)).order_by(User.created_at.desc()).limit(limit).offset(offset)
    result = await session.scalars(stmt)
    users: Sequence[User] = result.all()
    out: list[UserOut] = []
    for u in users:
        referred_by_tg_id = u.referred_by.tg_id if u.referred_by else None
        out.append(
            UserOut(
                id=u.id,
                tg_id=u.tg_id,
                username=u.username,
                first_name=u.first_name,
                last_name=u.last_name,
                is_active=u.is_active,
                balance=u.balance,
                referral_code=u.referral_code,
                referred_by_tg_id=referred_by_tg_id,
                trial_used=u.trial_used,
                has_active_subscription=u.has_active_subscription,
                subscription_ends_at=u.subscription_ends_at,
                selected_server_id=u.selected_server_id,
                created_at=u.created_at,
            )
        )
    return out


@app.get("/users/count")
async def users_count(session: AsyncSession = Depends(get_session)) -> dict[str, int]:
    total = await session.scalar(select(func.count()).select_from(User))
    return {"total": int(total or 0)}


@app.put("/users/by_tg/{tg_id}/auto-renew")
async def toggle_auto_renew(
    tg_id: int,
    payload: dict,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–í–∫–ª—é—á–∏—Ç—å/–≤—ã–∫–ª—é—á–∏—Ç—å –∞–≤—Ç–æ–ø—Ä–æ–¥–ª–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∫–∏"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    auto_renew = payload.get("auto_renew", True)
    user.auto_renew_subscription = bool(auto_renew)
    await session.commit()
    
    return {
        "tg_id": tg_id,
        "auto_renew_subscription": user.auto_renew_subscription,
    }


@app.get("/users/by_tg/{tg_id}")
async def get_user_by_tg(tg_id: int, session: AsyncSession = Depends(get_session)) -> UserOut:
    stmt = select(User).options(selectinload(User.referred_by)).where(User.tg_id == tg_id)
    user = await session.scalar(stmt)
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    referred_by_tg_id = user.referred_by.tg_id if user.referred_by else None
    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Å–ª—É—á–∞–π, –∫–æ–≥–¥–∞ auto_renew_subscription –º–æ–∂–µ—Ç –±—ã—Ç—å None
    auto_renew = user.auto_renew_subscription if user.auto_renew_subscription is not None else True
    return UserOut(
        id=user.id,
        tg_id=user.tg_id,
        username=user.username,
        first_name=user.first_name,
        last_name=user.last_name,
        is_active=user.is_active,
        balance=user.balance,
        referral_code=user.referral_code,
        referred_by_tg_id=referred_by_tg_id,
        trial_used=user.trial_used,
        has_active_subscription=user.has_active_subscription,
        subscription_ends_at=user.subscription_ends_at,
        selected_server_id=user.selected_server_id,
        auto_renew_subscription=auto_renew,
        created_at=user.created_at,
    )


@app.post("/users/upsert")
async def upsert_user(payload: UserUpsertIn, session: AsyncSession = Depends(get_session)) -> UserOut:
    stmt = select(User).options(selectinload(User.referred_by)).where(User.tg_id == payload.tg_id)
    user = await session.scalar(stmt)
    is_new = False
    if not user:
        is_new = True
        user = User(tg_id=payload.tg_id, is_active=True, balance=0)
        user.referral_code = await _ensure_ref_code_unique(session)

        if payload.referral_code:
            ref = await session.scalar(select(User).where(User.referral_code == payload.referral_code))
            if ref and ref.tg_id != payload.tg_id:
                user.referred_by = ref

        session.add(user)
        await session.flush()  # –ü–æ–ª—É—á–∞–µ–º ID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
        session.add(
            AuditLog(
                action=AuditLogAction.user_registered,
                user_tg_id=payload.tg_id,
                details=f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω. –†–µ—Ñ–µ—Ä–∞–ª—å–Ω—ã–π –∫–æ–¥: {user.referral_code}",
            )
        )
        
        # –ù–∞—á–∏—Å–ª—è–µ–º —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω—É—é –Ω–∞–≥—Ä–∞–¥—É –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–µ–º—É –∏ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–Ω–æ–º—É
        if user.referred_by:
            referrer_reward_cents = await _get_referral_reward_referrer_amount(session)
            referred_reward_cents = await _get_referral_reward_referred_amount(session)
            
            # –ù–∞–≥—Ä–∞–¥–∞ –¥–ª—è –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–µ–≥–æ
            if referrer_reward_cents > 0:
                user.referred_by.balance += referrer_reward_cents
                session.add(
                    ReferralReward(
                        referrer_user_id=user.referred_by.id,
                        referred_user_id=user.id,
                        amount_cents=referrer_reward_cents,
                        is_for_referrer=True,
                    )
                )
                session.add(
                    BalanceTransaction(
                        user_id=user.referred_by.id,
                        amount=referrer_reward_cents,
                        reason=f"–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –Ω–∞–≥—Ä–∞–¥–∞ –∑–∞ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {payload.tg_id}",
                    )
                )
                session.add(
                    AuditLog(
                        action=AuditLogAction.balance_credited,
                        user_tg_id=user.referred_by.tg_id,
                        details=f"–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –Ω–∞–≥—Ä–∞–¥–∞: {referrer_reward_cents / 100:.2f} RUB –∑–∞ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {payload.tg_id}",
                    )
                )
                # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–µ–º—É (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ)
                notify_on_referral = await session.scalar(select(SystemSetting).where(SystemSetting.key == "notify_on_referral"))
                if not notify_on_referral or notify_on_referral.value != "false":
                    notification_text = (
                        f"üéÅ <b>–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –Ω–∞–≥—Ä–∞–¥–∞!</b>\n\n"
                        f"–í—ã –ø–æ–ª—É—á–∏–ª–∏ <b>{referrer_reward_cents / 100:.2f} RUB</b> –∑–∞ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è.\n"
                        f"–í–∞—à –±–∞–ª–∞–Ω—Å: <b>{user.referred_by.balance / 100:.2f} RUB</b>"
                    )
                    asyncio.create_task(_send_user_notification(user.referred_by.tg_id, notification_text))
            
            # –ù–∞–≥—Ä–∞–¥–∞ –¥–ª—è –ø—Ä–∏–≥–ª–∞—à–µ–Ω–Ω–æ–≥–æ
            if referred_reward_cents > 0:
                user.balance += referred_reward_cents
                session.add(
                    ReferralReward(
                        referrer_user_id=user.referred_by.id,
                        referred_user_id=user.id,
                        amount_cents=referred_reward_cents,
                        is_for_referrer=False,
                    )
                )
                session.add(
                    BalanceTransaction(
                        user_id=user.id,
                        amount=referred_reward_cents,
                        reason=f"–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –Ω–∞–≥—Ä–∞–¥–∞ –∑–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—é –ø–æ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–∏—é",
                    )
                )
                session.add(
                    AuditLog(
                        action=AuditLogAction.balance_credited,
                        user_tg_id=payload.tg_id,
                        details=f"–†–µ—Ñ–µ—Ä–∞–ª—å–Ω–∞—è –Ω–∞–≥—Ä–∞–¥–∞: {referred_reward_cents / 100:.2f} RUB –∑–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—é –ø–æ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–∏—é",
                    )
                )
                # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–Ω–æ–º—É (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ)
                notify_on_referral = await session.scalar(select(SystemSetting).where(SystemSetting.key == "notify_on_referral"))
                if not notify_on_referral or notify_on_referral.value != "false":
                    notification_text = (
                        f"üéÅ <b>–î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å!</b>\n\n"
                        f"–í—ã –ø–æ–ª—É—á–∏–ª–∏ <b>{referred_reward_cents / 100:.2f} RUB</b> –∑–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—é –ø–æ —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω–æ–π —Å—Å—ã–ª–∫–µ.\n"
                        f"–í–∞—à –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>"
                    )
                    asyncio.create_task(_send_user_notification(payload.tg_id, notification_text))
    # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (username –º–æ–∂–µ—Ç –∏–∑–º–µ–Ω–∏—Ç—å—Å—è)
    if payload.username is not None:
        user.username = payload.username
    if payload.first_name is not None:
        user.first_name = payload.first_name
    if payload.last_name is not None:
        user.last_name = payload.last_name
    
    if is_new:
        await session.commit()
    else:
        await session.commit()
    await session.refresh(user)
    # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º —Å relationship –ø–æ—Å–ª–µ refresh
    user = await session.scalar(stmt)
    referred_by_tg_id = user.referred_by.tg_id if user.referred_by else None
    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Å–ª—É—á–∞–π, –∫–æ–≥–¥–∞ auto_renew_subscription –º–æ–∂–µ—Ç –±—ã—Ç—å None
    auto_renew = user.auto_renew_subscription if user.auto_renew_subscription is not None else True
    return UserOut(
        id=user.id,
        tg_id=user.tg_id,
        username=user.username,
        first_name=user.first_name,
        last_name=user.last_name,
        is_active=user.is_active,
        balance=user.balance,
        referral_code=user.referral_code,
        referred_by_tg_id=referred_by_tg_id,
        trial_used=user.trial_used,
        has_active_subscription=user.has_active_subscription,
        subscription_ends_at=user.subscription_ends_at,
        selected_server_id=user.selected_server_id,
        auto_renew_subscription=auto_renew,
        created_at=user.created_at,
    )


@app.get("/subscriptions")
async def list_subscriptions(session: AsyncSession = Depends(get_session)) -> list[dict]:
    result = await session.scalars(select(Subscription))
    subs: Sequence[Subscription] = result.all()
    return [
        {
            "id": s.id,
            "user_id": s.user_id,
            "plan_name": s.plan_name,
            "status": s.status,
            "starts_at": s.starts_at,
            "ends_at": s.ends_at,
        }
        for s in subs
    ]


@app.get("/users/{tg_id}/referral/rewards")
async def get_user_referral_rewards(
    tg_id: int,
    limit: int = Query(default=10, ge=1, le=50),
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å –∏—Å—Ç–æ—Ä–∏—é —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω—ã—Ö –Ω–∞–≥—Ä–∞–¥ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    # –ü–æ–ª—É—á–∞–µ–º –Ω–∞–≥—Ä–∞–¥—ã, –≥–¥–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –±—ã–ª –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–∏–º
    rewards_result = await session.scalars(
        select(ReferralReward)
        .where(ReferralReward.referrer_user_id == user.id)
        .order_by(ReferralReward.created_at.desc())
        .limit(limit)
    )
    rewards = rewards_result.all()
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–∏–≥–ª–∞—à–µ–Ω–Ω—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è—Ö
    referred_ids = [r.referred_user_id for r in rewards]
    referred_users = {}
    if referred_ids:
        users_result = await session.scalars(select(User).where(User.id.in_(referred_ids)))
        for u in users_result.all():
            referred_users[u.id] = u
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    return [
        {
            "id": r.id,
            "amount": r.amount_cents / 100,
            "referred_user_tg_id": referred_users.get(r.referred_user_id, User(tg_id=0)).tg_id if r.referred_user_id in referred_users else 0,
            "is_for_referrer": r.is_for_referrer,
            "created_at": fmt(r.created_at),
        }
        for r in rewards
    ]


@app.get("/users/{tg_id}/payments")
async def get_user_payments(
    tg_id: int,
    limit: int = Query(default=10, ge=1, le=200),
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å –∏—Å—Ç–æ—Ä–∏—é –ø–ª–∞—Ç–µ–∂–µ–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    payments_result = await session.scalars(
        select(Payment)
        .where(Payment.user_id == user.id)
        .order_by(Payment.created_at.desc())
        .limit(limit)
    )
    payments = payments_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    # amount_cents —É–∂–µ —Ö—Ä–∞–Ω–∏—Ç—Å—è –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
    return [
        {
            "id": p.id,
            "provider": p.provider,
            "amount": p.amount_cents / 100,  # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –∫–æ–ø–µ–π–∫–∏ –≤ —Ä—É–±–ª–∏
            "amount_cents": p.amount_cents,
            "currency": p.currency,
            "status": p.status.value if hasattr(p.status, "value") else str(p.status),
            "created_at": fmt(p.created_at),
        }
        for p in payments
    ]


@app.get("/api/payments/{payment_id}")
async def get_payment_detail(
    payment_id: int,
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å –¥–µ—Ç–∞–ª–∏ –ø–ª–∞—Ç–µ–∂–∞ –ø–æ ID"""
    payment = await session.scalar(select(Payment).where(Payment.id == payment_id))
    if not payment:
        raise HTTPException(status_code=404, detail="payment_not_found")
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    return {
        "id": payment.id,
        "provider": payment.provider,
        "amount_cents": payment.amount_cents,
        "currency": payment.currency,
        "status": payment.status.value if hasattr(payment.status, "value") else str(payment.status),
        "external_id": payment.external_id,
        "raw_response": payment.raw_response,
        "created_at": fmt(payment.created_at),
    }


@app.get("/payments")
async def list_payments(session: AsyncSession = Depends(get_session)) -> list[dict]:
    result = await session.scalars(select(Payment))
    pays: Sequence[Payment] = result.all()
    return [
        {
            "id": p.id,
            "user_id": p.user_id,
            "provider": p.provider,
            "amount_cents": p.amount_cents,
            "currency": p.currency,
            "status": p.status,
            "created_at": p.created_at,
        }
        for p in pays
    ]


@app.get("/admin/payments")
async def admin_list_payments(
    limit: int = Query(default=20, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    status: str | None = Query(default=None),
    provider: str | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    admin_user: dict | None = Depends(_require_admin_or_web),
) -> dict:
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –ø–ª–∞—Ç–µ–∂–µ–π –¥–ª—è –∞–¥–º–∏–Ω–∫–∏"""
    from sqlalchemy import or_
    
    stmt = select(Payment).options(selectinload(Payment.user))
    
    # –§–∏–ª—å—Ç—Ä—ã
    if status:
        try:
            status_enum = PaymentStatus[status]
            stmt = stmt.where(Payment.status == status_enum)
        except (KeyError, ValueError):
            pass
    
    if provider:
        stmt = stmt.where(Payment.provider == provider)
    
    # –ü–æ–¥—Å—á–µ—Ç –æ–±—â–µ–≥–æ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞
    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = await session.scalar(count_stmt)
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–ª–∞—Ç–µ–∂–∏ —Å –ø–∞–≥–∏–Ω–∞—Ü–∏–µ–π
    stmt = stmt.order_by(Payment.created_at.desc()).limit(limit).offset(offset)
    payments_result = await session.scalars(stmt)
    payments = payments_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    payments_data = []
    for p in payments:
        user_tg_id = p.user.tg_id if p.user else None
        payments_data.append({
            "id": p.id,
            "user_tg_id": user_tg_id,
            "provider": p.provider,
            "amount": p.amount_cents / 100,  # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –∫–æ–ø–µ–π–∫–∏ –≤ —Ä—É–±–ª–∏
            "amount_cents": p.amount_cents,
            "currency": p.currency,
            "status": p.status.value if hasattr(p.status, "value") else str(p.status),
            "created_at": fmt(p.created_at),
        })
    
    return {
        "payments": payments_data,
        "total": total or 0,
        "limit": limit,
        "offset": offset,
    }


@app.post("/payments/create")
async def create_payment(
    payload: PaymentCreateIn,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–°–æ–∑–¥–∞–Ω–∏–µ –ø–ª–∞—Ç–µ–∂–∞ –¥–ª—è –ø–æ–ø–æ–ª–Ω–µ–Ω–∏—è –±–∞–ª–∞–Ω—Å–∞"""
    from core.config import get_settings
    settings = get_settings()
    
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    # –°–æ–∑–¥–∞–µ–º –∑–∞–ø–∏—Å—å –æ –ø–ª–∞—Ç–µ–∂–µ
    payment = Payment(
        user_id=user.id,
        provider=payload.provider,
        amount_cents=payload.amount_cents,
        currency=payload.currency,
        status=PaymentStatus.pending,
    )
    session.add(payment)
    await session.commit()
    await session.refresh(payment)
    
    # –õ–æ–≥–∏—Ä—É–µ–º —Å–æ–∑–¥–∞–Ω–∏–µ –ø–ª–∞—Ç–µ–∂–∞
    amount_rub = payload.amount_cents / 100
    session.add(
        AuditLog(
            action=AuditLogAction.payment_created,
            user_tg_id=user.tg_id,
            admin_tg_id=None,
            details=f"–°–æ–∑–¥–∞–Ω –ø–ª–∞—Ç–µ–∂ #{payment.id} —á–µ—Ä–µ–∑ {payload.provider}. –°—É–º–º–∞: {amount_rub:.2f} RUB ({payload.currency}). –°—Ç–∞—Ç—É—Å: {payment.status.value}",
        )
    )
    await session.commit()
    
    result = {
        "payment_id": payment.id,
        "provider": payment.provider,
        "amount_cents": payment.amount_cents,
        "currency": payment.currency,
        "status": payment.status.value,
    }
    
    # –ï—Å–ª–∏ —ç—Ç–æ CryptoBot, —Å–æ–∑–¥–∞–µ–º –∏–Ω–≤–æ–π—Å
    if payload.provider == "cryptobot" and settings.cryptobot_token:
        try:
            from core.cryptobot import CryptoBotAPI
            cryptobot = CryptoBotAPI(settings.cryptobot_token)
            
            # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —Ä—É–±–ª–∏ –≤ USD –¥–ª—è CryptoBot
            # payload.amount_cents —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
            amount_rub = payload.amount_cents / 100
            
            # –ü–æ–ª—É—á–∞–µ–º –∫—É—Ä—Å USD/RUB –¥–ª—è –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏
            from core.currency import get_usd_to_rub_rate
            usd_rate = await get_usd_to_rub_rate()
            amount_usd = amount_rub / usd_rate
            
            # CryptoBot —Ç—Ä–µ–±—É–µ—Ç –º–∏–Ω–∏–º—É–º 0.01 USD –¥–ª—è –∏–Ω–≤–æ–π—Å–∞
            MIN_INVOICE_AMOUNT_USD = 0.01
            min_rub = MIN_INVOICE_AMOUNT_USD * usd_rate
            if amount_usd < MIN_INVOICE_AMOUNT_USD:
                raise HTTPException(
                    status_code=400,
                    detail=f"–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —Å—É–º–º–∞ –ø–æ–ø–æ–ª–Ω–µ–Ω–∏—è: {min_rub:.0f} RUB (—ç–∫–≤–∏–≤–∞–ª–µ–Ω—Ç {MIN_INVOICE_AMOUNT_USD} USD)"
                )
            
            # –ü–æ–ª—É—á–∞–µ–º –∫—É—Ä—Å –∫—Ä–∏–ø—Ç–æ–≤–∞–ª—é—Ç—ã –∫ USD —á–µ—Ä–µ–∑ CryptoBot API
            try:
                exchange_rates = await cryptobot.get_exchange_rates()
                if exchange_rates.get("ok") and exchange_rates.get("result"):
                    rates = exchange_rates["result"]
                    # –ò—â–µ–º –∫—É—Ä—Å –¥–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–π –≤–∞–ª—é—Ç—ã
                    currency_rate = None
                    for rate in rates:
                        if isinstance(rate, dict):
                            source = rate.get("source")
                            target = rate.get("target")
                            rate_value = rate.get("rate")
                            
                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Ç–∏–ø—ã –ø–µ—Ä–µ–¥ —Å—Ä–∞–≤–Ω–µ–Ω–∏–µ–º
                            if (isinstance(source, str) and isinstance(target, str) and 
                                source == payload.currency and target == "USD"):
                                # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º rate –≤ —á–∏—Å–ª–æ, –µ—Å–ª–∏ —ç—Ç–æ —Å—Ç—Ä–æ–∫–∞
                                try:
                                    currency_rate = float(rate_value) if rate_value else None
                                except (ValueError, TypeError):
                                    currency_rate = None
                                break
                    
                    if currency_rate and isinstance(currency_rate, (int, float)) and currency_rate > 0:
                        # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º: USD -> –∫—Ä–∏–ø—Ç–æ–≤–∞–ª—é—Ç–∞
                        crypto_amount = amount_usd / currency_rate
                    else:
                        # Fallback: —É–ø—Ä–æ—â–µ–Ω–Ω–∞—è –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è
                        crypto_amount = amount_usd
                else:
                    # Fallback: —É–ø—Ä–æ—â–µ–Ω–Ω–∞—è –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏—è
                    crypto_amount = amount_usd
            except Exception as e:
                import logging
                logging.warning(f"Failed to get exchange rates, using fallback: {e}")
                # Fallback: –¥–ª—è USDT –∏—Å–ø–æ–ª—å–∑—É–µ–º –∫—É—Ä—Å 1:1, –¥–ª—è –¥—Ä—É–≥–∏—Ö –≤–∞–ª—é—Ç –Ω—É–∂–µ–Ω –∫—É—Ä—Å
                # –ù–æ –ª—É—á—à–µ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å —É–ø—Ä–æ—â–µ–Ω–Ω—ã–π –∫—É—Ä—Å, —á–µ–º –Ω–µ–ø—Ä–∞–≤–∏–ª—å–Ω—É—é —Å—É–º–º—É
                if payload.currency == "USDT":
                    # USDT –ø—Ä–∏–≤—è–∑–∞–Ω –∫ USD, –ø–æ—ç—Ç–æ–º—É 1 USD ‚âà 1 USDT
                    crypto_amount = amount_usd
                else:
                    # –î–ª—è –¥—Ä—É–≥–∏—Ö –≤–∞–ª—é—Ç –Ω—É–∂–µ–Ω –∫—É—Ä—Å, –∏—Å–ø–æ–ª—å–∑—É–µ–º –∫–æ–Ω—Å–µ—Ä–≤–∞—Ç–∏–≤–Ω—ã–π fallback
                    logging.error(f"Cannot convert to {payload.currency} without exchange rate")
                    crypto_amount = amount_usd  # –í—Ä–µ–º–µ–Ω–Ω–æ–µ —Ä–µ—à–µ–Ω–∏–µ
            
            invoice = await cryptobot.create_invoice(
                amount=crypto_amount,
                currency=payload.currency,
                description=f"–ü–æ–ø–æ–ª–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ fioreVPN –Ω–∞ {amount_rub:.2f} RUB",
                paid_btn_name="callback",
                # paid_btn_url –Ω–µ –ø–µ—Ä–µ–¥–∞–µ–º, –µ—Å–ª–∏ –Ω–µ –Ω—É–∂–µ–Ω
                payload=f"payment_{payment.id}",
            )
            
            if invoice.get("ok") and invoice.get("result"):
                invoice_data = invoice["result"]
                invoice_id = invoice_data.get("invoice_id")
                payment.external_id = str(invoice_id) if invoice_id else None
                await session.commit()
                
                result["invoice_url"] = invoice_data.get("pay_url")
                result["invoice_id"] = invoice_id
                
                import logging
                logging.info("=" * 80)
                logging.info(f"=== PAYMENT CREATED ===")
                logging.info(f"Payment ID: #{payment.id}")
                logging.info(f"Invoice ID (external_id): {invoice_id} (type: {type(invoice_id).__name__})")
                logging.info(f"Invoice URL: {result.get('invoice_url')}")
                logging.info(f"User ID: {user.id}, User TG ID: {user.tg_id}")
                logging.info(f"Amount: {payment.amount_cents} cents ({payment.amount_cents / 100:.2f} RUB)")
                logging.info(f"Provider: {payment.provider}, Currency: {payment.currency}")
                logging.info("=" * 80)
        except httpx.ReadTimeout as e:
            import logging
            logging.error(f"CryptoBot API timeout when creating invoice: {e}")
            raise HTTPException(
                status_code=504,
                detail="–ü—Ä–µ–≤—ã—à–µ–Ω–æ –≤—Ä–µ–º—è –æ–∂–∏–¥–∞–Ω–∏—è –æ—Ç–≤–µ—Ç–∞ –æ—Ç CryptoBot API. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –ø–æ–ø—Ä–æ–±—É–π—Ç–µ –ø–æ–∑–∂–µ."
            )
        except httpx.ConnectTimeout as e:
            import logging
            logging.error(f"CryptoBot API connection timeout: {e}")
            raise HTTPException(
                status_code=504,
                detail="–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ CryptoBot API. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –∏–Ω—Ç–µ—Ä–Ω–µ—Ç-—Å–æ–µ–¥–∏–Ω–µ–Ω–∏–µ."
            )
        except Exception as e:
            import logging
            logging.error(f"Error creating CryptoBot invoice: {e}", exc_info=True)
            # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –±–µ–∑ –∏–Ω–≤–æ–π—Å–∞, –ø–ª–∞—Ç–µ–∂ –≤—Å–µ —Ä–∞–≤–Ω–æ —Å–æ–∑–¥–∞–Ω
    
    return result


@app.post("/payments/webhook")
async def payment_webhook(
    payload: PaymentWebhookIn,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Webhook –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π –æ—Ç –ø–ª–∞—Ç–µ–∂–Ω—ã—Ö —Å–∏—Å—Ç–µ–º"""
    import json
    
    # –ù–∞—Ö–æ–¥–∏–º –ø–ª–∞—Ç–µ–∂ –ø–æ external_id –∏–ª–∏ payment_id
    if payload.payment_id:
        payment = await session.scalar(select(Payment).where(Payment.id == payload.payment_id))
    else:
        payment = await session.scalar(select(Payment).where(Payment.external_id == payload.external_id))
    
    if not payment:
        raise HTTPException(status_code=404, detail="payment_not_found")
    
    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–ª–∞—Ç–µ–∂–∞
    old_status = payment.status
    if payload.status == "succeeded":
        payment.status = PaymentStatus.succeeded
    elif payload.status == "failed":
        payment.status = PaymentStatus.failed
    else:
        payment.status = PaymentStatus.pending
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º external_id –µ—Å–ª–∏ –µ–≥–æ –µ—â–µ –Ω–µ—Ç
    if not payment.external_id and payload.external_id:
        payment.external_id = payload.external_id
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º raw_response
    if payload.raw_data:
        payment.raw_response = json.dumps(payload.raw_data)
    
    await session.commit()
    
    # –õ–æ–≥–∏—Ä—É–µ–º –∏–∑–º–µ–Ω–µ–Ω–∏–µ —Å—Ç–∞—Ç—É—Å–∞ –ø–ª–∞—Ç–µ–∂–∞
    if old_status != payment.status:
        user = await session.scalar(select(User).where(User.id == payment.user_id))
        amount_rub = payment.amount_cents / 100
        session.add(
            AuditLog(
                action=AuditLogAction.payment_status_changed,
                user_tg_id=user.tg_id if user else None,
                admin_tg_id=None,
                details=f"–°—Ç–∞—Ç—É—Å –ø–ª–∞—Ç–µ–∂–∞ #{payment.id} –∏–∑–º–µ–Ω–µ–Ω: {old_status.value} -> {payment.status.value}. –ü—Ä–æ–≤–∞–π–¥–µ—Ä: {payment.provider}, —Å—É–º–º–∞: {amount_rub:.2f} RUB ({payment.currency})",
            )
        )
        await session.commit()
    
    # –õ–æ–≥–∏—Ä—É–µ–º –ø–æ–ª—É—á–µ–Ω–∏–µ webhook
    user = await session.scalar(select(User).where(User.id == payment.user_id))
    session.add(
        AuditLog(
            action=AuditLogAction.payment_webhook_received,
            user_tg_id=user.tg_id if user else None,
            admin_tg_id=None,
            details=f"–ü–æ–ª—É—á–µ–Ω webhook –¥–ª—è –ø–ª–∞—Ç–µ–∂–∞ #{payment.id}. –ù–æ–≤—ã–π —Å—Ç–∞—Ç—É—Å: {payment.status.value}, –ø—Ä–æ–≤–∞–π–¥–µ—Ä: {payment.provider}",
        )
    )
    await session.commit()
    
    # –ï—Å–ª–∏ –ø–ª–∞—Ç–µ–∂ —É—Å–ø–µ—à–µ–Ω, –Ω–∞—á–∏—Å–ª—è–µ–º –±–∞–ª–∞–Ω—Å
    if payment.status == PaymentStatus.succeeded and old_status != PaymentStatus.succeeded:
        user = await session.scalar(select(User).where(User.id == payment.user_id))
        if user:
            old_balance = user.balance
            user.balance += payment.amount_cents
            
            # –°–æ–∑–¥–∞–µ–º —Ç—Ä–∞–Ω–∑–∞–∫—Ü–∏—é –±–∞–ª–∞–Ω—Å–∞
            session.add(
                BalanceTransaction(
                    user_id=user.id,
                    admin_tg_id=None,  # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–æ–µ –ø–æ–ø–æ–ª–Ω–µ–Ω–∏–µ
                    amount=payment.amount_cents,
                    reason=f"–ü–æ–ø–æ–ª–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ —á–µ—Ä–µ–∑ {payment.provider}",
                )
            )
            
            # –õ–æ–≥–∏—Ä—É–µ–º
            session.add(
                AuditLog(
                    action=AuditLogAction.payment_processed,
                    user_tg_id=user.tg_id,
                    admin_tg_id=None,
                    details=f"–ü–ª–∞—Ç–µ–∂ #{payment.id} –æ–±—Ä–∞–±–æ—Ç–∞–Ω. –ë–∞–ª–∞–Ω—Å: {old_balance} -> {user.balance} —Ü–µ–Ω—Ç–æ–≤",
                )
            )
            
            await session.commit()
    
    return {"success": True, "payment_id": payment.id, "status": payment.status.value}


@app.get("/payments/cryptobot/info")
async def cryptobot_info() -> dict:
    """–ü–æ–ª—É—á–µ–Ω–∏–µ –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏ –æ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏–∏ CryptoBot"""
    from core.config import get_settings
    settings = get_settings()
    
    if not settings.cryptobot_token:
        raise HTTPException(status_code=400, detail="CRYPTOBOT_TOKEN –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω")
    
    try:
        from core.cryptobot import CryptoBotAPI
        cryptobot = CryptoBotAPI(settings.cryptobot_token)
        
        result = await cryptobot.get_me()
        
        if result.get("ok"):
            return {
                "success": True,
                "app_info": result.get("result"),
            }
        else:
            return {
                "success": False,
                "error": result.get("error", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞"),
            }
    except Exception as e:
        import logging
        logging.error(f"Error getting CryptoBot info: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–∏: {str(e)}")


@app.post("/payments/cryptobot/setup")
async def cryptobot_setup_webhook(
    webhook_url: str = Query(..., description="–ü–æ–ª–Ω—ã–π URL –¥–ª—è webhook (–Ω–∞–ø—Ä–∏–º–µ—Ä: https://yourdomain.com/payments/webhook/cryptobot)"),
) -> dict:
    """–ù–∞—Å—Ç—Ä–æ–π–∫–∞ webhook –¥–ª—è CryptoBot"""
    from core.config import get_settings
    settings = get_settings()
    
    if not settings.cryptobot_token:
        raise HTTPException(status_code=400, detail="CRYPTOBOT_TOKEN –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω")
    
    try:
        from core.cryptobot import CryptoBotAPI
        cryptobot = CryptoBotAPI(settings.cryptobot_token)
        
        result = await cryptobot.set_webhook(webhook_url)
        
        if result.get("ok"):
            return {
                "success": True,
                "message": "Webhook —É—Å–ø–µ—à–Ω–æ –Ω–∞—Å—Ç—Ä–æ–µ–Ω",
                "webhook_url": webhook_url,
                "result": result.get("result"),
            }
        else:
            # –ï—Å–ª–∏ –º–µ—Ç–æ–¥ –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç—Å—è, —Å–æ–æ–±—â–∞–µ–º —á—Ç–æ –Ω—É–∂–Ω–æ –Ω–∞—Å—Ç—Ä–∞–∏–≤–∞—Ç—å —á–µ—Ä–µ–∑ –±–æ—Ç–∞
            error_code = result.get("error_code")
            if error_code == 405:
                return {
                    "success": False,
                    "error": "CryptoBot API –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç –Ω–∞—Å—Ç—Ä–æ–π–∫—É webhook —á–µ—Ä–µ–∑ API. –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –Ω–∞—Å—Ç—Ä–æ–π—Ç–µ webhook —á–µ—Ä–µ–∑ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –±–æ—Ç–∞ @CryptoBot.",
                    "webhook_url": webhook_url,
                    "instruction": f"–û—Ç–∫—Ä–æ–π—Ç–µ @CryptoBot –≤ Telegram –∏ –≤–≤–µ–¥–∏—Ç–µ URL: {webhook_url}",
                }
            return {
                "success": False,
                "error": result.get("error", "–ù–µ–∏–∑–≤–µ—Å—Ç–Ω–∞—è –æ—à–∏–±–∫–∞"),
            }
    except Exception as e:
        import logging
        logging.error(f"Error setting CryptoBot webhook: {e}", exc_info=True)
        # –ù–µ –≤—ã–±—Ä–∞—Å—ã–≤–∞–µ–º –∏—Å–∫–ª—é—á–µ–Ω–∏–µ, –∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é
        return {
            "success": False,
            "error": f"–û—à–∏–±–∫–∞ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ webhook: {str(e)}",
            "webhook_url": webhook_url,
            "instruction": "–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –Ω–∞—Å—Ç—Ä–æ–π—Ç–µ webhook —á–µ—Ä–µ–∑ –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –±–æ—Ç–∞ @CryptoBot",
        }


@app.post("/payments/webhook/cryptobot")
async def cryptobot_webhook(
    request: Request,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """Webhook –æ—Ç CryptoBot –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –ø–ª–∞—Ç–µ–∂–µ–π"""
    import json
    import logging
    import sys
    
    # –õ–æ–≥–∏—Ä—É–µ–º —Å—Ä–∞–∑—É –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ –∑–∞–ø—Ä–æ—Å–∞ (–ø—Ä–∏–Ω—É–¥–∏—Ç–µ–ª—å–Ω–æ –≤ stdout)
    print("=" * 80, file=sys.stderr, flush=True)
    print("=== CRYPTOBOT WEBHOOK RECEIVED ===", file=sys.stderr, flush=True)
    logging.info("=" * 80)
    logging.info("=== CRYPTOBOT WEBHOOK RECEIVED ===")
    
    try:
        # –õ–æ–≥–∏—Ä—É–µ–º –≤–µ—Å—å –∑–∞–ø—Ä–æ—Å –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
        body = await request.body()
        body_str = body.decode('utf-8')
        print(f"Raw body length: {len(body_str)}", file=sys.stderr, flush=True)
        print(f"Raw body (first 500): {body_str[:500]}", file=sys.stderr, flush=True)
        logging.info(f"Raw body length: {len(body_str)}")
        logging.info(f"Raw body: {body_str[:500]}")
        
        data = await request.json()
        print(f"CryptoBot webhook data: {json.dumps(data, indent=2)}", file=sys.stderr, flush=True)
        logging.info(f"CryptoBot webhook data: {json.dumps(data, indent=2)}")
        
        # CryptoBot –æ—Ç–ø—Ä–∞–≤–ª—è–µ—Ç update —Å –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏–µ–π –æ –ø–ª–∞—Ç–µ–∂–µ
        update_type = data.get("update_type")
        print(f"Update type: {update_type}", file=sys.stderr, flush=True)
        print(f"Data keys: {list(data.keys())}", file=sys.stderr, flush=True)
        logging.info(f"Update type: {update_type}")
        logging.info(f"Data keys: {list(data.keys())}")
        
        if update_type == "invoice_paid":
            # –°–æ–≥–ª–∞—Å–Ω–æ –¥–æ–∫—É–º–µ–Ω—Ç–∞—Ü–∏–∏ CryptoBot API, –¥–∞–Ω–Ω—ã–µ –∏–Ω–≤–æ–π—Å–∞ –Ω–∞—Ö–æ–¥—è—Ç—Å—è –≤ –ø–æ–ª–µ "payload"
            # –°—Ç—Ä—É–∫—Ç—É—Ä–∞: { "update_type": "invoice_paid", "payload": Invoice }
            invoice_data = data.get("payload", {})
            
            # Fallback: –µ—Å–ª–∏ payload –ø—É—Å—Ç–æ–π, –ø—Ä–æ–±—É–µ–º invoice (–¥–ª—è –æ–±—Ä–∞—Ç–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏)
            if not invoice_data or (isinstance(invoice_data, dict) and len(invoice_data) == 0):
                invoice_data = data.get("invoice", {})
                if invoice_data:
                    print("Using data['invoice'] as fallback", file=sys.stderr, flush=True)
            
            # –ï—Å–ª–∏ invoice –ø—É—Å—Ç–æ–π, –≤–æ–∑–º–æ–∂–Ω–æ –¥–∞–Ω–Ω—ã–µ –Ω–∞—Ö–æ–¥—è—Ç—Å—è –Ω–∞ –≤–µ—Ä—Ö–Ω–µ–º —É—Ä–æ–≤–Ω–µ –∏–ª–∏ –≤ result
            if not invoice_data or (isinstance(invoice_data, dict) and len(invoice_data) == 0):
                print("WARNING: invoice_data is empty, trying alternative locations", file=sys.stderr, flush=True)
                # –ü—Ä–æ–±—É–µ–º result
                if "result" in data and isinstance(data["result"], dict) and len(data["result"]) > 0:
                    invoice_data = data["result"]
                    print(f"Using data['result'] as invoice_data (keys: {list(invoice_data.keys())})", file=sys.stderr, flush=True)
                # –ï—Å–ª–∏ –≤—Å–µ –µ—â–µ –ø—É—Å—Ç–æ, –ø—Ä–æ–±—É–µ–º –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å –≤–µ—Å—å data –∫–∞–∫ invoice_data
                elif "payload" in data or "invoice_id" in data or "status" in data:
                    invoice_data = data
                    print(f"Using top-level data as invoice_data (keys: {list(invoice_data.keys())})", file=sys.stderr, flush=True)
            # –õ–æ–≥–∏—Ä—É–µ–º –ø–æ–ª–Ω—É—é —Å—Ç—Ä—É–∫—Ç—É—Ä—É –¥–∞–Ω–Ω—ã—Ö –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
            import sys
            print("=" * 80, file=sys.stderr, flush=True)
            print("=== CRYPTOBOT WEBHOOK: INVOICE PAID ===", file=sys.stderr, flush=True)
            print(f"Full data structure: {json.dumps(data, indent=2)}", file=sys.stderr, flush=True)
            print(f"Invoice data type: {type(invoice_data).__name__}", file=sys.stderr, flush=True)
            print(f"Invoice data keys: {list(invoice_data.keys()) if isinstance(invoice_data, dict) else 'NOT A DICT'}", file=sys.stderr, flush=True)
            logging.info("=" * 80)
            logging.info("=== CRYPTOBOT WEBHOOK: INVOICE PAID ===")
            logging.info(f"Full data structure: {json.dumps(data, indent=2)}")
            
            # –ü—Ä–æ–±—É–µ–º —Ä–∞–∑–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã –ø–æ–ª—É—á–µ–Ω–∏—è invoice_id –∏ payload
            # invoice_id –º–æ–∂–µ—Ç –±—ã—Ç—å —Å—Ç—Ä–æ–∫–æ–π (–Ω–∞–ø—Ä–∏–º–µ—Ä, "IVZhvKKyl5Ce") –∏–ª–∏ —á–∏—Å–ª–æ–º
            invoice_id = None
            payload_str = ""
            
            if isinstance(invoice_data, dict):
                invoice_id = invoice_data.get("invoice_id") or invoice_data.get("id")
                payload_str = invoice_data.get("payload", "") or invoice_data.get("payload_str", "")
            else:
                # –ï—Å–ª–∏ invoice_data –Ω–µ —Å–ª–æ–≤–∞—Ä—å, –≤–æ–∑–º–æ–∂–Ω–æ –¥–∞–Ω–Ω—ã–µ –≤ –¥—Ä—É–≥–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
                print(f"WARNING: invoice_data is not a dict: {type(invoice_data)}", file=sys.stderr, flush=True)
                # –ü—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –Ω–∞–ø—Ä—è–º—É—é –∏–∑ data
                invoice_id = data.get("invoice_id") or data.get("id")
                payload_str = data.get("payload", "") or data.get("payload_str", "")
            
            # –ï—Å–ª–∏ invoice_id –Ω–µ –Ω–∞–π–¥–µ–Ω, –ø—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –∏–∑ –¥—Ä—É–≥–∏—Ö –ø–æ–ª–µ–π
            if not invoice_id:
                # –ò–Ω–æ–≥–¥–∞ invoice_id –º–æ–∂–µ—Ç –±—ã—Ç—å –≤ –¥—Ä—É–≥–æ–º —Ñ–æ—Ä–º–∞—Ç–µ
                invoice_id = data.get("invoice_id") or data.get("id")
            
            print(f"Invoice ID: {invoice_id} (type: {type(invoice_id).__name__})", file=sys.stderr, flush=True)
            print(f"Payload: '{payload_str}' (type: {type(payload_str).__name__}, length: {len(payload_str) if payload_str else 0})", file=sys.stderr, flush=True)
            if isinstance(invoice_data, dict):
                print(f"Full invoice data: {json.dumps(invoice_data, indent=2)}", file=sys.stderr, flush=True)
            logging.info(f"Invoice ID: {invoice_id} (type: {type(invoice_id).__name__})")
            logging.info(f"Payload: '{payload_str}' (type: {type(payload_str).__name__})")
            if isinstance(invoice_data, dict):
                logging.info(f"Full invoice data: {json.dumps(invoice_data, indent=2)}")
            print("=" * 80, file=sys.stderr, flush=True)
            
            # –ò–∑–≤–ª–µ–∫–∞–µ–º payment_id –∏–∑ payload
            payment_id = None
            if payload_str:
                print(f"Processing payload: '{payload_str}'", file=sys.stderr, flush=True)
                if payload_str.startswith("payment_"):
                    try:
                        payment_id = int(payload_str.split("_")[1])
                        print(f"‚úÖ‚úÖ‚úÖ Extracted payment_id from payload: {payment_id}", file=sys.stderr, flush=True)
                        logging.info(f"Extracted payment_id from payload: {payment_id}")
                    except (ValueError, IndexError) as e:
                        print(f"ERROR: Failed to extract payment_id from payload '{payload_str}': {e}", file=sys.stderr, flush=True)
                        logging.error(f"Failed to extract payment_id from payload '{payload_str}': {e}")
                else:
                    print(f"WARNING: Payload does not start with 'payment_': '{payload_str}'", file=sys.stderr, flush=True)
            else:
                print(f"WARNING: payload_str is empty or None", file=sys.stderr, flush=True)
            
            # –ï—Å–ª–∏ payload –ø—É—Å—Ç–æ–π, –ø—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ –ø–ª–∞—Ç–µ–∂ –ø–æ external_id (invoice_id)
            if not payment_id and invoice_id:
                logging.info(f"Payload empty, searching payment by external_id (invoice_id): {invoice_id}")
                
                # –ü—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ –ø–æ —Ç–æ—á–Ω–æ–º—É —Å–æ–≤–ø–∞–¥–µ–Ω–∏—é
                payment_by_external = await session.scalar(
                    select(Payment).where(Payment.external_id == str(invoice_id))
                )
                
                # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏, –ø—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ –ø–æ —á–∏—Å–ª–æ–≤–æ–º—É –∑–Ω–∞—á–µ–Ω–∏—é
                if not payment_by_external:
                    try:
                        invoice_id_int = int(invoice_id)
                        payment_by_external = await session.scalar(
                            select(Payment).where(Payment.external_id.cast(Integer) == invoice_id_int)
                        )
                    except (ValueError, TypeError):
                        pass
                
                # –ï—Å–ª–∏ –≤—Å–µ –µ—â–µ –Ω–µ –Ω–∞—à–ª–∏, –ø—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ –≤—Å–µ –ø–ª–∞—Ç–µ–∂–∏ CryptoBot –∏ —Å—Ä–∞–≤–Ω–∏—Ç—å
                if not payment_by_external:
                    logging.info(f"üîç Searching in all recent cryptobot payments...")
                    all_recent = await session.scalars(
                        select(Payment)
                        .where(Payment.provider == "cryptobot")
                        .order_by(Payment.created_at.desc())
                        .limit(20)
                    )
                    for p in all_recent:
                        logging.info(f"  Checking Payment #{p.id}: external_id='{p.external_id}' (type: {type(p.external_id)}), invoice_id='{invoice_id}' (type: {type(invoice_id)})")
                        if p.external_id and str(p.external_id) == str(invoice_id):
                            payment_by_external = p
                            logging.info(f"  ‚úÖ Found match! Payment #{p.id}")
                            break
                
                if payment_by_external:
                    payment_id = payment_by_external.id
                    logging.info(f"‚úÖ Found payment by external_id: payment_id={payment_id}, external_id={payment_by_external.external_id}")
                else:
                    logging.warning(f"‚ùå Payment not found by external_id={invoice_id}")
                    # –õ–æ–≥–∏—Ä—É–µ–º –≤—Å–µ –ø–æ—Å–ª–µ–¥–Ω–∏–µ –ø–ª–∞—Ç–µ–∂–∏ –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
                    recent_payments = await session.scalars(
                        select(Payment)
                        .where(Payment.provider == "cryptobot")
                        .order_by(Payment.created_at.desc())
                        .limit(10)
                    )
                    logging.info(f"üìã Recent cryptobot payments (last 10):")
                    for p in recent_payments:
                        logging.info(f"  Payment #{p.id}: external_id='{p.external_id}' (type: {type(p.external_id)}), status={p.status}, created_at={p.created_at}, user_id={p.user_id}")
            
            if payment_id:
                payment = await session.scalar(select(Payment).where(Payment.id == payment_id))
                if payment:
                    logging.info(f"Found payment #{payment_id}, current status: {payment.status}")
                    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–ª–∞—Ç–µ–∂–∞
                    old_status = payment.status
                    payment.status = PaymentStatus.succeeded
                    payment.external_id = str(invoice_id)
                    payment.raw_response = json.dumps(data)
                    await session.commit()
                    logging.info(f"Payment #{payment_id} status updated to succeeded")
                    
                    # –ù–∞—á–∏—Å–ª—è–µ–º –±–∞–ª–∞–Ω—Å –µ—Å–ª–∏ –µ—â–µ –Ω–µ –Ω–∞—á–∏—Å–ª–µ–Ω
                    if old_status != PaymentStatus.succeeded:
                        logging.info(f"Processing balance credit for payment #{payment_id}")
                        user = await session.scalar(select(User).where(User.id == payment.user_id))
                        if user:
                            old_balance = user.balance
                            user.balance += payment.amount_cents
                            
                            logging.info(f"User {user.tg_id}: balance {old_balance} -> {user.balance} cents")
                            
                            session.add(
                                BalanceTransaction(
                                    user_id=user.id,
                                    admin_tg_id=None,
                                    amount=payment.amount_cents,
                                    reason=f"–ü–æ–ø–æ–ª–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ —á–µ—Ä–µ–∑ {payment.provider}",
                                )
                            )
                            
                            session.add(
                                AuditLog(
                                    action=AuditLogAction.payment_processed,
                                    user_tg_id=user.tg_id,
                                    admin_tg_id=None,
                                    details=f"–ü–ª–∞—Ç–µ–∂ #{payment.id} –æ–±—Ä–∞–±–æ—Ç–∞–Ω —á–µ—Ä–µ–∑ CryptoBot. –ë–∞–ª–∞–Ω—Å: {old_balance} -> {user.balance} —Ü–µ–Ω—Ç–æ–≤",
                                )
                            )
                            
                            await session.commit()
                            logging.info(f"Balance credited successfully for user {user.tg_id}")
                            
                            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é —á–µ—Ä–µ–∑ –±–æ—Ç–∞ (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ)
                            notify_on_payment = await session.scalar(select(SystemSetting).where(SystemSetting.key == "notify_on_payment"))
                            if not notify_on_payment or notify_on_payment.value != "false":
                                try:
                                    # amount_cents –∏ balance —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
                                    amount_rub = payment.amount_cents / 100
                                    new_balance_rub = user.balance / 100
                                    
                                    notification_text = (
                                        f"‚úÖ <b>–ü–ª–∞—Ç–µ–∂ —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω!</b>\n\n"
                                        f"üí∞ –ü–æ–ø–æ–ª–Ω–µ–Ω–æ: <b>{amount_rub:.2f} RUB</b>\n"
                                        f"üíµ –¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{new_balance_rub:.2f} RUB</b>"
                                    )
                                    
                                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º BOT_TOKEN –∏–∑ –æ–∫—Ä—É–∂–µ–Ω–∏—è
                                    bot_token = os.getenv("BOT_TOKEN")
                                    if bot_token:
                                        import asyncio
                                        asyncio.create_task(_send_user_notification(
                                            user.tg_id,
                                            notification_text,
                                            bot_token
                                        ))
                                        logging.info(f"Notification task created for user {user.tg_id}")
                                    else:
                                        logging.warning(f"BOT_TOKEN not found, cannot send notification to user {user.tg_id}")
                                except Exception as e:
                                    logging.error(f"Failed to send notification to user {user.tg_id}: {e}", exc_info=True)
                        else:
                            logging.error(f"User not found for payment #{payment_id}, user_id={payment.user_id}")
                else:
                    logging.error(f"Payment #{payment_id} not found")
            else:
                import sys
                print(f"WARNING: Could not extract payment_id from payload: '{payload_str}'", file=sys.stderr, flush=True)
                print(f"INFO: invoice_id from webhook: {invoice_id}", file=sys.stderr, flush=True)
                logging.warning(f"Could not extract payment_id from payload: {payload_str}")
                logging.info(f"invoice_id from webhook: {invoice_id}")
                # –ü—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ –ø–ª–∞—Ç–µ–∂ –ø–æ invoice_id –Ω–∞–ø—Ä—è–º—É—é, –µ—Å–ª–∏ payload –ø—É—Å—Ç–æ–π
                if invoice_id:
                    print(f"üîç Trying to find payment by invoice_id: {invoice_id}", file=sys.stderr, flush=True)
                    logging.info(f"üîç Trying to find payment by invoice_id directly: {invoice_id} (type: {type(invoice_id)})")
                    # –ü—Ä–æ–±—É–µ–º —Ä–∞–∑–Ω—ã–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã –ø–æ–∏—Å–∫–∞
                    payment_by_invoice = await session.scalar(
                        select(Payment).where(
                            Payment.external_id == str(invoice_id)
                        ).order_by(Payment.created_at.desc())
                    )
                    
                    # –ï—Å–ª–∏ –Ω–µ –Ω–∞—à–ª–∏, –ø—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ —Å—Ä–µ–¥–∏ –≤—Å–µ—Ö –ø–ª–∞—Ç–µ–∂–µ–π CryptoBot
                    if not payment_by_invoice:
                        logging.info(f"üîç Searching in all cryptobot payments...")
                        all_payments = await session.scalars(
                            select(Payment)
                            .where(Payment.provider == "cryptobot")
                            .order_by(Payment.created_at.desc())
                            .limit(50)
                        )
                        for p in all_payments:
                            p_external_str = str(p.external_id) if p.external_id else "None"
                            invoice_str = str(invoice_id) if invoice_id else "None"
                            if p.external_id and str(p.external_id) == str(invoice_id):
                                payment_by_invoice = p
                                logging.info(f"‚úÖ‚úÖ‚úÖ FOUND PAYMENT #{p.id} by invoice_id in all payments: external_id='{p_external_str}' == invoice_id='{invoice_str}'")
                                break
                            else:
                                logging.info(f"  Payment #{p.id}: external_id='{p_external_str}' != invoice_id='{invoice_str}'")
                    if payment_by_invoice:
                        logging.info(f"Found payment #{payment_by_invoice.id} by invoice_id, updating status")
                        old_status = payment_by_invoice.status
                        payment_by_invoice.status = PaymentStatus.succeeded
                        payment_by_invoice.raw_response = json.dumps(data)
                        await session.commit()
                        logging.info(f"Payment #{payment_by_invoice.id} status updated from {old_status} to succeeded")
                        
                        # –ù–∞—á–∏—Å–ª—è–µ–º –±–∞–ª–∞–Ω—Å –µ—Å–ª–∏ –µ—â–µ –Ω–µ –Ω–∞—á–∏—Å–ª–µ–Ω
                        if old_status != PaymentStatus.succeeded:
                            user = await session.scalar(select(User).where(User.id == payment_by_invoice.user_id))
                            if user:
                                old_balance = user.balance
                                user.balance += payment_by_invoice.amount_cents
                                session.add(
                                    BalanceTransaction(
                                        user_id=user.id,
                                        admin_tg_id=None,
                                        amount=payment_by_invoice.amount_cents,
                                        reason=f"–ü–æ–ø–æ–ª–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ —á–µ—Ä–µ–∑ {payment_by_invoice.provider}",
                                    )
                                )
                                session.add(
                                    AuditLog(
                                        action=AuditLogAction.payment_processed,
                                        user_tg_id=user.tg_id,
                                        admin_tg_id=None,
                                        details=f"–ü–ª–∞—Ç–µ–∂ #{payment_by_invoice.id} –æ–±—Ä–∞–±–æ—Ç–∞–Ω —á–µ—Ä–µ–∑ CryptoBot. –ë–∞–ª–∞–Ω—Å: {old_balance} -> {user.balance} —Ü–µ–Ω—Ç–æ–≤",
                                    )
                                )
                                await session.commit()
                                
                                # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ
                                try:
                                    amount_rub = payment_by_invoice.amount_cents / 100
                                    new_balance_rub = user.balance / 100
                                    notification_text = (
                                        f"‚úÖ <b>–ü–ª–∞—Ç–µ–∂ —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω!</b>\n\n"
                                        f"üí∞ –ü–æ–ø–æ–ª–Ω–µ–Ω–æ: <b>{amount_rub:.2f} RUB</b>\n"
                                        f"üíµ –¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{new_balance_rub:.2f} RUB</b>"
                                    )
                                    bot_token = os.getenv("BOT_TOKEN")
                                    if bot_token:
                                        import asyncio
                                        asyncio.create_task(_send_user_notification(
                                            user.tg_id,
                                            notification_text,
                                            bot_token
                                        ))
                                except Exception as e:
                                    logging.error(f"Failed to send notification: {e}")
                    else:
                        logging.warning(f"Payment not found by invoice_id: {invoice_id}")
        else:
            logging.info(f"Update type '{update_type}' is not invoice_paid, ignoring")
        
        return {"ok": True}
    except Exception as e:
        import logging
        logging.error(f"Error processing CryptoBot webhook: {e}", exc_info=True)
        return {"ok": False, "error": str(e)}


@app.post("/payments/cryptobot/check/{payment_id}")
async def cryptobot_check_payment(
    payment_id: int,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–†—É—á–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ —Å—Ç–∞—Ç—É—Å–∞ –ø–ª–∞—Ç–µ–∂–∞ —á–µ—Ä–µ–∑ CryptoBot API"""
    from core.config import get_settings
    settings = get_settings()
    
    if not settings.cryptobot_token:
        raise HTTPException(status_code=400, detail="CRYPTOBOT_TOKEN –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω")
    
    try:
        payment = await session.scalar(select(Payment).where(Payment.id == payment_id))
        if not payment:
            raise HTTPException(status_code=404, detail="Payment not found")
        
        if not payment.external_id:
            return {
                "success": False,
                "error": "Payment has no external_id (invoice_id)",
                "payment_id": payment_id,
            }
        
        from core.cryptobot import CryptoBotAPI
        cryptobot = CryptoBotAPI(settings.cryptobot_token)
        
        invoice_id = int(payment.external_id)
        invoice_status = await cryptobot.get_invoice_status(invoice_id)
        
        if invoice_status.get("ok") and invoice_status.get("result"):
            invoices = invoice_status["result"]
            if invoices and len(invoices) > 0:
                invoice = invoices[0]
                status = invoice.get("status")
                
                # –ï—Å–ª–∏ –∏–Ω–≤–æ–π—Å –æ–ø–ª–∞—á–µ–Ω, –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –ø–ª–∞—Ç–µ–∂
                if status == "paid" and payment.status != PaymentStatus.succeeded:
                    import json
                    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–ª–∞—Ç–µ–∂–∞
                    payment.status = PaymentStatus.succeeded
                    payment.raw_response = json.dumps(invoice)
                    await session.commit()
                    
                    # –ù–∞—á–∏—Å–ª—è–µ–º –±–∞–ª–∞–Ω—Å
                    user = await session.scalar(select(User).where(User.id == payment.user_id))
                    if user:
                        old_balance = user.balance
                        user.balance += payment.amount_cents
                        
                        session.add(
                            BalanceTransaction(
                                user_id=user.id,
                                admin_tg_id=None,
                                amount=payment.amount_cents,
                                reason=f"–ü–æ–ø–æ–ª–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ —á–µ—Ä–µ–∑ {payment.provider}",
                            )
                        )
                        
                        session.add(
                            AuditLog(
                                action=AuditLogAction.payment_processed,
                                user_tg_id=user.tg_id,
                                admin_tg_id=None,
                                details=f"–ü–ª–∞—Ç–µ–∂ #{payment.id} –æ–±—Ä–∞–±–æ—Ç–∞–Ω —á–µ—Ä–µ–∑ CryptoBot (—Ä—É—á–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞). –ë–∞–ª–∞–Ω—Å: {old_balance} -> {user.balance} —Ü–µ–Ω—Ç–æ–≤",
                            )
                        )
                        
                        await session.commit()
                        
                        return {
                            "success": True,
                            "message": "Payment processed successfully",
                            "payment_id": payment_id,
                            "status": "succeeded",
                            "balance_credited": True,
                        }
                
                return {
                    "success": True,
                    "payment_id": payment_id,
                    "invoice_status": status,
                    "payment_status": payment.status.value,
                }
        
        return {
            "success": False,
            "error": "Failed to get invoice status",
            "response": invoice_status,
        }
    except Exception as e:
        import logging
        logging.error(f"Error checking CryptoBot payment: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏ –ø–ª–∞—Ç–µ–∂–∞: {str(e)}")


@app.get("/subscriptions/status/by_tg/{tg_id}")
async def subscription_status_by_tg(tg_id: int, session: AsyncSession = Depends(get_session)) -> SubscriptionStatusOut:
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return SubscriptionStatusOut(has_active=False)

    sub = await session.scalar(
        select(Subscription)
        .where(Subscription.user_id == user.id)
        .where(Subscription.status == SubscriptionStatus.active)
        .order_by(Subscription.ends_at.desc().nullslast())
    )

    if not sub:
        return SubscriptionStatusOut(has_active=False)

    return SubscriptionStatusOut(has_active=True, plan_name=sub.plan_name, ends_at=sub.ends_at)


async def _get_subscription_plans_from_db(session: AsyncSession) -> dict[int, dict]:
    """–ü–æ–ª—É—á–∏—Ç—å —Ç–∞—Ä–∏—Ñ—ã –∏–∑ –ë–î"""
    plans = await session.scalars(
        select(SubscriptionPlan)
        .where(SubscriptionPlan.is_active == True)
        .order_by(SubscriptionPlan.display_order, SubscriptionPlan.days)
    )
    return {
        plan.days: {
            "name": plan.name,
            "price_cents": plan.price_cents,
            "description": plan.description or "",
        }
        for plan in plans.all()
    }


async def _ensure_default_plans(session: AsyncSession) -> None:
    """–°–æ–∑–¥–∞—Ç—å —Ç–∞—Ä–∏—Ñ—ã –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é, –µ—Å–ª–∏ –∏—Ö –Ω–µ—Ç"""
    existing = await session.scalar(select(func.count()).select_from(SubscriptionPlan))
    if existing and existing > 0:
        return
    
    default_plans = [
        SubscriptionPlan(days=1, name="1 –¥–µ–Ω—å", price_cents=500, description="–ü—Ä–æ–±–Ω—ã–π –¥–µ–Ω—å", display_order=1),
        SubscriptionPlan(days=7, name="7 –¥–Ω–µ–π", price_cents=3000, description="–ù–µ–¥–µ–ª—å–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞", display_order=2),
        SubscriptionPlan(days=30, name="1 –º–µ—Å—è—Ü", price_cents=10000, description="–ú–µ—Å—è—á–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞", display_order=3),
        SubscriptionPlan(days=90, name="3 –º–µ—Å—è—Ü–∞", price_cents=27000, description="–¢—Ä–µ—Ö–º–µ—Å—è—á–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π 10%", display_order=4),
        SubscriptionPlan(days=180, name="6 –º–µ—Å—è—Ü–µ–≤", price_cents=48000, description="–ü–æ–ª—É–≥–æ–¥–æ–≤–∞—è –ø–æ–¥–ø–∏—Å–∫–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π 20%", display_order=5),
        SubscriptionPlan(days=365, name="12 –º–µ—Å—è—Ü–µ–≤", price_cents=84000, description="–ì–æ–¥–æ–≤–∞—è –ø–æ–¥–ø–∏—Å–∫–∞ —Å–æ —Å–∫–∏–¥–∫–æ–π 30%", display_order=6),
    ]
    for plan in default_plans:
        session.add(plan)
    await session.commit()


@app.get("/subscriptions/plans")
async def get_subscription_plans(session: AsyncSession = Depends(get_session)) -> dict:
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –¥–æ—Å—Ç—É–ø–Ω—ã—Ö —Ç–∞—Ä–∏—Ñ–æ–≤ –ø–æ–¥–ø–∏—Å–∫–∏"""
    await _ensure_default_plans(session)
    plans = await session.scalars(
        select(SubscriptionPlan)
        .where(SubscriptionPlan.is_active == True)
        .order_by(SubscriptionPlan.display_order, SubscriptionPlan.days)
    )
    return {
        "plans": [
            {
                "id": plan.id,
                "days": plan.days,
                "name": plan.name,
                "description": plan.description or "",
                "price_cents": plan.price_cents,
                "price_rub": plan.price_cents / 100,
                "is_active": plan.is_active,
                "display_order": plan.display_order,
            }
            for plan in plans.all()
        ]
    }


@app.post("/subscriptions/purchase")
async def purchase_subscription(
    payload: SubscriptionPurchaseIn,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–ü–æ–∫—É–ø–∫–∞ –ø–æ–¥–ø–∏—Å–∫–∏ —á–µ—Ä–µ–∑ –±–∞–ª–∞–Ω—Å"""
    from datetime import datetime, timedelta, timezone
    
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    await _ensure_default_plans(session)
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–ª–∞–Ω –∏–∑ –ë–î
    plan_db = await session.scalar(
        select(SubscriptionPlan)
        .where(SubscriptionPlan.days == payload.plan_days)
        .where(SubscriptionPlan.is_active == True)
    )
    
    if not plan_db:
        raise HTTPException(status_code=400, detail="invalid_plan")
    
    plan_name = plan_db.name
    price_cents = plan_db.price_cents
    
    # –ü—Ä–æ–º–æ–∫–æ–¥—ã –Ω–∞ —Å–∫–∏–¥–∫—É (–ø—Ä–æ—Ü–µ–Ω—Ç) –±–æ–ª—å—à–µ –Ω–µ –ø—Ä–∏–º–µ–Ω—è—é—Ç—Å—è –ø—Ä–∏ –ø–æ–∫—É–ø–∫–µ –ø–æ–¥–ø–∏—Å–∫–∏
    # –ü—Ä–æ–º–æ–∫–æ–¥—ã –Ω–∞ —Ñ–∏–∫—Å —Å—É–º–º—É –ø—Ä–∏–º–µ–Ω—è—é—Ç—Å—è —á–µ—Ä–µ–∑ –æ—Ç–¥–µ–ª—å–Ω—ã–π endpoint /promo-codes/apply
    final_price_cents = price_cents
    promo_code_used = None
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –±–∞–ª–∞–Ω—Å
    if user.balance < final_price_cents:
        raise HTTPException(
            status_code=400,
            detail=f"insufficient_balance. Required: {final_price_cents / 100:.2f} RUB, Available: {user.balance / 100:.2f} RUB"
        )
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –∞–∫—Ç–∏–≤–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞
    active_sub = await session.scalar(
        select(Subscription)
        .where(Subscription.user_id == user.id)
        .where(Subscription.status == SubscriptionStatus.active)
        .order_by(Subscription.ends_at.desc().nullslast())
    )
    
    # –ï—Å–ª–∏ –µ—Å—Ç—å –∞–∫—Ç–∏–≤–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞, –ø—Ä–æ–¥–ª–µ–≤–∞–µ–º –µ—ë
    now = datetime.now(timezone.utc)
    if active_sub and active_sub.ends_at and active_sub.ends_at > now:
        # –ü—Ä–æ–¥–ª–µ–≤–∞–µ–º –æ—Ç —Ç–µ–∫—É—â–µ–π –¥–∞—Ç—ã –æ–∫–æ–Ω—á–∞–Ω–∏—è
        starts_at = active_sub.ends_at
        ends_at = starts_at + timedelta(days=payload.plan_days)
    else:
        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –ø–æ–¥–ø–∏—Å–∫—É
        starts_at = now
        ends_at = now + timedelta(days=payload.plan_days)
    
    # –°–ø–∏—Å—ã–≤–∞–µ–º –±–∞–ª–∞–Ω—Å
    user.balance -= final_price_cents
    
    # –°–æ–∑–¥–∞–µ–º —Ç—Ä–∞–Ω–∑–∞–∫—Ü–∏—é –±–∞–ª–∞–Ω—Å–∞
    session.add(
        BalanceTransaction(
            user_id=user.id,
            admin_tg_id=None,
            amount=-final_price_cents,  # –û—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ = —Å–ø–∏—Å–∞–Ω–∏–µ
            reason=f"–ü–æ–∫—É–ø–∫–∞ –ø–æ–¥–ø–∏—Å–∫–∏: {plan_name}",
        )
    )
    
    # –°–æ–∑–¥–∞–µ–º –∏–ª–∏ –æ–±–Ω–æ–≤–ª—è–µ–º –ø–æ–¥–ø–∏—Å–∫—É
    if active_sub and active_sub.status == SubscriptionStatus.active:
        # –ü—Ä–æ–¥–ª–µ–≤–∞–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é
        active_sub.ends_at = ends_at
        active_sub.price_cents = final_price_cents  # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∏–Ω–∞–ª—å–Ω—É—é —Ü–µ–Ω—É —Å —É—á–µ—Ç–æ–º —Å–∫–∏–¥–∫–∏
        subscription = active_sub
    else:
        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –ø–æ–¥–ø–∏—Å–∫—É
        subscription = Subscription(
            user_id=user.id,
            plan_name=plan_name,
            price_cents=final_price_cents,  # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∏–Ω–∞–ª—å–Ω—É—é —Ü–µ–Ω—É —Å —É—á–µ—Ç–æ–º —Å–∫–∏–¥–∫–∏
            currency="RUB",
            status=SubscriptionStatus.active,
            starts_at=starts_at,
            ends_at=ends_at,
        )
        session.add(subscription)
    
    # –õ–æ–≥–∏—Ä—É–µ–º –ø–æ–∫—É–ø–∫—É
    log_details = f"–ü–æ–∫—É–ø–∫–∞ –ø–æ–¥–ø–∏—Å–∫–∏: {plan_name}. –¶–µ–Ω–∞: {final_price_cents / 100:.2f} RUB. –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_at.strftime('%d.%m.%Y %H:%M')} (UTC)"
    session.add(
        AuditLog(
            action=AuditLogAction.subscription_created,
            user_tg_id=user.tg_id,
            admin_tg_id=None,
            details=log_details,
        )
    )
    
    await session.commit()
    
    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–¥–ø–∏—Å–∫–∏ —É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ—Å–ª–µ –∫–æ–º–º–∏—Ç–∞
    await _update_user_subscription_status(user.id, session)
    await session.commit()
    await session.refresh(user)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–µ—Å—Å–∏–∏
    
    # –ù–µ –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º VPN –∫–æ–Ω—Ñ–∏–≥–∏ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ - –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –≤—ã–±–µ—Ä–µ—Ç —Å–µ—Ä–≤–µ—Ä –∏ —Å–≥–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç –∫–ª—é—á —Å–∞–º
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ)
    notify_on_subscription = await session.scalar(select(SystemSetting).where(SystemSetting.key == "notify_on_subscription"))
    if not notify_on_subscription or notify_on_subscription.value != "false":
        try:
            ends_at_moscow = ends_at.astimezone(ZoneInfo("Europe/Moscow"))
            ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M")
            notification_text = (
                f"‚úÖ <b>–ü–æ–¥–ø–∏—Å–∫–∞ —É—Å–ø–µ—à–Ω–æ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω–∞!</b>\n\n"
                f"üì¶ –¢–∞—Ä–∏—Ñ: <b>{plan_name}</b>\n"
                f"üí∞ –°—Ç–æ–∏–º–æ—Å—Ç—å: {final_price_cents / 100:.2f} RUB\n"
                f"üìÖ –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_str} –ú–°–ö\n"
                f"üíµ –û—Å—Ç–∞—Ç–æ–∫ –±–∞–ª–∞–Ω—Å–∞: {user.balance / 100:.2f} RUB"
            )
            asyncio.create_task(_send_user_notification(user.tg_id, notification_text))
        except Exception:
            pass  # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –æ—à–∏–±–∫–∏ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π
    
    return {
        "subscription_id": subscription.id,
        "plan_name": plan_name,
        "price_cents": final_price_cents,
        "price_rub": final_price_cents / 100,
        "starts_at": starts_at.isoformat(),
        "ends_at": ends_at.isoformat(),
        "balance_remaining": user.balance / 100,
    }


@app.post("/subscriptions/trial")
async def activate_trial(
    payload: SubscriptionTrialIn,
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–ê–∫—Ç–∏–≤–∞—Ü–∏—è –±–µ—Å–ø–ª–∞—Ç–Ω–æ–≥–æ –ø—Ä–æ–±–Ω–æ–≥–æ –ø–µ—Ä–∏–æ–¥–∞ –Ω–∞ 7 –¥–Ω–µ–π (–µ–¥–∏–Ω–æ—Ä–∞–∑–æ–≤–æ)"""
    from datetime import datetime, timedelta, timezone
    
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω –ª–∏ —É–∂–µ –ø—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥
    if user.trial_used:
        raise HTTPException(status_code=400, detail="trial_already_used")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ –∞–∫—Ç–∏–≤–Ω–æ–π –ø–æ–¥–ø–∏—Å–∫–∏
    active_sub = await session.scalar(
        select(Subscription)
        .where(Subscription.user_id == user.id)
        .where(Subscription.status == SubscriptionStatus.active)
    )
    
    if active_sub:
        raise HTTPException(status_code=400, detail="active_subscription_exists")
    
    # –ê–∫—Ç–∏–≤–∏—Ä—É–µ–º –ø—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥
    now = datetime.now(timezone.utc)
    starts_at = now
    ends_at = now + timedelta(days=7)
    
    subscription = Subscription(
        user_id=user.id,
        plan_name="–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ (7 –¥–Ω–µ–π)",
        price_cents=0,
        currency="RUB",
        status=SubscriptionStatus.active,
        starts_at=starts_at,
        ends_at=ends_at,
    )
    session.add(subscription)
    
    # –û—Ç–º–µ—á–∞–µ–º, —á—Ç–æ –ø—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω
    user.trial_used = True
    
    # –õ–æ–≥–∏—Ä—É–µ–º –∞–∫—Ç–∏–≤–∞—Ü–∏—é –ø—Ä–æ–±–Ω–æ–≥–æ –ø–µ—Ä–∏–æ–¥–∞
    session.add(
        AuditLog(
            action=AuditLogAction.subscription_activated,
            user_tg_id=user.tg_id,
            admin_tg_id=None,
            details=f"–ê–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω –ø—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ –Ω–∞ 7 –¥–Ω–µ–π. –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_at.strftime('%d.%m.%Y %H:%M')} (UTC)",
        )
    )
    
    await session.commit()
    
    # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–¥–ø–∏—Å–∫–∏ —É –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ—Å–ª–µ –∫–æ–º–º–∏—Ç–∞
    await _update_user_subscription_status(user.id, session)
    await session.commit()
    await session.refresh(user)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–µ—Å—Å–∏–∏
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é (–µ—Å–ª–∏ –≤–∫–ª—é—á–µ–Ω–æ)
    notify_on_subscription = await session.scalar(select(SystemSetting).where(SystemSetting.key == "notify_on_subscription"))
    if not notify_on_subscription or notify_on_subscription.value != "false":
        try:
            ends_at_moscow = ends_at.astimezone(ZoneInfo("Europe/Moscow"))
            ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M")
            notification_text = (
                f"üéÅ <b>–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ –∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω!</b>\n\n"
                f"üì¶ –¢–∞—Ä–∏—Ñ: <b>–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ (7 –¥–Ω–µ–π)</b>\n"
                f"üìÖ –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_str} –ú–°–ö\n\n"
                f"–ü–æ—Å–ª–µ –æ–∫–æ–Ω—á–∞–Ω–∏—è –ø—Ä–æ–±–Ω–æ–≥–æ –ø–µ—Ä–∏–æ–¥–∞ –≤—ã —Å–º–æ–∂–µ—Ç–µ –ø—Ä–∏–æ–±—Ä–µ—Å—Ç–∏ –ø–æ–¥–ø–∏—Å–∫—É."
            )
            asyncio.create_task(_send_user_notification(user.tg_id, notification_text))
        except Exception:
            pass  # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –æ—à–∏–±–∫–∏ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π
    
    return {
        "subscription_id": subscription.id,
        "plan_name": "–ü—Ä–æ–±–Ω—ã–π –ø–µ—Ä–∏–æ–¥ (7 –¥–Ω–µ–π)",
        "starts_at": starts_at.isoformat(),
        "ends_at": ends_at.isoformat(),
    }


@app.get("/users/referral/by_tg/{tg_id}")
async def referral_info_by_tg(tg_id: int, session: AsyncSession = Depends(get_session)) -> ReferralInfoOut:
    stmt = select(User).options(selectinload(User.referred_by)).where(User.tg_id == tg_id)
    user = await session.scalar(stmt)
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    if not user.referral_code:
        user.referral_code = await _ensure_ref_code_unique(session)
        await session.commit()
        await session.refresh(user)
        # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º —Å relationship –ø–æ—Å–ª–µ refresh
        user = await session.scalar(stmt)

    referrals_count = await session.scalar(select(func.count()).select_from(User).where(User.referred_by_user_id == user.id))
    referred_by_tg_id = user.referred_by.tg_id if user.referred_by else None
    
    # –ü–æ–¥—Å—á–∏—Ç—ã–≤–∞–µ–º –æ–±—â—É—é —Å—É–º–º—É —Ä–µ—Ñ–µ—Ä–∞–ª—å–Ω—ã—Ö –Ω–∞–≥—Ä–∞–¥ (—Ç–æ–ª—å–∫–æ –¥–ª—è –ø—Ä–∏–≥–ª–∞—Å–∏–≤—à–µ–≥–æ)
    total_rewards_result = await session.scalar(
        select(func.sum(ReferralReward.amount_cents))
        .select_from(ReferralReward)
        .where(
            ReferralReward.referrer_user_id == user.id,
            ReferralReward.is_for_referrer == True
        )
    )
    total_rewards_cents = int(total_rewards_result or 0)
    
    return ReferralInfoOut(
        tg_id=user.tg_id,
        referral_code=user.referral_code,
        referred_by_tg_id=referred_by_tg_id,
        referrals_count=int(referrals_count or 0),
        total_rewards_cents=total_rewards_cents,
    )


# --- Admin actions (optionally protected by ADMIN_TOKEN) ---
async def _send_user_notification_with_menu_update(tg_id: int, text: str, bot_token: str | None = None) -> None:
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é –∏ –æ–±–Ω–æ–≤–ª—è–µ—Ç –µ–≥–æ –º–µ–Ω—é"""
    try:
        from bot.keyboards import user_menu
        from aiogram import Bot
        from aiogram.types import ReplyKeyboardMarkup
        
        settings = get_settings()
        if not bot_token:
            bot_token = os.getenv("BOT_TOKEN", "") or settings.bot_token
        
        if not bot_token:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –æ—Ç–ø—Ä–∞–≤–∏—Ç—å —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {tg_id}: bot_token –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω")
            return
        
        bot = Bot(token=bot_token)
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–¥–ø–∏—Å–∫–∏ –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –º–µ–Ω—é
        has_subscription = False
        try:
            async with SessionLocal() as session:
                user = await session.scalar(select(User).where(User.tg_id == tg_id))
                if user:
                    has_subscription = user.has_active_subscription
        except Exception:
            pass
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —è–≤–ª—è–µ—Ç—Å—è –ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∞–¥–º–∏–Ω–æ–º
        is_admin = tg_id in set(settings.admin_ids)
        
        # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –æ–±–Ω–æ–≤–ª–µ–Ω–Ω—ã–º –º–µ–Ω—é
        await bot.send_message(
            chat_id=tg_id,
            text=text,
            parse_mode="HTML",
            reply_markup=user_menu(is_admin=is_admin, has_subscription=has_subscription)
        )
        
        await bot.session.close()
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–ø—Ä–∞–≤–∫–µ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏—è —Å –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ–º –º–µ–Ω—é –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é {tg_id}: {e}", exc_info=True)


async def _send_user_notification(tg_id: int, text: str, bot_token: str | None = None) -> None:
    """–û—Ç–ø—Ä–∞–≤–ª—è–µ—Ç —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é –≤ –±–æ—Ç–µ"""
    if not bot_token:
        bot_token = os.getenv("BOT_TOKEN")
    if not bot_token:
        return
    
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10.0) as client:
            await client.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": tg_id, "text": text, "parse_mode": "HTML"},
            )
    except Exception:
        pass  # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –æ—à–∏–±–∫–∏ –æ—Ç–ø—Ä–∞–≤–∫–∏ —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π


@app.post("/admin/users/credit")
async def admin_credit_user(
    payload: AdminCreditIn,
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> dict[str, int]:
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    old_balance = user.balance
    # payload.amount —É–∂–µ –≤ —Ä—É–±–ª—è—Ö, –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ –∫–æ–ø–µ–π–∫–∏
    amount_cents = int(float(payload.amount) * 100)
    user.balance += amount_cents
    amount_rub = amount_cents / 100
    session.add(
        BalanceTransaction(
            user_id=user.id,
            admin_tg_id=payload.admin_tg_id,
            amount=amount_cents,
            reason=payload.reason,
        )
    )
    session.add(
        AuditLog(
            action=AuditLogAction.balance_credited,
            user_tg_id=payload.tg_id,
            admin_tg_id=payload.admin_tg_id,
            details=f"–ë–∞–ª–∞–Ω—Å –∏–∑–º–µ–Ω–µ–Ω: {old_balance} -> {user.balance} –∫–æ–ø–µ–µ–∫ (RUB). –ü—Ä–∏—á–∏–Ω–∞: {payload.reason or '–Ω–µ —É–∫–∞–∑–∞–Ω–∞'}",
        )
    )
    await session.commit()
    await session.refresh(user)
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    if amount_rub > 0:
        notification_text = (
            f"üí∞ <b>–ë–∞–ª–∞–Ω—Å –ø–æ–ø–æ–ª–Ω–µ–Ω</b>\n\n"
            f"–°—É–º–º–∞: <b>+{amount_rub:.2f} RUB</b>\n"
            f"–¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>"
        )
        if payload.reason:
            notification_text += f"\n\n–ü—Ä–∏—á–∏–Ω–∞: {payload.reason}"
    else:
        notification_text = (
            f"üí∞ <b>–ò–∑–º–µ–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞</b>\n\n"
            f"–°—É–º–º–∞: <b>{amount_rub:.2f} RUB</b>\n"
            f"–¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>"
        )
        if payload.reason:
            notification_text += f"\n\n–ü—Ä–∏—á–∏–Ω–∞: {payload.reason}"
    
    asyncio.create_task(_send_user_notification(payload.tg_id, notification_text))
    
    return {"tg_id": user.tg_id, "balance": user.balance, "new_balance_cents": user.balance}


@app.post("/admin/users/set_active")
async def admin_set_active(
    payload: AdminSetActiveIn,
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> dict[str, int | bool]:
    user = await session.scalar(select(User).where(User.tg_id == payload.tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    old_status = user.is_active
    user.is_active = bool(payload.is_active)
    action_type = AuditLogAction.user_unblocked if payload.is_active else AuditLogAction.user_blocked
    session.add(
        AuditLog(
            action=action_type,
            user_tg_id=payload.tg_id,
            details=f"–°—Ç–∞—Ç—É—Å –∏–∑–º–µ–Ω–µ–Ω: {'–∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω' if old_status else '–∞–∫—Ç–∏–≤–µ–Ω'} -> {'–∞–∫—Ç–∏–≤–µ–Ω' if payload.is_active else '–∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω'}",
        )
    )
    await session.commit()
    return {"tg_id": user.tg_id, "is_active": user.is_active}


@app.post("/admin/users/block")
async def admin_block_user(
    payload: dict[str, int],
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> dict[str, int | bool]:
    tg_id = payload.get("tg_id")
    if not tg_id:
        raise HTTPException(status_code=400, detail="tg_id required")
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    user.is_active = False
    session.add(
        AuditLog(
            action=AuditLogAction.user_blocked,
            user_tg_id=tg_id,
            details="–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω",
        )
    )
    await session.commit()
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    notification_text = (
        f"‚ùå <b>–ê–∫–∫–∞—É–Ω—Ç –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
        f"–í–∞—à –∞–∫–∫–∞—É–Ω—Ç –±—ã–ª –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º."
    )
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    return {"tg_id": user.tg_id, "is_active": False}


@app.post("/admin/users/unblock")
async def admin_unblock_user(
    payload: dict[str, int],
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> dict[str, int | bool]:
    tg_id = payload.get("tg_id")
    if not tg_id:
        raise HTTPException(status_code=400, detail="tg_id required")
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    user.is_active = True
    session.add(
        AuditLog(
            action=AuditLogAction.user_unblocked,
            user_tg_id=tg_id,
            details="–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω",
        )
    )
    await session.commit()
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    notification_text = (
        f"‚úÖ <b>–ê–∫–∫–∞—É–Ω—Ç —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
        f"–í–∞—à –∞–∫–∫–∞—É–Ω—Ç –±—ã–ª —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º."
    )
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    return {"tg_id": user.tg_id, "is_active": True}


@app.get("/admin/users/export.csv")
async def admin_export_users_csv(
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> StreamingResponse:
    result = await session.scalars(select(User).order_by(User.created_at.desc()))
    users: Sequence[User] = result.all()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "tg_id", "is_active", "balance", "referral_code", "referred_by_tg_id", "created_at"])
    for u in users:
        writer.writerow(
            [
                u.id,
                u.tg_id,
                u.is_active,
                u.balance,
                u.referral_code or "",
                (u.referred_by.tg_id if u.referred_by else ""),
                u.created_at.isoformat(),
            ]
        )
    buf.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="users_export.csv"'}
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/admin/users/export.xlsx")
async def admin_export_users_xlsx(
    session: AsyncSession = Depends(get_session),
    _admin: None = Depends(_require_admin),
):
    """–≠–∫—Å–ø–æ—Ä—Ç –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –≤ Excel —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º"""
    result = await session.scalars(select(User).order_by(User.created_at.desc()))
    users = result.all()
    
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏"
    
    # –ó–∞–≥–æ–ª–æ–≤–∫–∏
    headers = ["ID", "Telegram ID", "–ò–º—è", "Username", "–ë–∞–ª–∞–Ω—Å (USD)", "–°—Ç–∞—Ç—É—Å", "–†–æ–ª—å", "–†–µ—Ñ–µ—Ä–∞–ª—å–Ω—ã–π –∫–æ–¥", "–†–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏—è"]
    ws.append(headers)
    
    # –°—Ç–∏–ª–∏ –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # –î–∞–Ω–Ω—ã–µ
    for user in users:
        balance_rub = user.balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
        status = "–ê–∫—Ç–∏–≤–µ–Ω" if user.is_active else "–ó–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω"
        role = "–ê–¥–º–∏–Ω" if user.tg_id in admin_ids else "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å"
        
        row = [
            user.id,
            user.tg_id,
            f"{user.first_name or ''} {user.last_name or ''}".strip() or "‚Äî",
            user.username or "‚Äî",
            balance_rub,
            status,
            role,
            user.referral_code or "‚Äî",
            user.created_at.strftime("%d.%m.%Y %H:%M") if user.created_at else "‚Äî",
        ]
        ws.append(row)
        
        # –°—Ç–∏–ª–∏ –¥–ª—è —Å—Ç—Ä–æ–∫
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # –¶–≤–µ—Ç –¥–ª—è —Å—Ç–∞—Ç—É—Å–∞
            if cell.column == 6:  # –°—Ç–∞—Ç—É—Å
                if cell.value == "–ê–∫—Ç–∏–≤–µ–Ω":
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    
    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –ø–∞–º—è—Ç—å
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="users_export.xlsx"'},
    )


@app.get("/admin/web/export/logs.csv")
async def admin_web_export_logs_csv(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
) -> StreamingResponse:
    """–≠–∫—Å–ø–æ—Ä—Ç –ª–æ–≥–æ–≤ –≤ CSV"""
    result = await session.scalars(
        select(AuditLog)
        .order_by(AuditLog.created_at.desc())
    )
    logs = result.all()
    
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "action", "user_tg_id", "admin_tg_id", "details", "created_at"])
    for log in logs:
        writer.writerow([
            log.id,
            log.action.value if hasattr(log.action, "value") else str(log.action),
            log.user_tg_id or "",
            log.admin_tg_id or "",
            (log.details or "").replace("\n", " ").replace("\r", " "),
            log.created_at.isoformat(),
        ])
    buf.seek(0)
    
    headers = {"Content-Disposition": 'attachment; filename="logs_export.csv"'}
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/admin/web/export/logs.xlsx")
async def admin_web_export_logs_xlsx(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–≠–∫—Å–ø–æ—Ä—Ç –ª–æ–≥–æ–≤ –≤ Excel —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º"""
    result = await session.scalars(
        select(AuditLog)
        .order_by(AuditLog.created_at.desc())
    )
    logs = result.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "–õ–æ–≥–∏"
    
    headers = ["ID", "–î–µ–π—Å—Ç–≤–∏–µ", "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å", "–ê–¥–º–∏–Ω", "–î–µ—Ç–∞–ª–∏", "–í—Ä–µ–º—è"]
    ws.append(headers)
    
    # –°—Ç–∏–ª–∏ –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # –¶–≤–µ—Ç–∞ –¥–ª—è –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –¥–µ–π—Å—Ç–≤–∏–π
    critical_actions = ['user_blocked', 'user_unblocked', 'balance_credited', 'role_changed']
    warning_actions = ['ticket_created', 'subscription_created']
    
    for log in logs:
        action_str = log.action.value if hasattr(log.action, "value") else str(log.action)
        is_critical = action_str in critical_actions
        is_warning = action_str in warning_actions
        
        row = [
            log.id,
            action_str,
            log.user_tg_id or "‚Äî",
            log.admin_tg_id or "‚Äî",
            (log.details or "‚Äî").replace("\n", " ").replace("\r", " "),
            log.created_at.strftime("%d.%m.%Y %H:%M") if log.created_at else "‚Äî",
        ]
        ws.append(row)
        
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # –¶–≤–µ—Ç –¥–ª—è –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –¥–µ–π—Å—Ç–≤–∏–π
            if cell.column == 2:  # –î–µ–π—Å—Ç–≤–∏–µ
                if is_critical:
                    cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
                elif is_warning:
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    
    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="logs_export.xlsx"'},
    )


@app.get("/admin/web/export/tickets.csv")
async def admin_web_export_tickets_csv(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
) -> StreamingResponse:
    """–≠–∫—Å–ø–æ—Ä—Ç —Ç–∏–∫–µ—Ç–æ–≤ –≤ CSV"""
    result = await session.scalars(
        select(Ticket)
        .order_by(Ticket.created_at.desc())
    )
    tickets = result.all()
    
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "user_tg_id", "topic", "status", "created_at", "updated_at", "closed_at"])
    for ticket in tickets:
        writer.writerow([
            ticket.id,
            ticket.user_tg_id,
            (ticket.topic or "").replace("\n", " ").replace("\r", " "),
            ticket.status.value if hasattr(ticket.status, "value") else str(ticket.status),
            ticket.created_at.isoformat() if ticket.created_at else "",
            ticket.updated_at.isoformat() if ticket.updated_at else "",
            ticket.closed_at.isoformat() if ticket.closed_at else "",
        ])
    buf.seek(0)
    
    headers = {"Content-Disposition": 'attachment; filename="tickets_export.csv"'}
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/admin/web/export/tickets.xlsx")
async def admin_web_export_tickets_xlsx(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–≠–∫—Å–ø–æ—Ä—Ç —Ç–∏–∫–µ—Ç–æ–≤ –≤ Excel —Å —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ–º"""
    result = await session.scalars(
        select(Ticket)
        .order_by(Ticket.created_at.desc())
    )
    tickets = result.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "–¢–∏–∫–µ—Ç—ã"
    
    headers = ["ID", "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å", "–¢–µ–º–∞", "–°—Ç–∞—Ç—É—Å", "–°–æ–∑–¥–∞–Ω", "–û–±–Ω–æ–≤–ª—ë–Ω"]
    ws.append(headers)
    
    # –°—Ç–∏–ª–∏ –¥–ª—è –∑–∞–≥–æ–ª–æ–≤–∫–æ–≤
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for ticket in tickets:
        status_str = ticket.status.value if hasattr(ticket.status, "value") else str(ticket.status)
        
        row = [
            ticket.id,
            ticket.user_tg_id,
            ticket.topic or "‚Äî",
            status_str,
            ticket.created_at.strftime("%d.%m.%Y %H:%M") if ticket.created_at else "‚Äî",
            ticket.updated_at.strftime("%d.%m.%Y %H:%M") if ticket.updated_at else "‚Äî",
        ]
        ws.append(row)
        
        last_row = ws.max_row
        for cell in ws[last_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # –¶–≤–µ—Ç –¥–ª—è —Å—Ç–∞—Ç—É—Å–∞
            if cell.column == 4:  # –°—Ç–∞—Ç—É—Å
                if status_str == "closed":
                    cell.fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
                elif status_str == "in_progress":
                    cell.fill = PatternFill(start_color="d1ecf1", end_color="d1ecf1", fill_type="solid")
                elif status_str == "new":
                    cell.fill = PatternFill(start_color="fff3cd", end_color="fff3cd", fill_type="solid")
    
    # –ê–≤—Ç–æ–ø–æ–¥–±–æ—Ä —à–∏—Ä–∏–Ω—ã –∫–æ–ª–æ–Ω–æ–∫
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tickets_export.xlsx"'},
    )


@app.get("/admin/logs")
async def admin_get_logs(
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    action: AuditLogAction | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> list[AuditLogOut]:
    stmt = select(AuditLog).order_by(AuditLog.created_at.desc()).limit(limit).offset(offset)
    if action:
        stmt = stmt.where(AuditLog.action == action)
    result = await session.scalars(stmt)
    logs: Sequence[AuditLog] = result.all()
    return [
        AuditLogOut(
            id=log.id,
            action=log.action.value,
            user_tg_id=log.user_tg_id,
            admin_tg_id=log.admin_tg_id,
            details=log.details,
            created_at=log.created_at,
        )
        for log in logs
    ]


@app.get("/admin/logs/count")
async def admin_logs_count(
    action: AuditLogAction | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    _: None = Depends(_require_admin),
) -> dict[str, int]:
    stmt = select(func.count()).select_from(AuditLog)
    if action:
        stmt = stmt.where(AuditLog.action == action)
    total = await session.scalar(stmt)
    return {"total": int(total or 0)}


# --- Web Admin Interface ---
def _require_web_admin(request: Request) -> dict:
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ –≤–µ–±-–∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ —Å–µ—Å—Å–∏—é"""
    session_data = request.session.get("admin_user")
    if not session_data:
        raise HTTPException(status_code=403, detail="not_authenticated")
    return session_data


@app.get("/admin/login", response_class=HTMLResponse)
async def admin_login_page(request: Request):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ Telegram"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    # –ï—Å–ª–∏ —É–∂–µ –∞–≤—Ç–æ—Ä–∏–∑–æ–≤–∞–Ω, —Ä–µ–¥–∏—Ä–µ–∫—Ç –Ω–∞ –∞–¥–º–∏–Ω–∫—É
    if request.session.get("admin_user"):
        return RedirectResponse(url="/admin/web", status_code=303)
    
    # –ü–æ–ª—É—á–∞–µ–º bot_username –∏–∑ –ø–µ—Ä–µ–º–µ–Ω–Ω–æ–π –æ–∫—Ä—É–∂–µ–Ω–∏—è
    bot_username_raw = os.getenv("BOT_USERNAME", "").strip()
    bot_username = bot_username_raw
    
    # –ï—Å–ª–∏ –Ω–µ –∑–∞–¥–∞–Ω, –ø—ã—Ç–∞–µ–º—Å—è –ø–æ–ª—É—á–∏—Ç—å —á–µ—Ä–µ–∑ Bot API
    if not bot_username:
        bot_token = os.getenv("BOT_TOKEN", "").strip()
        if bot_token:
            try:
                import httpx
                async with httpx.AsyncClient(timeout=10.0) as client:
                    r = await client.get(f"https://api.telegram.org/bot{bot_token}/getMe")
                    if r.status_code == 200:
                        data = r.json()
                        if data.get("ok"):
                            bot_username = data.get("result", {}).get("username", "").strip()
            except Exception as e:
                # –õ–æ–≥–∏—Ä—É–µ–º –æ—à–∏–±–∫—É –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
                import logging
                logging.warning(f"Failed to get bot username via API: {e}")
                pass
    
    # –ù–æ—Ä–º–∞–ª–∏–∑—É–µ–º username
    if bot_username.startswith("@"):
        bot_username = bot_username[1:]
    bot_username = bot_username.strip()

    # Telegram Login Widget –ø—Ä–∏–Ω–∏–º–∞–µ—Ç username –±–æ—Ç–∞ (–æ–±—ã—á–Ω–æ –∑–∞–∫–∞–Ω—á–∏–≤–∞–µ—Ç—Å—è –Ω–∞ "bot").
    # –ï—Å–ª–∏ —É –Ω–∞—Å —è–≤–Ω–æ –∑–∞–¥–∞–Ω–æ —á—Ç–æ-—Ç–æ –≤—Ä–æ–¥–µ "fioreVPN" (–∏–º—è, –∞ –Ω–µ username), –ª—É—á—à–µ –ø–æ–∫–∞–∑–∞—Ç—å –æ—à–∏–±–∫—É.
    if bot_username and not bot_username.lower().endswith("bot"):
        bot_username = ""
    
    return templates.TemplateResponse("login.html", {
        "request": request,
        "bot_username": bot_username or "",
        "bot_username_raw": bot_username_raw or "",
    })


def _verify_telegram_auth(data: dict, bot_token: str) -> bool:
    """–ü—Ä–æ–≤–µ—Ä–∫–∞ –ø–æ–¥–ø–∏—Å–∏ –¥–∞–Ω–Ω—ã—Ö –æ—Ç Telegram Login Widget"""
    try:
        # –ü–æ–ª—É—á–∞–µ–º hash –∏–∑ –¥–∞–Ω–Ω—ã—Ö
        received_hash = data.pop("hash", "")
        if not received_hash:
            return False
        
        # –°–æ—Ä—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –∏ —Å–æ–∑–¥–∞–µ–º —Å—Ç—Ä–æ–∫—É –¥–ª—è –ø—Ä–æ–≤–µ—Ä–∫–∏
        data_check_string = "\n".join(f"{k}={v}" for k, v in sorted(data.items()))
        
        # Telegram Login Widget:
        # secret_key = SHA256(bot_token)
        secret_key = hashlib.sha256(bot_token.encode()).digest()
        
        # –í—ã—á–∏—Å–ª—è–µ–º hash
        calculated_hash = hmac.new(
            key=secret_key,
            msg=data_check_string.encode(),
            digestmod=hashlib.sha256
        ).hexdigest()
        
        return calculated_hash == received_hash
    except Exception:
        return False


@app.get("/admin/auth/telegram")
async def admin_auth_telegram(request: Request):
    """–û–±—Ä–∞–±–æ—Ç–∫–∞ –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ Telegram Login Widget"""
    # –ë–µ—Ä—ë–º —Ä–æ–≤–Ω–æ —Ç–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã, –∫–æ—Ç–æ—Ä—ã–µ –ø—Ä–∏—Å–ª–∞–ª Telegram (–≤–∞–∂–Ω–æ –¥–ª—è –ø–æ–¥–ø–∏—Å–∏!)
    qp = dict(request.query_params)
    try:
        tg_id = int(qp.get("id", "0"))
        auth_date = int(qp.get("auth_date", "0"))
        received_hash = qp.get("hash", "")
        first_name = qp.get("first_name", "")
        last_name = qp.get("last_name", "")
        username = qp.get("username", "")
        photo_url = qp.get("photo_url", "")
    except Exception:
        return RedirectResponse(url="/admin/login?error=invalid_request", status_code=303)

    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∞–¥–º–∏–Ω
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()
    
    if tg_id not in admin_ids:
        return RedirectResponse(url="/admin/login?error=not_admin", status_code=303)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ–¥–ø–∏—Å—å Telegram
    bot_token = os.getenv("BOT_TOKEN", "")
    if not bot_token:
        return RedirectResponse(url="/admin/login?error=server_misconfigured", status_code=303)

    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ–¥–ø–∏—Å—å –ø–æ –æ—Ä–∏–≥–∏–Ω–∞–ª—å–Ω—ã–º query params
    auth_data = dict(qp)
    auth_data["hash"] = received_hash
    if not _verify_telegram_auth(auth_data.copy(), bot_token):
        return RedirectResponse(url="/admin/login?error=invalid_signature", status_code=303)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –¥–∞–Ω–Ω—ã–µ –Ω–µ —É—Å—Ç–∞—Ä–µ–ª–∏ (–Ω–µ —Å—Ç–∞—Ä—à–µ 24 —á–∞—Å–æ–≤)
    current_time = int(time.time())
    if current_time - auth_date > 86400:  # 24 —á–∞—Å–∞
        return RedirectResponse(url="/admin/login?error=expired", status_code=303)
    
    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–∞–Ω–Ω—ã–µ –≤ —Å–µ—Å—Å–∏—é
    request.session["admin_user"] = {
        "tg_id": tg_id,
        "first_name": first_name,
        "last_name": last_name,
        "username": username,
        "photo_url": photo_url,
    }
    _get_csrf_token(request)
    
    return RedirectResponse(url="/admin/web", status_code=303)


@app.get("/admin/logout")
async def admin_logout(request: Request):
    """–í—ã—Ö–æ–¥ –∏–∑ –∞–¥–º–∏–Ω–∫–∏"""
    request.session.clear()
    resp = RedirectResponse(
        url="/admin/login?force_reauth=1" if "drop_session" in request.query_params else "/admin/login",
        status_code=303,
    )
    # –Ø–≤–Ω–æ —É–¥–∞–ª—è–µ–º cookie —Å–µ—Å—Å–∏–∏, —á—Ç–æ–±—ã –Ω–µ –æ—Å—Ç–∞–ª–æ—Å—å –ø–æ–¥–ø–∏—Å–∞–Ω–Ω–æ–≥–æ —Å–æ—Å—Ç–æ—è–Ω–∏—è
    resp.delete_cookie(key="session")
    return resp


@app.get("/admin/web/dashboard", response_class=HTMLResponse)
async def admin_web_dashboard(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–î–∞—à–±–æ—Ä–¥ —Å–æ —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–æ–π"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    from datetime import timedelta
    now_utc = datetime.utcnow()
    today_start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = now_utc - timedelta(days=7)
    month_start = now_utc - timedelta(days=30)
    
    # –û–±—â–∞—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞
    total_users = await session.scalar(select(func.count()).select_from(User))
    total_active = await session.scalar(select(func.count()).select_from(User).where(User.is_active == True))
    total_blocked = await session.scalar(select(func.count()).select_from(User).where(User.is_active == False))
    total_balance = await session.scalar(select(func.sum(User.balance)).select_from(User)) or 0
    total_balance_rub = total_balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
    
    # –ü–æ–¥–ø–∏—Å–∫–∏
    total_subscriptions = await session.scalar(select(func.count()).select_from(Subscription))
    active_subscriptions = await session.scalar(
        select(func.count()).select_from(Subscription).where(Subscription.status == SubscriptionStatus.active)
    )
    
    # –ü–ª–∞—Ç–µ–∂–∏
    total_payments = await session.scalar(select(func.count()).select_from(Payment))
    succeeded_payments = await session.scalar(
        select(func.count()).select_from(Payment).where(Payment.status == PaymentStatus.succeeded)
    )
    total_revenue = await session.scalar(
        select(func.sum(Payment.amount_cents)).select_from(Payment).where(Payment.status == PaymentStatus.succeeded)
    ) or 0
    total_revenue_rub = total_revenue / 100  # –í—ã—Ä—É—á–∫–∞ —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
    
    # –¢–∏–∫–µ—Ç—ã
    total_tickets = await session.scalar(select(func.count()).select_from(Ticket))
    new_tickets = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.status == TicketStatus.new)
    )
    in_progress_tickets = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.status == TicketStatus.in_progress)
    )
    closed_tickets = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.status == TicketStatus.closed)
    )
    
    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –∑–∞ –ø–µ—Ä–∏–æ–¥—ã
    users_today = await session.scalar(
        select(func.count()).select_from(User).where(User.created_at >= today_start)
    )
    users_week = await session.scalar(
        select(func.count()).select_from(User).where(User.created_at >= week_start)
    )
    users_month = await session.scalar(
        select(func.count()).select_from(User).where(User.created_at >= month_start)
    )
    
    payments_today = await session.scalar(
        select(func.count()).select_from(Payment)
        .where(Payment.created_at >= today_start, Payment.status == PaymentStatus.succeeded)
    )
    payments_week = await session.scalar(
        select(func.count()).select_from(Payment)
        .where(Payment.created_at >= week_start, Payment.status == PaymentStatus.succeeded)
    )
    payments_month = await session.scalar(
        select(func.count()).select_from(Payment)
        .where(Payment.created_at >= month_start, Payment.status == PaymentStatus.succeeded)
    )
    
    revenue_today = await session.scalar(
        select(func.sum(Payment.amount_cents)).select_from(Payment)
        .where(Payment.created_at >= today_start, Payment.status == PaymentStatus.succeeded)
    ) or 0
    revenue_week = await session.scalar(
        select(func.sum(Payment.amount_cents)).select_from(Payment)
        .where(Payment.created_at >= week_start, Payment.status == PaymentStatus.succeeded)
    ) or 0
    revenue_month = await session.scalar(
        select(func.sum(Payment.amount_cents)).select_from(Payment)
        .where(Payment.created_at >= month_start, Payment.status == PaymentStatus.succeeded)
    ) or 0
    
    tickets_today = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.created_at >= today_start)
    )
    tickets_week = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.created_at >= week_start)
    )
    tickets_month = await session.scalar(
        select(func.count()).select_from(Ticket).where(Ticket.created_at >= month_start)
    )
    
    # –¢–æ–ø –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –ø–æ –±–∞–ª–∞–Ω—Å—É
    top_users_result = await session.scalars(
        select(User)
        .order_by(User.balance.desc())
        .limit(10)
    )
    top_users = [
        {
            "tg_id": u.tg_id,
            "username": u.username or "‚Äî",
            "balance": u.balance / 100,
        }
        for u in top_users_result.all()
    ]
    
    # –ü–æ—Å–ª–µ–¥–Ω–∏–µ –¥–µ–π—Å—Ç–≤–∏—è
    recent_logs_result = await session.scalars(
        select(AuditLog)
        .order_by(AuditLog.created_at.desc())
        .limit(10)
    )
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    recent_logs = [
        {
            "id": log.id,
            "action": log.action.value if hasattr(log.action, "value") else str(log.action),
            "user_tg_id": log.user_tg_id,
            "admin_tg_id": log.admin_tg_id,
            "details": log.details or "‚Äî",
            "created_at": fmt(log.created_at),
        }
        for log in recent_logs_result.all()
    ]
    
    # –ü–æ–ª—É—á–∞–µ–º –∞–ª–µ—Ä—Ç—ã
    hour_ago = now_utc - timedelta(hours=1)
    day_ago = now_utc - timedelta(days=1)
    
    dashboard_alerts = []
    
    # –ú–Ω–æ–≥–æ –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤
    if new_tickets >= 5:
        dashboard_alerts.append({
            "type": "warning",
            "title": "–ú–Ω–æ–≥–æ –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤",
            "message": f"–°–æ–∑–¥–∞–Ω–æ {new_tickets} –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤. –¢—Ä–µ–±—É–µ—Ç—Å—è –≤–Ω–∏–º–∞–Ω–∏–µ!",
            "link": "/admin/web/tickets?status=all",
        })
    
    # –ó–∞–≤–∏—Å—à–∏–µ —Ç–∏–∫–µ—Ç—ã
    stale_tickets = await session.scalars(
        select(Ticket)
        .where(
            Ticket.status == TicketStatus.in_progress,
            Ticket.updated_at < day_ago
        )
    )
    stale_count = len(stale_tickets.all())
    if stale_count > 0:
        dashboard_alerts.append({
            "type": "warning",
            "title": "–ó–∞–≤–∏—Å—à–∏–µ —Ç–∏–∫–µ—Ç—ã",
            "message": f"{stale_count} —Ç–∏–∫–µ—Ç–æ–≤ –≤ —Ä–∞–±–æ—Ç–µ –Ω–µ –æ–±–Ω–æ–≤–ª—è–ª–∏—Å—å –±–æ–ª–µ–µ 24 —á–∞—Å–æ–≤",
            "link": "/admin/web/tickets?status=all",
        })
    
    # –û—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–π –±–∞–ª–∞–Ω—Å
    negative_balance_count = await session.scalar(
        select(func.count()).select_from(User).where(User.balance < 0)
    )
    if negative_balance_count and negative_balance_count > 0:
        dashboard_alerts.append({
            "type": "danger",
            "title": "–û—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–π –±–∞–ª–∞–Ω—Å",
            "message": f"{negative_balance_count} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–º–µ—é—Ç –æ—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–π –±–∞–ª–∞–Ω—Å",
            "link": "/admin/web/users?balance_max=0",
        })
    
    # –ù–∏–∑–∫–∏–π –æ–±—â–∏–π –±–∞–ª–∞–Ω—Å (–º–µ–Ω–µ–µ $100)
    if total_balance_rub < 100:
        dashboard_alerts.append({
            "type": "info",
            "title": "–ù–∏–∑–∫–∏–π –æ–±—â–∏–π –±–∞–ª–∞–Ω—Å",
            "message": f"–û–±—â–∏–π –±–∞–ª–∞–Ω—Å –≤—Å–µ—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {total_balance_rub:.2f} RUB",
        })
    
    # –ü–æ–ª—É—á–∞–µ–º —Å—Ç–∞—Ç—É—Å —Å–µ—Ä–≤–µ—Ä–æ–≤
    servers = await session.scalars(select(Server).where(Server.is_enabled == True))
    servers_list = servers.all()
    servers_status = []
    
    for server in servers_list:
        # –ü–æ–ª—É—á–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–π —Å—Ç–∞—Ç—É—Å —Å–µ—Ä–≤–µ—Ä–∞
        latest_status = await session.scalar(
            select(ServerStatus)
            .where(ServerStatus.server_id == server.id)
            .order_by(ServerStatus.checked_at.desc())
            .limit(1)
        )
        
        servers_status.append({
            "id": server.id,
            "name": server.name,
            "host": server.host,
            "location": server.location or "‚Äî",
            "is_online": latest_status.is_online if latest_status else False,
            "response_time_ms": latest_status.response_time_ms if latest_status else None,
            "active_connections": latest_status.active_connections if latest_status else 0,
            "capacity": server.capacity,
            "checked_at": latest_status.checked_at if latest_status else None,
            "error_message": latest_status.error_message if latest_status else None,
        })
        
        # –î–æ–±–∞–≤–ª—è–µ–º –∞–ª–µ—Ä—Ç, –µ—Å–ª–∏ —Å–µ—Ä–≤–µ—Ä –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω
        if latest_status and not latest_status.is_online:
            dashboard_alerts.append({
                "type": "danger",
                "title": f"–°–µ—Ä–≤–µ—Ä {server.name} –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω",
                "message": f"–°–µ—Ä–≤–µ—Ä {server.name} ({server.host}) –Ω–µ –æ—Ç–≤–µ—á–∞–µ—Ç. {latest_status.error_message or '–ü—Ä–æ–≤–µ—Ä—å—Ç–µ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–µ'}",
                "link": "/admin/web/servers",
            })
    
    csrf_token = _get_csrf_token(request)
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "admin_user": admin_user,
        "csrf_token": csrf_token,
        "total_users": int(total_users or 0),
        "total_active": int(total_active or 0),
        "total_blocked": int(total_blocked or 0),
        "total_balance_rub": total_balance_rub,
        "total_subscriptions": int(total_subscriptions or 0),
        "active_subscriptions": int(active_subscriptions or 0),
        "total_payments": int(total_payments or 0),
        "succeeded_payments": int(succeeded_payments or 0),
        "total_revenue_rub": total_revenue_rub,
        "total_tickets": int(total_tickets or 0),
        "new_tickets": int(new_tickets or 0),
        "in_progress_tickets": int(in_progress_tickets or 0),
        "closed_tickets": int(closed_tickets or 0),
        "users_today": int(users_today or 0),
        "users_week": int(users_week or 0),
        "users_month": int(users_month or 0),
        "payments_today": int(payments_today or 0),
        "payments_week": int(payments_week or 0),
        "payments_month": int(payments_month or 0),
        "revenue_today": revenue_today / 100,
        "revenue_week": revenue_week / 100,
        "revenue_month": revenue_month / 100,
        "tickets_today": int(tickets_today or 0),
        "tickets_week": int(tickets_week or 0),
        "tickets_month": int(tickets_month or 0),
        "top_users": top_users,
        "recent_logs": recent_logs,
        "alerts": dashboard_alerts,
        "servers_status": servers_status,
    })


@app.get("/admin/web/api/backups/recent")
async def admin_api_recent_backups(
    limit: int = Query(default=5, ge=1, le=10),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –ø–æ—Å–ª–µ–¥–Ω–∏—Ö –±—ç–∫–∞–ø–æ–≤"""
    backups_result = await session.scalars(
        select(Backup)
        .order_by(Backup.created_at.desc())
        .limit(limit)
    )
    backups = backups_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    backups_data = []
    for backup in backups:
        file_size_mb = backup.file_size_bytes / (1024 * 1024) if backup.file_size_bytes > 0 else 0
        
        backups_data.append({
            "id": backup.id,
            "backup_type": backup.backup_type,
            "file_size_mb": file_size_mb,
            "status": backup.status,
            "error_message": backup.error_message,
            "created_at": fmt(backup.created_at),
        })
    
    return {"backups": backups_data}


@app.get("/admin/web/api/dashboard/stats")
async def admin_api_dashboard_stats(
    days: int = Query(default=30, ge=7, le=365),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏ –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–æ–≤"""
    from datetime import timedelta
    now_utc = datetime.utcnow()
    start_date = now_utc - timedelta(days=days)
    
    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–π –ø–æ –¥–Ω—è–º
    registrations_by_day = await session.execute(
        select(
            func.date(User.created_at).label('date'),
            func.count(User.id).label('count')
        )
        .where(User.created_at >= start_date)
        .group_by(func.date(User.created_at))
        .order_by(func.date(User.created_at))
    )
    registrations_data = registrations_by_day.all()
    
    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ –ø–ª–∞—Ç–µ–∂–µ–π –ø–æ –¥–Ω—è–º
    payments_by_day = await session.execute(
        select(
            func.date(Payment.created_at).label('date'),
            func.count(Payment.id).label('count'),
            func.sum(Payment.amount_cents).label('revenue')
        )
        .where(
            Payment.created_at >= start_date,
            Payment.status == PaymentStatus.succeeded
        )
        .group_by(func.date(Payment.created_at))
        .order_by(func.date(Payment.created_at))
    )
    payments_data = payments_by_day.all()
    
    # –°—Ç–∞—Ç–∏—Å—Ç–∏–∫–∞ —Ç–∏–∫–µ—Ç–æ–≤ –ø–æ –¥–Ω—è–º
    tickets_by_day = await session.execute(
        select(
            func.date(Ticket.created_at).label('date'),
            func.count(Ticket.id).label('count')
        )
        .where(Ticket.created_at >= start_date)
        .group_by(func.date(Ticket.created_at))
        .order_by(func.date(Ticket.created_at))
    )
    tickets_data = tickets_by_day.all()
    
    # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ
    def format_date(dt):
        if dt:
            # –ï—Å–ª–∏ —ç—Ç–æ date, –ø—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –≤ —Å—Ç—Ä–æ–∫—É –Ω–∞–ø—Ä—è–º—É—é
            if isinstance(dt, datetime):
                try:
                    from zoneinfo import ZoneInfo
                    moscow_tz = ZoneInfo("Europe/Moscow")
                    return dt.astimezone(moscow_tz).strftime("%d.%m")
                except Exception:
                    return dt.strftime("%d.%m")
            else:
                # –ï—Å–ª–∏ —ç—Ç–æ date –æ–±—ä–µ–∫—Ç
                return dt.strftime("%d.%m")
        return ""
    
    return {
        "registrations": [
            {"date": format_date(r.date), "count": r.count}
            for r in registrations_data
        ],
        "payments": [
            {"date": format_date(p.date), "count": p.count, "revenue": (p.revenue or 0) / 100}
            for p in payments_data
        ],
        "tickets": [
            {"date": format_date(t.date), "count": t.count}
            for t in tickets_data
        ],
    }


@app.get("/admin/web", response_class=HTMLResponse)
async def admin_web(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ì–ª–∞–≤–Ω–∞—è —Å—Ç—Ä–∞–Ω–∏—Ü–∞ - –ø–µ—Ä–µ–Ω–∞–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –Ω–∞ –¥–∞—à–±–æ—Ä–¥"""
    return RedirectResponse(url="/admin/web/dashboard", status_code=303)


@app.get("/admin/web/api/users")
async def admin_api_users(
    q: str | None = Query(default=None),
    status: str | None = Query(default="all"),
    role: str | None = Query(default="all"),
    reg_period: str | None = Query(default="all"),
    balance_min: str | None = Query(default=None),
    balance_max: str | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è —Å–ø–∏—Å–∫–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (–¥–ª—è –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è)"""
    # –í–∞–ª–∏–¥–∏—Ä—É–µ–º —Ñ–∏–ª—å—Ç—Ä—ã (—Ç–∞ –∂–µ –ª–æ–≥–∏–∫–∞, —á—Ç–æ –∏ –≤ admin_web_users)
    status_filter = (status or "all").lower()
    if status_filter not in {"all", "active", "blocked"}:
        status_filter = "all"
    
    role_filter = (role or "all").lower()
    if role_filter not in {"all", "superadmin", "admin", "moderator", "user"}:
        role_filter = "all"
    
    reg_filter = (reg_period or "all").lower()
    if reg_filter not in {"all", "today", "7d", "30d"}:
        reg_filter = "all"
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (—Å –ø–æ–∏—Å–∫–æ–º/—Ñ–∏–ª—å—Ç—Ä–∞–º–∏)
    stmt = select(User).options(selectinload(User.referred_by)).order_by(User.created_at.desc())
    if status_filter == "active":
        stmt = stmt.where(User.is_active == True)
    elif status_filter == "blocked":
        stmt = stmt.where(User.is_active == False)
    
    # –§–∏–ª—å—Ç—Ä –ø–æ –ø–µ—Ä–∏–æ–¥—É —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏
    if reg_filter != "all":
        now_utc = datetime.utcnow()
        if reg_filter == "today":
            start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
        elif reg_filter == "7d":
            start = now_utc - timedelta(days=7)
        else:  # "30d"
            start = now_utc - timedelta(days=30)
        stmt = stmt.where(User.created_at >= start)
    
    if q:
        qq = q.strip()
        if qq.isdigit():
            stmt = stmt.where(User.tg_id == int(qq))
        else:
            like = f"%{qq.lower()}%"
            stmt = stmt.where(
                func.lower(User.username).like(like)
                | func.lower(User.first_name).like(like)
                | func.lower(User.last_name).like(like)
            )
    
    # –§–∏–ª—å—Ç—Ä –ø–æ –±–∞–ª–∞–Ω—Å—É
    if balance_min:
        try:
            min_balance = int(float(balance_min) * 100)
            stmt = stmt.where(User.balance >= min_balance)
        except (ValueError, TypeError):
            pass
    if balance_max:
        try:
            max_balance = int(float(balance_max) * 100)
            stmt = stmt.where(User.balance <= max_balance)
        except (ValueError, TypeError):
            pass
    
    result = await session.scalars(stmt)
    users: Sequence[User] = result.all()
    
    # –ü–æ–¥—Ç—è–≥–∏–≤–∞–µ–º overrides —Ä–æ–ª–µ–π
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
    
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()
    
    # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ
    users_data = []
    for u in users:
        referred_by_tg_id = u.referred_by.tg_id if u.referred_by else None
        full_name = " ".join(filter(None, [u.first_name or "", u.last_name or ""])) or "‚Äî"
        tag = f"@{u.username}" if u.username else "‚Äî"
        eff_role = _get_effective_role(u.tg_id, admin_ids, overrides_map)
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–∏–ª—å—Ç—Ä –ø–æ —Ä–æ–ª–∏
        if role_filter != "all" and eff_role != role_filter:
            continue
        
        role = "–ì–ª–∞–≤–Ω—ã–π –∞–¥–º–∏–Ω" if eff_role == "superadmin" else ("–ê–¥–º–∏–Ω" if eff_role == "admin" else ("–ú–æ–¥–µ—Ä–∞—Ç–æ—Ä" if eff_role == "moderator" else "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å"))
        balance_rub = u.balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
        
        try:
            from zoneinfo import ZoneInfo
            moscow_tz = ZoneInfo("Europe/Moscow")
            dt_moscow = u.created_at.astimezone(moscow_tz)
            created_str = dt_moscow.strftime("%d.%m.%Y %H:%M")
        except:
            created_str = str(u.created_at)[:16]
        
        users_data.append({
            "id": u.id,
            "tg_id": u.tg_id,
            "username": u.username or "‚Äî",
            "full_name": full_name,
            "tag": tag,
            "role": role,
            "is_active": u.is_active,
            "has_active_subscription": u.has_active_subscription,
            "balance": balance_rub,
            "referral_code": u.referral_code or "‚Äî",
            "referred_by_tg_id": referred_by_tg_id,
            "created_at": created_str,
        })
    
    # –°—á—ë—Ç—á–∏–∫–∏
    total_users = await session.scalar(select(func.count()).select_from(User))
    total_active = await session.scalar(select(func.count()).select_from(User).where(User.is_active == True))
    total_blocked = await session.scalar(select(func.count()).select_from(User).where(User.is_active == False))
    
    return {
        "users": users_data,
        "total_users": int(total_users or 0),
        "total_active": int(total_active or 0),
        "total_blocked": int(total_blocked or 0),
    }


@app.get("/admin/web/api/notifications")
async def admin_api_notifications(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–π –∏ –∞–ª–µ—Ä—Ç–æ–≤"""
    from datetime import timedelta
    now_utc = datetime.utcnow()
    hour_ago = now_utc - timedelta(hours=1)
    day_ago = now_utc - timedelta(days=1)
    
    notifications = []
    alerts = []
    
    # –ù–æ–≤—ã–µ —Ç–∏–∫–µ—Ç—ã (–∑–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å)
    new_tickets_count = await session.scalar(
        select(func.count()).select_from(Ticket)
        .where(Ticket.status == TicketStatus.new, Ticket.created_at >= hour_ago)
    )
    if new_tickets_count and new_tickets_count > 0:
        notifications.append({
            "type": "ticket",
            "severity": "info",
            "title": f"–ù–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤: {new_tickets_count}",
            "message": f"–ó–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å —Å–æ–∑–¥–∞–Ω–æ {new_tickets_count} –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤",
            "link": "/admin/web/tickets?status=all",
            "count": new_tickets_count,
        })
        if new_tickets_count >= 5:
            alerts.append({
                "type": "warning",
                "title": "–ú–Ω–æ–≥–æ –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤",
                "message": f"–°–æ–∑–¥–∞–Ω–æ {new_tickets_count} –Ω–æ–≤—ã—Ö —Ç–∏–∫–µ—Ç–æ–≤ –∑–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å. –¢—Ä–µ–±—É–µ—Ç—Å—è –≤–Ω–∏–º–∞–Ω–∏–µ!",
            })
    
    # –¢–∏–∫–µ—Ç—ã –≤ —Ä–∞–±–æ—Ç–µ –±–µ–∑ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –±–æ–ª–µ–µ 24 —á–∞—Å–æ–≤
    stale_tickets = await session.scalars(
        select(Ticket)
        .where(
            Ticket.status == TicketStatus.in_progress,
            Ticket.updated_at < day_ago
        )
    )
    stale_count = len(stale_tickets.all())
    if stale_count > 0:
        alerts.append({
            "type": "warning",
            "title": "–ó–∞–≤–∏—Å—à–∏–µ —Ç–∏–∫–µ—Ç—ã",
            "message": f"{stale_count} —Ç–∏–∫–µ—Ç–æ–≤ –≤ —Ä–∞–±–æ—Ç–µ –Ω–µ –æ–±–Ω–æ–≤–ª—è–ª–∏—Å—å –±–æ–ª–µ–µ 24 —á–∞—Å–æ–≤",
        })
    
    # –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ —Å –æ—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–º –±–∞–ª–∞–Ω—Å–æ–º
    negative_balance_count = await session.scalar(
        select(func.count()).select_from(User).where(User.balance < 0)
    )
    if negative_balance_count and negative_balance_count > 0:
        alerts.append({
            "type": "danger",
            "title": "–û—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–π –±–∞–ª–∞–Ω—Å",
            "message": f"{negative_balance_count} –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π –∏–º–µ—é—Ç –æ—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω—ã–π –±–∞–ª–∞–Ω—Å",
        })
    
    # –ù–æ–≤—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏ –∑–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å
    new_users_count = await session.scalar(
        select(func.count()).select_from(User).where(User.created_at >= hour_ago)
    )
    if new_users_count and new_users_count > 0:
        notifications.append({
            "type": "user",
            "severity": "success",
            "title": f"–ù–æ–≤—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π: {new_users_count}",
            "message": f"–ó–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å –∑–∞—Ä–µ–≥–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–æ {new_users_count} –Ω–æ–≤—ã—Ö –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π",
            "link": "/admin/web/users?reg_period=today",
            "count": new_users_count,
        })
    
    # –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∏–µ –¥–µ–π—Å—Ç–≤–∏—è –≤ –ª–æ–≥–∞—Ö (–±–ª–æ–∫–∏—Ä–æ–≤–∫–∏, –∏–∑–º–µ–Ω–µ–Ω–∏—è –±–∞–ª–∞–Ω—Å–∞)
    critical_actions = await session.scalars(
        select(AuditLog)
        .where(
            AuditLog.created_at >= hour_ago,
            AuditLog.action.in_([
                AuditLogAction.user_blocked,
                AuditLogAction.balance_credited,
            ])
        )
        .order_by(AuditLog.created_at.desc())
        .limit(10)
    )
    critical_count = len(critical_actions.all())
    if critical_count > 0:
        notifications.append({
            "type": "action",
            "severity": "warning",
            "title": f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –¥–µ–π—Å—Ç–≤–∏–π: {critical_count}",
            "message": f"–ó–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å –≤—ã–ø–æ–ª–Ω–µ–Ω–æ {critical_count} –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –¥–µ–π—Å—Ç–≤–∏–π",
            "link": "/admin/web/logs",
            "count": critical_count,
        })
    
    # –ü–ª–∞—Ç–µ–∂–∏ –∑–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å
    recent_payments = await session.scalar(
        select(func.count()).select_from(Payment)
        .where(
            Payment.created_at >= hour_ago,
            Payment.status == PaymentStatus.succeeded
        )
    )
    if recent_payments and recent_payments > 0:
        notifications.append({
            "type": "payment",
            "severity": "success",
            "title": f"–ù–æ–≤—ã—Ö –ø–ª–∞—Ç–µ–∂–µ–π: {recent_payments}",
            "message": f"–ó–∞ –ø–æ—Å–ª–µ–¥–Ω–∏–π —á–∞—Å –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ {recent_payments} —É—Å–ø–µ—à–Ω—ã—Ö –ø–ª–∞—Ç–µ–∂–µ–π",
            "link": "/admin/web/logs?action=payment_processed",
            "count": recent_payments,
        })
    
    return {
        "notifications": notifications,
        "alerts": alerts,
        "unread_count": len(notifications) + len(alerts),
    }


@app.get("/admin/web/users", response_class=HTMLResponse)
async def admin_web_users(
    request: Request,
    q: str | None = Query(default=None),
    status: str | None = Query(default="all"),
    role: str | None = Query(default="all"),
    reg_period: str | None = Query(default="all"),
    balance_min: str | None = Query(default=None),
    balance_max: str | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–í–µ–±-–∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å –∞–¥–º–∏–Ω–∫–∏ –¥–ª—è –ø—Ä–æ—Å–º–æ—Ç—Ä–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    # –í–∞–ª–∏–¥–∏—Ä—É–µ–º —Ñ–∏–ª—å—Ç—Ä —Å—Ç–∞—Ç—É—Å–∞
    status_filter = (status or "all").lower()
    if status_filter not in {"all", "active", "blocked"}:
        status_filter = "all"

    # –í–∞–ª–∏–¥–∏—Ä—É–µ–º —Ñ–∏–ª—å—Ç—Ä —Ä–æ–ª–∏
    role_filter = (role or "all").lower()
    if role_filter not in {"all", "superadmin", "admin", "moderator", "user"}:
        role_filter = "all"

    # –í–∞–ª–∏–¥–∏—Ä—É–µ–º –ø–µ—Ä–∏–æ–¥ —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏
    reg_filter = (reg_period or "all").lower()
    if reg_filter not in {"all", "today", "7d", "30d"}:
        reg_filter = "all"

    # –°—á—ë—Ç—á–∏–∫–∏
    total_users = await session.scalar(select(func.count()).select_from(User))
    total_active = await session.scalar(select(func.count()).select_from(User).where(User.is_active == True))
    total_blocked = await session.scalar(select(func.count()).select_from(User).where(User.is_active == False))

    # –ü–æ–ª—É—á–∞–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π (—Å –ø–æ–∏—Å–∫–æ–º/—Ñ–∏–ª—å—Ç—Ä–∞–º–∏)
    stmt = select(User).options(selectinload(User.referred_by)).order_by(User.created_at.desc())
    if status_filter == "active":
        stmt = stmt.where(User.is_active == True)
    elif status_filter == "blocked":
        stmt = stmt.where(User.is_active == False)

    # –§–∏–ª—å—Ç—Ä –ø–æ –ø–µ—Ä–∏–æ–¥—É —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏ (–ø–æ created_at, –≤ UTC)
    if reg_filter != "all":
        now_utc = datetime.utcnow()
        if reg_filter == "today":
            start = now_utc.replace(hour=0, minute=0, second=0, microsecond=0)
        elif reg_filter == "7d":
            start = now_utc - timedelta(days=7)
        else:  # "30d"
            start = now_utc - timedelta(days=30)
        stmt = stmt.where(User.created_at >= start)

    if q:
        qq = q.strip()
        if qq.isdigit():
            stmt = stmt.where(User.tg_id == int(qq))
        else:
            # –ø–æ–∏—Å–∫ –ø–æ username/first/last
            like = f"%{qq.lower()}%"
            stmt = stmt.where(
                func.lower(User.username).like(like)
                | func.lower(User.first_name).like(like)
                | func.lower(User.last_name).like(like)
            )
    
    # –§–∏–ª—å—Ç—Ä –ø–æ –±–∞–ª–∞–Ω—Å—É
    if balance_min:
        try:
            min_balance = int(float(balance_min) * 100)  # –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º USD –≤ —Ü–µ–Ω—Ç—ã
            stmt = stmt.where(User.balance >= min_balance)
        except (ValueError, TypeError):
            pass
    if balance_max:
        try:
            max_balance = int(float(balance_max) * 100)  # –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º RUB –≤ –∫–æ–ø–µ–π–∫–∏
            stmt = stmt.where(User.balance <= max_balance)
        except (ValueError, TypeError):
            pass
    
    result = await session.scalars(stmt)
    users: Sequence[User] = result.all()

    # –ü–æ–¥—Ç—è–≥–∏–≤–∞–µ–º overrides —Ä–æ–ª–µ–π
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
    
    # –ü–æ–ª—É—á–∞–µ–º admin_ids –∏–∑ –ø–µ—Ä–µ–º–µ–Ω–Ω–æ–π –æ–∫—Ä—É–∂–µ–Ω–∏—è
    import os
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()
    
    # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è —à–∞–±–ª–æ–Ω–∞ –∏ —Ñ–∏–ª—å—Ç—Ä—É–µ–º –ø–æ —Ä–æ–ª—è–º
    users_data = []
    for u in users:
        referred_by_tg_id = u.referred_by.tg_id if u.referred_by else None
        full_name = " ".join(filter(None, [u.first_name or "", u.last_name or ""])) or "‚Äî"
        tag = f"@{u.username}" if u.username else "‚Äî"
        eff_role = _get_effective_role(u.tg_id, admin_ids, overrides_map)
        
        # –ü—Ä–∏–º–µ–Ω—è–µ–º —Ñ–∏–ª—å—Ç—Ä –ø–æ —Ä–æ–ª–∏ (–µ—Å–ª–∏ –∑–∞–¥–∞–Ω)
        if role_filter != "all" and eff_role != role_filter:
            continue
        
        role = "–ì–ª–∞–≤–Ω—ã–π –∞–¥–º–∏–Ω" if eff_role == "superadmin" else ("–ê–¥–º–∏–Ω" if eff_role == "admin" else ("–ú–æ–¥–µ—Ä–∞—Ç–æ—Ä" if eff_role == "moderator" else "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å"))
        balance_rub = u.balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –≤—Ä–µ–º—è –≤ –ú–°–ö
        try:
            from zoneinfo import ZoneInfo
            moscow_tz = ZoneInfo("Europe/Moscow")
            dt_moscow = u.created_at.astimezone(moscow_tz)
            created_str = dt_moscow.strftime("%d.%m.%Y %H:%M")
        except:
            created_str = str(u.created_at)[:16]
        
        users_data.append({
            "id": u.id,
            "tg_id": u.tg_id,
            "username": u.username or "‚Äî",
            "full_name": full_name,
            "tag": tag,
            "role": role,
            "is_active": u.is_active,
            "has_active_subscription": u.has_active_subscription,
            "balance": balance_rub,
            "referral_code": u.referral_code or "‚Äî",
            "referred_by_tg_id": referred_by_tg_id,
            "created_at": created_str,
        })
    
    return templates.TemplateResponse("admin.html", {
        "request": request,
        "users": users_data,
        "total_users": int(total_users or 0),
        "total_active": int(total_active or 0),
        "total_blocked": int(total_blocked or 0),
        "q": q or "",
        "status": status_filter,
        "role_filter": role_filter,
        "reg_filter": reg_filter,
        "balance_min": balance_min or "",
        "balance_max": balance_max or "",
        "admin_user": admin_user,
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/web/users/{tg_id}", response_class=HTMLResponse)
async def admin_web_user_detail(
    tg_id: int,
    request: Request,
    logs_page: int = Query(default=1, ge=1),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)

    stmt = select(User).options(selectinload(User.referred_by)).where(User.tg_id == tg_id)
    user = await session.scalar(stmt)
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
    override_role = overrides_map.get(tg_id)
    eff_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
    role = "–ì–ª–∞–≤–Ω—ã–π –∞–¥–º–∏–Ω" if eff_role == "superadmin" else ("–ê–¥–º–∏–Ω" if eff_role == "admin" else ("–ú–æ–¥–µ—Ä–∞—Ç–æ—Ä" if eff_role == "moderator" else "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å"))
    balance_rub = user.balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)

    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        created_str = user.created_at.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M")
    except Exception:
        created_str = str(user.created_at)[:16]

    referrals_count = await session.scalar(select(func.count()).select_from(User).where(User.referred_by_user_id == user.id))

    # --- –õ–æ–≥–∏ —Å –ø–∞–≥–∏–Ω–∞—Ü–∏–µ–π
    logs_page_size = 20
    logs_stmt = (
        select(AuditLog)
        .where(AuditLog.user_tg_id == tg_id)
        .order_by(AuditLog.created_at.desc())
    )
    total_logs = await session.scalar(select(func.count()).select_from(logs_stmt.subquery()))
    logs_result = await session.scalars(
        logs_stmt
        .limit(logs_page_size)
        .offset((logs_page - 1) * logs_page_size)
    )
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    logs = [
        {
            "id": log.id,
            "action": log.action.value if hasattr(log.action, "value") else str(log.action),
            "admin_tg_id": log.admin_tg_id,
            "details": log.details,
            "created_at": fmt(log.created_at),
        }
        for log in logs_result.all()
    ]
    logs_has_next = (logs_page * logs_page_size) < (total_logs or 0)
    logs_has_prev = logs_page > 1

    # --- –í—Å–µ —Ç–∏–∫–µ—Ç—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    tickets_result = await session.scalars(
        select(Ticket)
        .where(Ticket.user_tg_id == tg_id)
        .order_by(Ticket.updated_at.desc())
    )
    tickets_list = tickets_result.all()
    
    tickets_data = []
    for t in tickets_list:
        tickets_data.append({
            "id": t.id,
            "topic": t.topic or "‚Äî",
            "status": t.status.value if hasattr(t.status, "value") else str(t.status),
            "created_at": fmt(t.created_at),
            "updated_at": fmt(t.updated_at),
        })

    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –≤—ã–±—Ä–∞–Ω–Ω–æ–º —Å–µ—Ä–≤–µ—Ä–µ
    selected_server_name = None
    if user.selected_server_id:
        selected_server = await session.get(Server, user.selected_server_id)
        if selected_server:
            selected_server_name = selected_server.name
    
    user_data = {
        "id": user.id,
        "tg_id": user.tg_id,
        "username": user.username,
        "first_name": user.first_name,
        "last_name": user.last_name,
        "role": role,
        "is_admin_env": user.tg_id in admin_ids,
        "role_override": override_role,
        "is_active": user.is_active,
        "has_active_subscription": user.has_active_subscription,
        "subscription_ends_at": fmt(user.subscription_ends_at) if user.subscription_ends_at else "‚Äî",
        "selected_server_name": selected_server_name or "‚Äî",
        "balance": balance_rub,
        "referral_code": user.referral_code or "‚Äî",
        "referred_by_tg_id": user.referred_by.tg_id if user.referred_by else None,
        "created_at": created_str,
    }

    photo_url = await _fetch_avatar_url(tg_id, os.getenv("BOT_TOKEN", ""))
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –±–∞–Ω–∞—Ö
    active_ban = await session.scalar(
        select(UserBan)
        .where(UserBan.user_id == user.id)
        .where(UserBan.is_active == True)
        .order_by(UserBan.banned_at.desc())
    )
    
    ban_info = None
    if active_ban:
        ban_info = {
            "id": active_ban.id,
            "reason": active_ban.reason,
            "details": active_ban.details,
            "banned_at": fmt(active_ban.banned_at),
            "banned_until": fmt(active_ban.banned_until) if active_ban.banned_until else "–ü–µ—Ä–º–∞–Ω–µ–Ω—Ç–Ω–æ",
            "auto_ban": active_ban.auto_ban,
        }
    
    # –ü–æ–ª—É—á–∞–µ–º –∏—Å—Ç–æ—Ä–∏—é IP –∞–¥—Ä–µ—Å–æ–≤
    ip_logs = await session.scalars(
        select(IpLog)
        .where(IpLog.user_id == user.id)
        .order_by(IpLog.last_seen.desc())
        .limit(20)
    )
    ip_history = []
    for ip_log in ip_logs.all():
        server = await session.get(Server, ip_log.server_id)
        ip_history.append({
            "ip": ip_log.ip_address,
            "server": server.name if server else "‚Äî",
            "first_seen": fmt(ip_log.first_seen),
            "last_seen": fmt(ip_log.last_seen),
            "count": ip_log.connection_count,
        })

    return templates.TemplateResponse("user_detail.html", {
        "request": request,
        "user": user_data,
        "referrals_count": int(referrals_count or 0),
        "logs": logs,
        "logs_page": logs_page,
        "logs_has_next": logs_has_next,
        "logs_has_prev": logs_has_prev,
        "tickets": tickets_data,
        "admin_user": admin_user,
        "csrf_token": _get_csrf_token(request),
        "can_toggle_role": (user.tg_id not in admin_ids),  # –Ω–µ–ª—å–∑—è –º–µ–Ω—è—Ç—å —Ä–æ–ª—å –¥–ª—è env-–∞–¥–º–∏–Ω–∞ (–≥–ª–∞–≤–Ω—ã–π)
        "is_protected_admin": (user.tg_id in admin_ids),
        "role_value": ("superadmin" if user.tg_id in admin_ids else (override_role if override_role else "user")),
        "role_options": ["admin", "moderator", "user"],
        "avatar_url": photo_url,
        "ban_info": ban_info,
        "ip_history": ip_history,
    })


@app.get("/admin/web/api/users/{tg_id}/logs")
async def admin_api_user_logs(
    tg_id: int,
    page: int = Query(default=1, ge=1),
    action: str = Query(default="all"),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –ª–æ–≥–æ–≤ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è —Å –ø–∞–≥–∏–Ω–∞—Ü–∏–µ–π."""
    logs_page_size = 20
    logs_stmt = (
        select(AuditLog)
        .where(AuditLog.user_tg_id == tg_id)
        .order_by(AuditLog.created_at.desc())
    )
    
    if action != "all":
        try:
            action_enum = AuditLogAction(action)
            logs_stmt = logs_stmt.where(AuditLog.action == action_enum)
        except ValueError:
            pass  # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –Ω–µ–≤–µ—Ä–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
    total_logs = await session.scalar(select(func.count()).select_from(logs_stmt.subquery()))
    logs_result = await session.scalars(
        logs_stmt
        .limit(logs_page_size)
        .offset((page - 1) * logs_page_size)
    )
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    logs = [
        {
            "id": log.id,
            "action": log.action.value if hasattr(log.action, "value") else str(log.action),
            "admin_tg_id": log.admin_tg_id,
            "details": log.details,
            "created_at": fmt(log.created_at),
        }
        for log in logs_result.all()
    ]
    logs_has_next = (page * logs_page_size) < (total_logs or 0)
    logs_has_prev = page > 1
    
    return {
        "logs": logs,
        "page": page,
        "has_next": logs_has_next,
        "has_prev": logs_has_prev,
        "total": total_logs or 0,
    }


@app.get("/admin/web/payments", response_class=HTMLResponse)
async def admin_web_payments(
    request: Request,
    page: int = Query(default=1, ge=1),
    status: str | None = Query(default=None),
    provider: str | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ –ø–ª–∞—Ç–µ–∂–µ–π –≤ –∞–¥–º–∏–Ω–∫–µ"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    page_size = 50
    offset = (page - 1) * page_size
    
    result = await admin_list_payments(
        limit=page_size,
        offset=offset,
        status=status,
        provider=provider,
        session=session,
        admin_user=admin_user,
    )
    
    payments = result.get("payments", [])
    total = result.get("total", 0)
    has_next = (page * page_size) < total
    has_prev = page > 1
    
    return templates.TemplateResponse("payments.html", {
        "request": request,
        "admin_user": admin_user,
        "payments": payments,
        "page": page,
        "has_next": has_next,
        "has_prev": has_prev,
        "total": total,
        "status_filter": status or "all",
        "provider_filter": provider or "all",
        "csrf_token": _get_csrf_token(request),
    })


@app.get("/admin/web/user/{tg_id}", response_class=HTMLResponse)
async def admin_web_user(
    tg_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)

    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()

    stmt = (
        select(User)
        .options(
            selectinload(User.referred_by),
            selectinload(User.referrals),
        )
        .where(User.tg_id == tg_id)
    )
    user = await session.scalar(stmt)
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    referred_by_tg_id = user.referred_by.tg_id if user.referred_by else None
    referrals_count = len(user.referrals or [])
    role = "–ê–¥–º–∏–Ω" if user.tg_id in admin_ids else "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å"
    balance_rub = user.balance / 100  # –ë–∞–ª–∞–Ω—Å —É–∂–µ –≤ —Ä—É–±–ª—è—Ö (–∫–æ–ø–µ–π–∫–∞—Ö)

    # –ü–æ–¥–ø–∏—Å–∫–∞
    sub = await session.scalar(
        select(Subscription)
        .where(Subscription.user_id == user.id, Subscription.status == SubscriptionStatus.active)
        .order_by(Subscription.ends_at.desc().nullslast())
    )
    sub_info = None
    if sub:
        try:
            from zoneinfo import ZoneInfo
            moscow_tz = ZoneInfo("Europe/Moscow")
            ends_msk = sub.ends_at.astimezone(moscow_tz) if sub.ends_at else None
            ends_str = ends_msk.strftime("%d.%m.%Y %H:%M") if ends_msk else "‚Äî"
        except Exception:
            ends_str = sub.ends_at.isoformat() if sub and sub.ends_at else "‚Äî"
        sub_info = {
            "plan": sub.plan_name or "‚Äî",
            "ends_at": ends_str,
            "status": sub.status.value if hasattr(sub.status, "value") else str(sub.status),
        }

    # –í—Ä–µ–º—è —Ä–µ–≥–∏—Å—Ç—Ä–∞—Ü–∏–∏ –≤ –ú–°–ö
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        created_msk = user.created_at.astimezone(moscow_tz)
        created_str = created_msk.strftime("%d.%m.%Y %H:%M")
    except Exception:
        created_str = str(user.created_at)[:16]

    user_data = {
        "id": user.id,
        "tg_id": user.tg_id,
        "username": user.username or "‚Äî",
        "full_name": " ".join(filter(None, [user.first_name or "", user.last_name or ""])) or "‚Äî",
        "tag": f"@{user.username}" if user.username else "‚Äî",
        "role": role,
        "is_active": user.is_active,
        "balance": balance_rub,
        "referral_code": user.referral_code or "‚Äî",
        "referred_by_tg_id": referred_by_tg_id,
        "referrals_count": referrals_count,
        "created_at": created_str,
    }

    # --- –í—Å–µ —Ç–∏–∫–µ—Ç—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    tickets_result = await session.scalars(
        select(Ticket)
        .where(Ticket.user_tg_id == tg_id)
        .order_by(Ticket.updated_at.desc())
    )
    tickets_list = tickets_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    tickets_data = []
    for t in tickets_list:
        tickets_data.append({
            "id": t.id,
            "topic": t.topic or "‚Äî",
            "status": t.status.value if hasattr(t.status, "value") else str(t.status),
            "created_at": fmt(t.created_at),
            "updated_at": fmt(t.updated_at),
        })

    return templates.TemplateResponse(
        "user_detail.html",
        {
            "request": request,
            "user": user_data,
            "sub": sub_info,
            "csrf_token": _get_csrf_token(request),
            "admin_user": admin_user,
            "tickets": tickets_data,
            "open_modal": request.query_params.get("open"),
        },
    )


@app.post("/admin/web/users/block")
async def admin_web_block_user(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    ticket_id = form.get("ticket_id")
    reason = str(form.get("reason", "")).strip()
    if not reason:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=reason_required", status_code=303)
    admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}

    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    actor_tg = admin_user.get("tg_id")
    actor_role = _get_effective_role(actor_tg, admin_ids, overrides_map)
    target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
    # –°–∞–º–æ–≥–æ —Å–µ–±—è –±–ª–æ–∫–∏—Ä–æ–≤–∞—Ç—å –Ω–µ–ª—å–∑—è
    if actor_tg == user.tg_id:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_self", status_code=303)
    # –ò–µ—Ä–∞—Ä—Ö–∏—è –¥–ª—è —á—É–∂–∏—Ö
    if _role_rank(actor_role) <= _role_rank(target_role):
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_rank", status_code=303)
    user.is_active = False
    session.add(
        AuditLog(
            action=AuditLogAction.user_blocked,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
        )
    )
    await session.commit()
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    notification_text = (
        f"‚ùå <b>–ê–∫–∫–∞—É–Ω—Ç –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
        f"–í–∞—à –∞–∫–∫–∞—É–Ω—Ç –±—ã–ª –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º."
    )
    if reason:
        notification_text += f"\n\n–ü—Ä–∏—á–∏–Ω–∞: {reason}"
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/support/webhook")
async def support_webhook(
    update: dict,
    session: AsyncSession = Depends(get_session),
    x_admin_token: str | None = Header(default=None),
):
    # –£–±–∏—Ä–∞–µ–º –æ–±—è–∑–∞—Ç–µ–ª—å–Ω—É—é –∞–≤—Ç–æ—Ä–∏–∑–∞—Ü–∏—é –∑–∞–≥–æ–ª–æ–≤–∫–æ–º ‚Äî Telegram –Ω–µ —à–ª—ë—Ç –µ–≥–æ –Ω–∞ –≤–µ–±—Ö—É–∫–µ
    settings = get_settings()

    # –ü–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ–º message –∏ edited_message
    message = update.get("message") or update.get("edited_message") or {}
    chat = message.get("chat") or {}
    text = message.get("text") or message.get("caption") or ""
    tg_id = chat.get("id")
    if not tg_id or not text:
        return {"status": "ignored"}

    # –°–æ–∑–¥–∞—ë–º/–æ–±–Ω–æ–≤–ª—è–µ–º —Ç–∏–∫–µ—Ç
    ticket = await session.scalar(
        select(Ticket).where(Ticket.user_tg_id == tg_id).order_by(Ticket.updated_at.desc())
    )
    now = datetime.utcnow()
    if not ticket:
        ticket = Ticket(user_tg_id=tg_id, status=TicketStatus.open, created_at=now, updated_at=now)
        session.add(ticket)
        await session.flush()
    else:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∑–∞–∫—Ä—ã—Ç –ª–∏ —Ç–∏–∫–µ—Ç
        if ticket.status == TicketStatus.closed:
            # –¢–∏–∫–µ—Ç –∑–∞–∫—Ä—ã—Ç, –Ω–µ –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ
            return {"status": "ignored", "reason": "ticket_closed"}
        ticket.updated_at = now

    session.add(
        TicketMessage(
            ticket_id=ticket.id,
            user_tg_id=tg_id,
            direction=MessageDirection.incoming,
            admin_tg_id=None,
            text=text,
        )
    )
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            details=f"User message to support: {text}",
        )
    )
    await session.commit()
    return {"status": "ok"}


@app.post("/admin/web/users/unblock")
async def admin_web_unblock_user(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    reason = str(form.get("reason", "")).strip()
    if not reason:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=reason_required", status_code=303)
    admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}

    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    actor_role = _get_effective_role(admin_user.get("tg_id"), admin_ids, overrides_map)
    target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
    if admin_user.get("tg_id") == user.tg_id:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_self", status_code=303)
    if _role_rank(actor_role) <= _role_rank(target_role):
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_rank", status_code=303)
    user.is_active = True
    session.add(
        AuditLog(
            action=AuditLogAction.user_unblocked,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
        )
    )
    await session.commit()
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    notification_text = (
        f"‚úÖ <b>–ê–∫–∫–∞—É–Ω—Ç —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
        f"–í–∞—à –∞–∫–∫–∞—É–Ω—Ç –±—ã–ª —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º."
    )
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/admin/web/users/vpn-ban")
async def admin_web_vpn_ban_user(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ó–∞–±–∞–Ω–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ VPN (–æ—Ç–∫–ª—é—á–∏—Ç—å –∫–ª–∏–µ–Ω—Ç–∞ –≤ 3x-UI)"""
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    reason = str(form.get("reason", "")).strip()
    duration_hours = int(str(form.get("duration_hours", "24")))
    
    if not reason:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=reason_required", status_code=303)
    
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ —É–∂–µ –∞–∫—Ç–∏–≤–Ω–æ–≥–æ –±–∞–Ω–∞
    existing_ban = await session.scalar(
        select(UserBan).where(UserBan.user_id == user.id, UserBan.is_active == True)
    )
    
    if existing_ban:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=already_banned", status_code=303)
    
    now = datetime.utcnow()
    banned_until = now + timedelta(hours=duration_hours) if duration_hours > 0 else None
    
    # –°–æ–∑–¥–∞–µ–º –±–∞–Ω
    ban = UserBan(
        user_id=user.id,
        reason="manual",
        details=reason,
        is_active=True,
        auto_ban=False,
        banned_until=banned_until,
    )
    session.add(ban)
    
    # –û—Ç–∫–ª—é—á–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –≤ 3x-UI
    credentials = await session.scalars(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.active == True)
        .options(selectinload(VpnCredential.server))
    )
    
    for cred in credentials.all():
        if not cred.server or not cred.user_uuid:
            continue
        
        server = cred.server
        if server.x3ui_api_url and server.x3ui_username and server.x3ui_password and server.x3ui_inbound_id:
            try:
                from core.x3ui_api import X3UIAPI
                x3ui = X3UIAPI(
                    api_url=server.x3ui_api_url,
                    username=server.x3ui_username,
                    password=server.x3ui_password,
                )
                try:
                    await x3ui.disable_client(server.x3ui_inbound_id, cred.user_uuid)
                finally:
                    await x3ui.close()
            except Exception as e:
                logger.error(f"Error disabling client for ban: {e}")
    
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"VPN –±–∞–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è. –ü—Ä–∏—á–∏–Ω–∞: {reason}. –°—Ä–æ–∫: {duration_hours}—á" if duration_hours > 0 else f"VPN –±–∞–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è (–ø–µ—Ä–º–∞–Ω–µ–Ω—Ç–Ω—ã–π). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
        )
    )
    await session.commit()
    
    # –£–≤–µ–¥–æ–º–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    duration_text = f"–Ω–∞ {duration_hours} —á–∞—Å–æ–≤" if duration_hours > 0 else "–Ω–∞ –Ω–µ–æ–ø—Ä–µ–¥–µ–ª–µ–Ω–Ω—ã–π —Å—Ä–æ–∫"
    notification_text = (
        f"‚õî <b>–í–∞—à VPN –¥–æ—Å—Ç—É–ø –∑–∞–±–ª–æ–∫–∏—Ä–æ–≤–∞–Ω</b>\n\n"
        f"–ü—Ä–∏—á–∏–Ω–∞: {reason}\n"
        f"–°—Ä–æ–∫: {duration_text}\n\n"
        "–ï—Å–ª–∏ –≤—ã —Å—á–∏—Ç–∞–µ—Ç–µ —ç—Ç–æ –æ—à–∏–±–∫–æ–π, –æ–±—Ä–∞—Ç–∏—Ç–µ—Å—å –≤ –ø–æ–¥–¥–µ—Ä–∂–∫—É."
    )
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/admin/web/users/vpn-unban")
async def admin_web_vpn_unban_user(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–†–∞–∑–±–∞–Ω–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ VPN"""
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)
    
    # –ù–∞—Ö–æ–¥–∏–º –∞–∫—Ç–∏–≤–Ω—ã–π –±–∞–Ω
    active_ban = await session.scalar(
        select(UserBan).where(UserBan.user_id == user.id, UserBan.is_active == True)
    )
    
    if not active_ban:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=not_banned", status_code=303)
    
    # –°–Ω–∏–º–∞–µ–º –±–∞–Ω
    active_ban.is_active = False
    active_ban.unbanned_at = datetime.utcnow()
    active_ban.unbanned_by_tg_id = admin_user.get("tg_id")
    
    # –í–∫–ª—é—á–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –æ–±—Ä–∞—Ç–Ω–æ –≤ 3x-UI
    credentials = await session.scalars(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.active == True)
        .options(selectinload(VpnCredential.server))
    )
    
    for cred in credentials.all():
        if not cred.server or not cred.user_uuid:
            continue
        
        server = cred.server
        if server.x3ui_api_url and server.x3ui_username and server.x3ui_password and server.x3ui_inbound_id:
            try:
                from core.x3ui_api import X3UIAPI
                x3ui = X3UIAPI(
                    api_url=server.x3ui_api_url,
                    username=server.x3ui_username,
                    password=server.x3ui_password,
                )
                try:
                    await x3ui.enable_client(server.x3ui_inbound_id, cred.user_uuid)
                finally:
                    await x3ui.close()
            except Exception as e:
                logger.error(f"Error enabling client for unban: {e}")
    
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"VPN —Ä–∞–∑–±–∞–Ω –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è",
        )
    )
    await session.commit()
    
    # –£–≤–µ–¥–æ–º–ª—è–µ–º –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    notification_text = (
        f"‚úÖ <b>–í–∞—à VPN –¥–æ—Å—Ç—É–ø –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω</b>\n\n"
        "–í—ã —Å–Ω–æ–≤–∞ –º–æ–∂–µ—Ç–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç—å—Å—è VPN."
    )
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/admin/web/users/credit")
async def admin_web_credit_user(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    amount_str = str(form.get("amount", "0")).replace(",", ".").strip()
    reason = str(form.get("reason", "")).strip()
    if not reason:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=reason_required", status_code=303)
    try:
        amount_rub = float(amount_str)
    except Exception:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=bad_amount", status_code=303)
    if amount_rub == 0:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=bad_amount", status_code=303)
    amount_cents = int(round(amount_rub * 100))

    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    # –ü—Ä–æ–≤–µ—Ä–∫–∞ –∏–µ—Ä–∞—Ä—Ö–∏–∏ —Ä–æ–ª–µ–π: –Ω–µ–ª—å–∑—è –º–µ–Ω—è—Ç—å –±–∞–ª–∞–Ω—Å —Ä–∞–≤–Ω—ã–º/—Å—Ç–∞—Ä—à–∏–º
    admin_ids_str = os.getenv("ADMIN_IDS", "")
    admin_ids = set(int(x.strip()) for x in admin_ids_str.split(",") if x.strip().isdigit()) if admin_ids_str else set()
    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
    actor_tg = admin_user.get("tg_id")
    actor_role = _get_effective_role(actor_tg, admin_ids, overrides_map)
    target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
    # –°–∞–º–æ–º—É —Å–µ–±–µ –≤—Å–µ–≥–¥–∞ –º–æ–∂–Ω–æ; –¥–ª—è —á—É–∂–∏—Ö ‚Äî –ø—Ä–æ–≤–µ—Ä–∫–∞ —Ä–∞–Ω–≥–∞
    if actor_tg != user.tg_id and _role_rank(actor_role) <= _role_rank(target_role):
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_rank", status_code=303)
    # –ó–∞–ø—Ä–µ—Ç –º–µ–Ω—è—Ç—å –±–∞–ª–∞–Ω—Å –≥–ª–∞–≤–Ω–æ–º—É –∞–¥–º–∏–Ω—É –¥—Ä—É–≥–∏–º–∏
    admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
    if tg_id in admin_ids and actor_tg != user.tg_id:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=role_protected", status_code=303)

    old_balance = user.balance
    user.balance += amount_cents
    session.add(
        BalanceTransaction(
            user_id=user.id,
            admin_tg_id=admin_user.get("tg_id"),
            amount=amount_cents,
            reason=reason,
        )
    )
    session.add(
        AuditLog(
            action=AuditLogAction.balance_credited,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–ë–∞–ª–∞–Ω—Å –∏–∑–º–µ–Ω–µ–Ω (web): {old_balance} -> {user.balance} —Ü–µ–Ω—Ç–æ–≤. –ü—Ä–∏—á–∏–Ω–∞: {reason}",
        )
    )
    await session.commit()
    
    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
    amount_rub = amount_cents / 100
    if amount_rub > 0:
        notification_text = (
            f"üí∞ <b>–ë–∞–ª–∞–Ω—Å –ø–æ–ø–æ–ª–Ω–µ–Ω</b>\n\n"
            f"–°—É–º–º–∞: <b>+{amount_rub:.2f} RUB</b>\n"
            f"–¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>"
        )
        if reason:
            notification_text += f"\n\n–ü—Ä–∏—á–∏–Ω–∞: {reason}"
    else:
        notification_text = (
            f"üí∞ <b>–ò–∑–º–µ–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞</b>\n\n"
            f"–°—É–º–º–∞: <b>{amount_rub:.2f} RUB</b>\n"
            f"–¢–µ–∫—É—â–∏–π –±–∞–ª–∞–Ω—Å: <b>{user.balance / 100:.2f} RUB</b>"
        )
        if reason:
            notification_text += f"\n\n–ü—Ä–∏—á–∏–Ω–∞: {reason}"
    
    asyncio.create_task(_send_user_notification(tg_id, notification_text))
    
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/admin/web/users/subscription")
async def admin_web_manage_subscription(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
) -> JSONResponse:
    """–£–ø—Ä–∞–≤–ª–µ–Ω–∏–µ –ø–æ–¥–ø–∏—Å–∫–æ–π –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è —á–µ—Ä–µ–∑ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å"""
    from datetime import timedelta, timezone
    
    await _require_csrf(request)
    try:
        data = await request.json()
        tg_id = data.get("tg_id")
        action = data.get("action")  # "add", "extend", "remove"
        days = data.get("days")
        reason = data.get("reason", "").strip()
        
        if not tg_id:
            return JSONResponse({"success": False, "error": "–ù–µ —É–∫–∞–∑–∞–Ω tg_id"})
        if not action or action not in ["add", "extend", "remove"]:
            return JSONResponse({"success": False, "error": "–ù–µ–≤–µ—Ä–Ω–æ–µ –¥–µ–π—Å—Ç–≤–∏–µ"})
        if action != "remove" and (not days or days < 1 or days > 9999):
            return JSONResponse({"success": False, "error": "–ù–µ–≤–µ—Ä–Ω–æ–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –¥–Ω–µ–π (–æ—Ç 1 –¥–æ 9999 –¥–Ω–µ–π)"})
        if not reason:
            return JSONResponse({"success": False, "error": "–ù—É–∂–Ω–æ —É–∫–∞–∑–∞—Ç—å –ø—Ä–∏—á–∏–Ω—É"})
        
        user = await session.scalar(select(User).where(User.tg_id == tg_id))
        if not user:
            return JSONResponse({"success": False, "error": "–ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –Ω–µ –Ω–∞–π–¥–µ–Ω"})
        
        actor_tg = admin_user.get("tg_id")
        now = datetime.now(timezone.utc)
        
        if action == "remove":
            # –û—Ç–º–µ–Ω—è–µ–º –≤—Å–µ –∞–∫—Ç–∏–≤–Ω—ã–µ –ø–æ–¥–ø–∏—Å–∫–∏
            active_subs = await session.scalars(
                select(Subscription)
                .where(Subscription.user_id == user.id)
                .where(Subscription.status == SubscriptionStatus.active)
            )
            canceled_count = 0
            for sub in active_subs.all():
                sub.status = SubscriptionStatus.canceled
                canceled_count += 1
            
            # –£–¥–∞–ª—è–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –∏–∑ 3x-UI –Ω–∞ –≤—Å–µ—Ö —Å–µ—Ä–≤–µ—Ä–∞—Ö
            credentials = await session.scalars(
                select(VpnCredential)
                .where(VpnCredential.user_id == user.id)
                .where(VpnCredential.active == True)
                .options(selectinload(VpnCredential.server))
            )
            
            deleted_clients = 0
            for cred in credentials.all():
                if not cred.server or not cred.user_uuid:
                    continue
                
                server = cred.server
                if server.x3ui_api_url and server.x3ui_username and server.x3ui_password:
                    try:
                        from core.x3ui_api import X3UIAPI
                        x3ui = X3UIAPI(
                            api_url=server.x3ui_api_url,
                            username=server.x3ui_username,
                            password=server.x3ui_password,
                        )
                        try:
                            client_email = f"tg_{user.tg_id}_server_{server.id}@fiorevpn"
                            inbound_id = server.x3ui_inbound_id
                            if inbound_id:
                                deleted = await x3ui.delete_client(inbound_id, client_email)
                                if deleted:
                                    logger.info(f"–£–¥–∞–ª–µ–Ω –∫–ª–∏–µ–Ω—Ç {client_email} –∏–∑ 3x-UI –ø—Ä–∏ –æ—Ç–º–µ–Ω–µ –ø–æ–¥–ø–∏—Å–∫–∏")
                                    deleted_clients += 1
                                # –ü–æ–º–µ—á–∞–µ–º credential –∫–∞–∫ –Ω–µ–∞–∫—Ç–∏–≤–Ω—ã–π
                                cred.active = False
                        finally:
                            await x3ui.close()
                    except Exception as e:
                        logger.error(f"Error deleting client from 3x-UI when canceling subscription: {e}")
            
            # –õ–æ–≥–∏—Ä—É–µ–º
            session.add(
                AuditLog(
                    action=AuditLogAction.admin_action,
                    user_tg_id=tg_id,
                    admin_tg_id=actor_tg,
                    details=f"–û—Ç–º–µ–Ω–µ–Ω–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º. –ü—Ä–∏—á–∏–Ω–∞: {reason}. –£–¥–∞–ª–µ–Ω–æ –∫–ª–∏–µ–Ω—Ç–æ–≤ –∏–∑ 3x-UI: {deleted_clients}",
                )
            )
            
            await session.commit()
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ—Å–ª–µ –∫–æ–º–º–∏—Ç–∞ –æ—Ç–º–µ–Ω—ã –ø–æ–¥–ø–∏—Å–∫–∏
            await _update_user_subscription_status(user.id, session)
            await session.commit()
            await session.refresh(user)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–µ—Å—Å–∏–∏
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é —Å –æ–±–Ω–æ–≤–ª–µ–Ω–Ω—ã–º –º–µ–Ω—é
            notification_text = (
                f"üìã <b>–ü–æ–¥–ø–∏—Å–∫–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞</b>\n\n"
                f"–í–∞—à–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –±—ã–ª–∞ –æ—Ç–º–µ–Ω–µ–Ω–∞ –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º.\n\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: {reason}"
            )
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –∏ –æ–±–Ω–æ–≤–ª—è–µ–º –º–µ–Ω—é –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
            asyncio.create_task(_send_user_notification_with_menu_update(tg_id, notification_text))
            
            return JSONResponse({"success": True, "message": f"–û—Ç–º–µ–Ω–µ–Ω–æ –ø–æ–¥–ø–∏—Å–æ–∫: {canceled_count}, —É–¥–∞–ª–µ–Ω–æ –∫–ª–∏–µ–Ω—Ç–æ–≤ –∏–∑ 3x-UI: {deleted_clients}"})
        
        elif action == "add":
            # –í—ã–¥–∞–µ–º –Ω–æ–≤—É—é –ø–æ–¥–ø–∏—Å–∫—É
            starts_at = now
            ends_at = now + timedelta(days=days)
            
            subscription = Subscription(
                user_id=user.id,
                plan_name=f"–ü–æ–¥–ø–∏—Å–∫–∞ –Ω–∞ {days} –¥–Ω–µ–π (–≤—ã–¥–∞–Ω–∞ –∞–¥–º–∏–Ω–æ–º)",
                price_cents=0,  # –ë–µ—Å–ø–ª–∞—Ç–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞ –æ—Ç –∞–¥–º–∏–Ω–∞
                currency="RUB",
                status=SubscriptionStatus.active,
                starts_at=starts_at,
                ends_at=ends_at,
            )
            session.add(subscription)
            
            # –õ–æ–≥–∏—Ä—É–µ–º
            session.add(
                AuditLog(
                    action=AuditLogAction.subscription_created,
                    user_tg_id=tg_id,
                    admin_tg_id=actor_tg,
                    details=f"–í—ã–¥–∞–Ω–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –Ω–∞ {days} –¥–Ω–µ–π –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º. –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {ends_at.strftime('%d.%m.%Y %H:%M')} (UTC). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            
            await session.commit()
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ—Å–ª–µ –∫–æ–º–º–∏—Ç–∞ –ø–æ–¥–ø–∏—Å–∫–∏
            await _update_user_subscription_status(user.id, session)
            await session.commit()
            await session.refresh(user)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–µ—Å—Å–∏–∏
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
            try:
                ends_at_moscow = ends_at.astimezone(ZoneInfo("Europe/Moscow"))
                ends_str = ends_at_moscow.strftime("%d.%m.%Y %H:%M")
            except:
                ends_str = ends_at.strftime("%d.%m.%Y %H:%M")
            
            notification_text = (
                f"‚úÖ <b>–ü–æ–¥–ø–∏—Å–∫–∞ –≤—ã–¥–∞–Ω–∞</b>\n\n"
                f"–ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä –≤—ã–¥–∞–ª –≤–∞–º –ø–æ–¥–ø–∏—Å–∫—É –Ω–∞ <b>{days} –¥–Ω–µ–π</b>.\n\n"
                f"üìÖ –ü–æ–¥–ø–∏—Å–∫–∞ –¥–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: <b>{ends_str} –ú–°–ö</b>\n\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: {reason}"
            )
            asyncio.create_task(_send_user_notification(tg_id, notification_text))
            
            return JSONResponse({"success": True, "message": f"–ü–æ–¥–ø–∏—Å–∫–∞ –≤—ã–¥–∞–Ω–∞ –Ω–∞ {days} –¥–Ω–µ–π"})
        
        elif action == "extend":
            # –ü—Ä–æ–¥–ª–µ–≤–∞–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â—É—é –ø–æ–¥–ø–∏—Å–∫—É –∏–ª–∏ —Å–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é
            active_sub = await session.scalar(
                select(Subscription)
                .where(Subscription.user_id == user.id)
                .where(Subscription.status == SubscriptionStatus.active)
                .order_by(Subscription.ends_at.desc().nullslast())
            )
            
            if active_sub and active_sub.ends_at and active_sub.ends_at > now:
                # –ü—Ä–æ–¥–ª–µ–≤–∞–µ–º –æ—Ç —Ç–µ–∫—É—â–µ–π –¥–∞—Ç—ã –æ–∫–æ–Ω—á–∞–Ω–∏—è
                new_ends_at = active_sub.ends_at + timedelta(days=days)
                active_sub.ends_at = new_ends_at
                subscription = active_sub
            else:
                # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –ø–æ–¥–ø–∏—Å–∫—É
                starts_at = now
                new_ends_at = now + timedelta(days=days)
                subscription = Subscription(
                    user_id=user.id,
                    plan_name=f"–ü–æ–¥–ø–∏—Å–∫–∞ –Ω–∞ {days} –¥–Ω–µ–π (–ø—Ä–æ–¥–ª–µ–Ω–∞ –∞–¥–º–∏–Ω–æ–º)",
                    price_cents=0,
                    currency="RUB",
                    status=SubscriptionStatus.active,
                    starts_at=starts_at,
                    ends_at=new_ends_at,
                )
                session.add(subscription)
            
            # –õ–æ–≥–∏—Ä—É–µ–º
            session.add(
                AuditLog(
                    action=AuditLogAction.subscription_created,
                    user_tg_id=tg_id,
                    admin_tg_id=actor_tg,
                    details=f"–ü—Ä–æ–¥–ª–µ–Ω–∞ –ø–æ–¥–ø–∏—Å–∫–∞ –Ω–∞ {days} –¥–Ω–µ–π –∞–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä–æ–º. –î–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: {new_ends_at.strftime('%d.%m.%Y %H:%M')} (UTC). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            
            await session.commit()
            
            # –û–±–Ω–æ–≤–ª—è–µ–º —Å—Ç–∞—Ç—É—Å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –ø–æ—Å–ª–µ –∫–æ–º–º–∏—Ç–∞ –ø–æ–¥–ø–∏—Å–∫–∏
            await _update_user_subscription_status(user.id, session)
            await session.commit()
            await session.refresh(user)  # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞–Ω–Ω—ã–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤ —Å–µ—Å—Å–∏–∏
            
            # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —É–≤–µ–¥–æ–º–ª–µ–Ω–∏–µ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—é
            try:
                new_ends_at_moscow = new_ends_at.astimezone(ZoneInfo("Europe/Moscow"))
                ends_str = new_ends_at_moscow.strftime("%d.%m.%Y %H:%M")
            except:
                ends_str = new_ends_at.strftime("%d.%m.%Y %H:%M")
            
            notification_text = (
                f"üîÑ <b>–ü–æ–¥–ø–∏—Å–∫–∞ –ø—Ä–æ–¥–ª–µ–Ω–∞</b>\n\n"
                f"–ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∞—Ç–æ—Ä –ø—Ä–æ–¥–ª–∏–ª –≤–∞—à—É –ø–æ–¥–ø–∏—Å–∫—É –Ω–∞ <b>{days} –¥–Ω–µ–π</b>.\n\n"
                f"üìÖ –ü–æ–¥–ø–∏—Å–∫–∞ —Ç–µ–ø–µ—Ä—å –¥–µ–π—Å—Ç–≤—É–µ—Ç –¥–æ: <b>{ends_str} –ú–°–ö</b>\n\n"
                f"–ü—Ä–∏—á–∏–Ω–∞: {reason}"
            )
            asyncio.create_task(_send_user_notification(tg_id, notification_text))
            
            return JSONResponse({"success": True, "message": f"–ü–æ–¥–ø–∏—Å–∫–∞ –ø—Ä–æ–¥–ª–µ–Ω–∞ –Ω–∞ {days} –¥–Ω–µ–π"})
        
    except Exception as e:
        import logging
        logging.error(f"Error managing subscription: {e}", exc_info=True)
        return JSONResponse({"success": False, "error": str(e)})


@app.post("/admin/web/users/message")
async def admin_web_send_message(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    ticket_id = form.get("ticket_id")
    text = str(form.get("text", "")).strip()
    if not text:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=reason_required", status_code=303)

    settings = get_settings()
    # –î–ª—è —Å–æ–æ–±—â–µ–Ω–∏–π –≤ —Ç–∏–∫–µ—Ç–µ –∏—Å–ø–æ–ª—å–∑—É–µ–º support-–±–æ—Ç, –∏–Ω–∞—á–µ –æ—Å–Ω–æ–≤–Ω–æ–π
    bot_token = None
    if ticket_id:
        bot_token = os.getenv("SUPPORT_BOT_TOKEN", "") or settings.support_bot_token or settings.admin_token
    if not bot_token:
        bot_token = os.getenv("BOT_TOKEN", "") or settings.admin_token
    if not bot_token:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=bot_token_missing", status_code=303)

    # –ü—Ä–æ–≤–µ—Ä–∫–∞, —á—Ç–æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    # –û—Ç–ø—Ä–∞–≤–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç –∏–º–µ–Ω–∏ –±–æ—Ç–∞
    try:
        import httpx
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.post(
                f"https://api.telegram.org/bot{bot_token}/sendMessage",
                json={"chat_id": tg_id, "text": text},
            )
            if r.status_code != 200 or not r.json().get("ok"):
                back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
                return RedirectResponse(url=f"{back}?error=send_failed", status_code=303)
    except Exception:
        back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
        return RedirectResponse(url=f"{back}?error=send_failed", status_code=303)

    # –°–æ–∑–¥–∞–µ–º –∏–ª–∏ –æ–±–Ω–æ–≤–ª—è–µ–º —Ç–∏–∫–µ—Ç
    ticket = None
    if ticket_id:
        try:
            ticket = await session.scalar(select(Ticket).where(Ticket.id == int(ticket_id)))
        except Exception:
            ticket = None
    if not ticket:
        ticket = await session.scalar(
            select(Ticket).where(Ticket.user_tg_id == tg_id).order_by(Ticket.updated_at.desc())
        )
    now = datetime.utcnow()
    if not ticket:
        ticket = Ticket(user_tg_id=tg_id, status=TicketStatus.open, created_at=now, updated_at=now)
        session.add(ticket)
        await session.flush()
    else:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ –∑–∞–∫—Ä—ã—Ç –ª–∏ —Ç–∏–∫–µ—Ç
        if ticket.status == TicketStatus.closed:
            back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
            return RedirectResponse(url=f"{back}?error=ticket_closed", status_code=303)
        ticket.updated_at = now

    # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ –≤ —Ç–∏–∫–µ—Ç–µ
    session.add(
        TicketMessage(
            ticket_id=ticket.id,
            user_tg_id=tg_id,
            direction=MessageDirection.outgoing,
            admin_tg_id=admin_user.get("tg_id"),
            text=text,
        )
    )

    # –õ–æ–≥–∏—Ä—É–µ–º –∫–∞–∫ admin_action
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"Admin sent message to user: {text}",
        )
    )
    await session.commit()
    back = request.headers.get("referer") or f"/admin/web/users/{tg_id}"
    return RedirectResponse(url=back, status_code=303)


@app.post("/admin/web/tickets/{ticket_id}/status")
async def admin_web_ticket_status(
    ticket_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    # –î–ª—è —Å—Ç–∞—Ç—É—Å–∞ —Ç–∏–∫–µ—Ç–∞ —É–±–∏—Ä–∞–µ–º —Å—Ç—Ä–æ–≥–∏–π CSRF, —á—Ç–æ–±—ã –∫–Ω–æ–ø–∫–∏ —Ä–∞–±–æ—Ç–∞–ª–∏ –±–µ–∑ ajax
    form = await request.form()
    action = str(form.get("action", "")).strip()
    if action not in {"close", "take"}:
        return RedirectResponse(url=f"/admin/web/tickets/{ticket_id}?error=bad_action", status_code=303)

    ticket = await session.scalar(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket:
        return RedirectResponse(url="/admin/web/tickets?error=ticket_not_found", status_code=303)
    # –ó–∞–∫—Ä—ã—Ç—ã–µ –Ω–µ –æ—Ç–∫—Ä—ã–≤–∞–µ–º/–Ω–µ –±–µ—Ä–µ–º
    if ticket.status == TicketStatus.closed:
        return RedirectResponse(url=f"/admin/web/tickets/{ticket_id}?error=already_closed", status_code=303)

    now = datetime.utcnow()
    system_text = None
    if action == "close":
        ticket.status = TicketStatus.closed
        ticket.closed_at = now
        system_text = f"–¢–∏–∫–µ—Ç –∑–∞–∫—Ä—ã—Ç –∞–¥–º–∏–Ω–æ–º {admin_user.get('tg_id')}"
    elif action == "take":
        ticket.status = TicketStatus.in_progress
        system_text = f"–¢–∏–∫–µ—Ç –≤–∑—è–ª –∞–¥–º–∏–Ω {admin_user.get('tg_id')}"
    ticket.updated_at = now

    if system_text:
        session.add(
            TicketMessage(
                ticket_id=ticket.id,
                user_tg_id=ticket.user_tg_id,
                direction=MessageDirection.system,
                admin_tg_id=admin_user.get("tg_id"),
                text=system_text,
            )
        )

    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=ticket.user_tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"Ticket {ticket.id} action: {action}",
        )
    )
    await session.commit()

    # –ü—ã—Ç–∞–µ–º—Å—è —É–≤–µ–¥–æ–º–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è —á–µ—Ä–µ–∑ support-–±–æ—Ç
    try:
        bot_token = os.getenv("SUPPORT_BOT_TOKEN", "") or get_settings().support_bot_token
        if bot_token:
            import httpx
            async with httpx.AsyncClient(timeout=10.0) as client:
                await client.post(
                    f"https://api.telegram.org/bot{bot_token}/sendMessage",
                    json={"chat_id": ticket.user_tg_id, "text": system_text},
                )
    except Exception:
        pass

    return RedirectResponse(url=f"/admin/web/tickets/{ticket_id}", status_code=303)


@app.get("/admin/web/api/tickets/{ticket_id}/messages")
async def admin_api_ticket_messages(
    ticket_id: int,
    since_id: int | None = Query(default=None),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –≤–µ—Ä–Ω—É—Ç—å —Å–æ–æ–±—â–µ–Ω–∏—è —Ç–∏–∫–µ—Ç–∞ (–¥–ª—è –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è —á–∞—Ç–∞ –≤ –∞–¥–º–∏–Ω–∫–µ)."""
    ticket = await session.scalar(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket:
        raise HTTPException(status_code=404, detail="ticket_not_found")

    stmt = select(TicketMessage).where(TicketMessage.ticket_id == ticket.id)
    if since_id is not None:
        stmt = stmt.where(TicketMessage.id > since_id)
    stmt = stmt.order_by(TicketMessage.id.asc())
    result = await session.scalars(stmt)
    items = []
    for m in result.all():
        items.append(
            {
                "id": m.id,
                "created_at": m.created_at.isoformat(),
                "direction": m.direction.value if hasattr(m.direction, "value") else str(m.direction),
                "user_tg_id": m.user_tg_id,
                "admin_tg_id": m.admin_tg_id,
                "text": m.text,
            }
        )
    return {"messages": items}

@app.post("/tickets/create")
async def create_ticket_endpoint(payload: dict, session: AsyncSession = Depends(get_session)):
    """–°–æ–∑–¥–∞–Ω–∏–µ —Ç–∏–∫–µ—Ç–∞ –ø–æ –∑–∞–ø—Ä–æ—Å—É –æ—Ç –æ—Å–Ω–æ–≤–Ω–æ–≥–æ –±–æ—Ç–∞."""
    tg_id = int(payload.get("tg_id", 0))
    topic = (payload.get("topic") or "").strip()
    if not tg_id or not topic:
        raise HTTPException(status_code=400, detail="tg_id_and_topic_required")

    now = datetime.utcnow()
    ticket = Ticket(
        user_tg_id=tg_id,
        topic=topic,
        status=TicketStatus.new,
        created_at=now,
        updated_at=now,
    )
    session.add(ticket)
    await session.commit()
    await session.refresh(ticket)
    return {"ticket_id": ticket.id}


@app.get("/admin/web/tickets", response_class=HTMLResponse)
async def admin_web_tickets(
    request: Request,
    status: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)

    status_filter = status.lower()
    if status_filter not in {"all", "open", "closed"}:
        status_filter = "all"

    page_size = 20
    stmt = select(Ticket).order_by(Ticket.updated_at.desc())
    if status_filter == "open":
        stmt = stmt.where(Ticket.status != TicketStatus.closed)
    elif status_filter == "closed":
        stmt = stmt.where(Ticket.status == TicketStatus.closed)

    total = await session.scalar(select(func.count()).select_from(stmt.subquery()))
    result = await session.scalars(stmt.limit(page_size).offset((page - 1) * page_size))
    tickets = result.all()

    has_next = (page * page_size) < (total or 0)
    has_prev = page > 1

    return templates.TemplateResponse(
        "tickets.html",
        {
            "request": request,
            "tickets": tickets,
            "status": status_filter,
            "page": page,
            "has_next": has_next,
            "has_prev": has_prev,
            "admin_user": admin_user,
        },
    )


@app.get("/admin/web/api/tickets")
async def admin_api_tickets(
    status: str = Query(default="all"),
    page: int = Query(default=1, ge=1),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """JSON-API –¥–ª—è —Å–ø–∏—Å–∫–∞ —Ç–∏–∫–µ—Ç–æ–≤ (–∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –¥–ª—è –∞–≤—Ç–æ–æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –∏–Ω—Ç–µ—Ä—Ñ–µ–π—Å–∞)."""
    status_filter = status.lower()
    if status_filter not in {"all", "open", "closed"}:
        status_filter = "all"

    page_size = 20
    stmt = select(Ticket).order_by(Ticket.updated_at.desc())
    if status_filter == "open":
        stmt = stmt.where(Ticket.status != TicketStatus.closed)
    elif status_filter == "closed":
        stmt = stmt.where(Ticket.status == TicketStatus.closed)

    total = await session.scalar(select(func.count()).select_from(stmt.subquery()))
    result = await session.scalars(stmt.limit(page_size).offset((page - 1) * page_size))
    tickets = result.all()

    items: list[dict[str, object]] = []
    # –ü—Ä–∏–≤–æ–¥–∏–º –≤—Ä–µ–º—è –∫ –ú–æ—Å–∫–≤–µ (UTC+3)
    try:
        from zoneinfo import ZoneInfo
        msk_tz = ZoneInfo("Europe/Moscow")
    except Exception:
        msk_tz = None

    for t in tickets:
        if t.updated_at and msk_tz:
            msk_time = t.updated_at.astimezone(msk_tz)
            updated_str = msk_time.strftime("%d.%m.%Y %H:%M")
        else:
            updated_str = t.updated_at.isoformat() if t.updated_at else None
        items.append(
            {
                "id": t.id,
                "user_tg_id": t.user_tg_id,
                "topic": t.topic,
                "status": t.status.value if hasattr(t.status, "value") else str(t.status),
                "updated_at": updated_str,
            }
        )

    has_next = (page * page_size) < (total or 0)
    has_prev = page > 1

    return {
        "tickets": items,
        "page": page,
        "has_next": has_next,
        "has_prev": has_prev,
        "status": status_filter,
    }


@app.get("/admin/web/logs", response_class=HTMLResponse)
async def admin_web_logs(
    request: Request,
    q: str | None = Query(default=None),
    action: str | None = Query(default="all"),
    page: int = Query(default=1, ge=1),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–í–µ–±-—Å—Ç—Ä–∞–Ω–∏—Ü–∞ –ª–æ–≥–æ–≤ —Å —Ñ–∏–ª—å—Ç—Ä–∞–º–∏."""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)

    action_filter = (action or "all")
    valid_actions = {
        "user_registered",
        "balance_credited",
        "user_blocked",
        "user_unblocked",
        "subscription_created",
        "subscription_activated",
        "payment_processed",
        "payment_created",
        "payment_status_changed",
        "payment_webhook_received",
        "admin_action",
        "backup_action",
    }
    if action_filter not in valid_actions and action_filter != "all":
        action_filter = "all"

    page_size = 50
    stmt = select(AuditLog).order_by(AuditLog.created_at.desc())
    if action_filter != "all":
        stmt = stmt.where(AuditLog.action == AuditLogAction(action_filter))

    if q:
        qq = q.strip()
        conds = []
        if qq.isdigit():
            v = int(qq)
            conds.append(AuditLog.user_tg_id == v)
            conds.append(AuditLog.admin_tg_id == v)
        like = f"%{qq.lower()}%"
        conds.append(func.lower(AuditLog.details).like(like))
        stmt = stmt.where(or_(*conds))

    total = await session.scalar(select(func.count()).select_from(stmt.subquery()))
    result = await session.scalars(stmt.limit(page_size).offset((page - 1) * page_size))
    rows = result.all()

    # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –≤—Ä–µ–º—è –≤ –ú–°–ö
    try:
        from zoneinfo import ZoneInfo
        msk_tz = ZoneInfo("Europe/Moscow")
    except Exception:
        msk_tz = None

    logs = []
    for log in rows:
        if msk_tz:
            created_str = log.created_at.astimezone(msk_tz).strftime("%d.%m.%Y %H:%M")
        else:
            created_str = log.created_at.isoformat()
        logs.append(
            {
                "id": log.id,
                "action": log.action.value if hasattr(log.action, "value") else str(log.action),
                "user_tg_id": log.user_tg_id,
                "admin_tg_id": log.admin_tg_id,
                "details": log.details,
                "created_at": created_str,
            }
        )

    has_next = (page * page_size) < (total or 0)
    has_prev = page > 1

    return templates.TemplateResponse(
        "logs.html",
        {
            "request": request,
            "logs": logs,
            "page": page,
            "has_next": has_next,
            "has_prev": has_prev,
            "action_filter": action_filter,
            "q": q or "",
            "admin_user": admin_user,
        },
    )

@app.get("/admin/web/tickets/{ticket_id}", response_class=HTMLResponse)
async def admin_web_ticket_detail(
    request: Request,
    ticket_id: int,
    page: int = Query(default=1, ge=1),
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ü–æ–¥—Ä–æ–±–Ω—ã–π –ø—Ä–æ—Å–º–æ—Ç—Ä —Ç–∏–∫–µ—Ç–∞ —Å –∏—Å—Ç–æ—Ä–∏–µ–π —Å–æ–æ–±—â–µ–Ω–∏–π."""
    ticket = await session.scalar(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket:
        return RedirectResponse(url="/admin/web/tickets?error=ticket_not_found", status_code=303)

    user = await session.scalar(select(User).where(User.tg_id == ticket.user_tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/tickets?error=user_not_found", status_code=303)

    try:
        from zoneinfo import ZoneInfo
        tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"

    info = {
        "id": ticket.id,
        "user_tg_id": ticket.user_tg_id,
        "topic": ticket.topic,
        "status": ticket.status.value if hasattr(ticket.status, "value") else str(ticket.status),
        "created_at": fmt(ticket.created_at),
        "updated_at": fmt(ticket.updated_at),
        "closed_at": fmt(ticket.closed_at),
    }

    page_size = 30
    total_messages = await session.scalar(
        select(func.count()).select_from(TicketMessage).where(TicketMessage.ticket_id == ticket.id)
    )
    # –ï—Å–ª–∏ —è–≤–Ω–æ –Ω–µ –ø–µ—Ä–µ–¥–∞–ª–∏ page, –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω—é—é —Å—Ç—Ä–∞–Ω–∏—Ü—É, —á—Ç–æ–±—ã –≤–∏–¥–µ—Ç—å –Ω–æ–≤—ã–µ —Å–æ–æ–±—â–µ–Ω–∏—è
    page_param = request.query_params.get("page")
    if not page_param:
        page = max(((total_messages or 0) - 1) // page_size + 1, 1)

    result = await session.scalars(
        select(TicketMessage)
        .where(TicketMessage.ticket_id == ticket.id)
        .order_by(TicketMessage.created_at.asc())
        .limit(page_size)
        .offset((page - 1) * page_size)
    )
    messages_raw = result.all()
    
    # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º messages –≤ —Å–ª–æ–≤–∞—Ä–∏ —Å –ø—Ä–∞–≤–∏–ª—å–Ω—ã–º direction (—Å—Ç—Ä–æ–∫–∞ –≤–º–µ—Å—Ç–æ enum)
    messages = []
    for m in messages_raw:
        messages.append({
            "id": m.id,
            "created_at": fmt(m.created_at),
            "direction": m.direction.value if hasattr(m.direction, "value") else str(m.direction),
            "user_tg_id": m.user_tg_id,
            "admin_tg_id": m.admin_tg_id,
            "text": m.text,
        })
    
    has_next = (page * page_size) < (total_messages or 0)
    has_prev = page > 1

    return templates.TemplateResponse(
        "ticket_detail.html",
        {
            "request": request,
            "ticket": info,
            "messages": messages,
            "page": page,
            "has_next": has_next,
            "has_prev": has_prev,
            "csrf_token": _get_csrf_token(request),
            "admin_user": admin_user,
            "user": user,
        },
    )


@app.post("/admin/web/users/role")
async def admin_web_set_role(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    await _require_csrf(request)
    form = await request.form()
    tg_id = int(str(form.get("tg_id", "0")))
    role = str(form.get("role", "user")).strip()
    reason = str(form.get("reason", "")).strip()
    if not reason:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=reason_required", status_code=303)

    allowed_roles = {"admin", "moderator", "user"}
    admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()

    if tg_id in admin_ids:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=role_protected", status_code=303)
    if role not in allowed_roles:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=bad_role", status_code=303)

    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        return RedirectResponse(url="/admin/web/users?error=user_not_found", status_code=303)

    overrides_result = await session.scalars(select(AdminOverride))
    overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}

    actor_role = _get_effective_role(admin_user.get("tg_id"), admin_ids, overrides_map)
    target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
    if admin_user.get("tg_id") == user.tg_id:
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_self", status_code=303)
    if _role_rank(actor_role) <= _role_rank(target_role):
        return RedirectResponse(url=f"/admin/web/users/{tg_id}?error=forbidden_rank", status_code=303)

    override_obj = await session.scalar(select(AdminOverride).where(AdminOverride.tg_id == tg_id))
    if not override_obj:
        override_obj = AdminOverride(tg_id=tg_id, role=role)
        session.add(override_obj)
    else:
        override_obj.role = role

    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            user_tg_id=tg_id,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"Role set to {role} (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
        )
    )
    await session.commit()
    return RedirectResponse(url=f"/admin/web/users/{tg_id}", status_code=303)


@app.post("/admin/web/users/bulk/block")
async def admin_web_bulk_block_users(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ú–∞—Å—Å–æ–≤–∞—è –±–ª–æ–∫–∏—Ä–æ–≤–∫–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    await _require_csrf(request)
    try:
        data = await request.json()
        user_ids = data.get("user_ids", [])
        reason = data.get("reason", "").strip()
        
        if not user_ids:
            return {"success": False, "error": "–ù–µ –≤—ã–±—Ä–∞–Ω—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏"}
        if not reason:
            return {"success": False, "error": "–ù—É–∂–Ω–æ —É–∫–∞–∑–∞—Ç—å –ø—Ä–∏—á–∏–Ω—É"}
        
        admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
        overrides_result = await session.scalars(select(AdminOverride))
        overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
        
        actor_tg = admin_user.get("tg_id")
        actor_role = _get_effective_role(actor_tg, admin_ids, overrides_map)
        
        processed = 0
        for user_id in user_ids:
            user = await session.scalar(select(User).where(User.tg_id == user_id))
            if not user:
                continue
            if actor_tg == user.tg_id:
                continue  # –ù–µ –±–ª–æ–∫–∏—Ä—É–µ–º —Å–µ–±—è
            target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
            if _role_rank(actor_role) <= _role_rank(target_role):
                continue  # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –µ—Å–ª–∏ –Ω–µ–¥–æ—Å—Ç–∞—Ç–æ—á–Ω–æ –ø—Ä–∞–≤
            user.is_active = False
            session.add(
                AuditLog(
                    action=AuditLogAction.user_blocked,
                    user_tg_id=user_id,
                    admin_tg_id=actor_tg,
                    details=f"–ú–∞—Å—Å–æ–≤–∞—è –±–ª–æ–∫–∏—Ä–æ–≤–∫–∞ (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            processed += 1
        
        await session.commit()
        return {"success": True, "processed": processed}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/admin/web/users/bulk/unblock")
async def admin_web_bulk_unblock_users(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ú–∞—Å—Å–æ–≤–∞—è —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∫–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    await _require_csrf(request)
    try:
        data = await request.json()
        user_ids = data.get("user_ids", [])
        reason = data.get("reason", "").strip()
        
        if not user_ids:
            return {"success": False, "error": "–ù–µ –≤—ã–±—Ä–∞–Ω—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏"}
        if not reason:
            return {"success": False, "error": "–ù—É–∂–Ω–æ —É–∫–∞–∑–∞—Ç—å –ø—Ä–∏—á–∏–Ω—É"}
        
        admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
        overrides_result = await session.scalars(select(AdminOverride))
        overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
        
        actor_tg = admin_user.get("tg_id")
        actor_role = _get_effective_role(actor_tg, admin_ids, overrides_map)
        
        processed = 0
        for user_id in user_ids:
            user = await session.scalar(select(User).where(User.tg_id == user_id))
            if not user:
                continue
            if actor_tg == user.tg_id:
                continue
            target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
            if _role_rank(actor_role) <= _role_rank(target_role):
                continue
            user.is_active = True
            session.add(
                AuditLog(
                    action=AuditLogAction.user_unblocked,
                    user_tg_id=user_id,
                    admin_tg_id=actor_tg,
                    details=f"–ú–∞—Å—Å–æ–≤–∞—è —Ä–∞–∑–±–ª–æ–∫–∏—Ä–æ–≤–∫–∞ (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            processed += 1
        
        await session.commit()
        return {"success": True, "processed": processed}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/admin/web/users/bulk/credit")
async def admin_web_bulk_credit_users(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ú–∞—Å—Å–æ–≤–æ–µ –∏–∑–º–µ–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    await _require_csrf(request)
    try:
        data = await request.json()
        user_ids = data.get("user_ids", [])
        amount = data.get("amount")
        reason = data.get("reason", "").strip()
        
        if not user_ids:
            return {"success": False, "error": "–ù–µ –≤—ã–±—Ä–∞–Ω—ã –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–∏"}
        if amount is None:
            return {"success": False, "error": "–ù–µ —É–∫–∞–∑–∞–Ω–∞ —Å—É–º–º–∞"}
        if not reason:
            return {"success": False, "error": "–ù—É–∂–Ω–æ —É–∫–∞–∑–∞—Ç—å –ø—Ä–∏—á–∏–Ω—É"}
        
        amount_cents = int(float(amount) * 100)
        admin_ids = set(int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()) if os.getenv("ADMIN_IDS") else set()
        overrides_result = await session.scalars(select(AdminOverride))
        overrides_map = {ov.tg_id: ov.role for ov in overrides_result.all()}
        
        actor_tg = admin_user.get("tg_id")
        actor_role = _get_effective_role(actor_tg, admin_ids, overrides_map)
        
        processed = 0
        for user_id in user_ids:
            user = await session.scalar(select(User).where(User.tg_id == user_id))
            if not user:
                continue
            target_role = _get_effective_role(user.tg_id, admin_ids, overrides_map)
            if actor_tg != user.tg_id and _role_rank(actor_role) <= _role_rank(target_role):
                continue
            old_balance = user.balance
            user.balance += amount_cents
            session.add(
                BalanceTransaction(
                    user_id=user.id,
                    amount=amount_cents,
                    balance_before=old_balance,
                    balance_after=user.balance,
                    reason=f"–ú–∞—Å—Å–æ–≤–æ–µ –∏–∑–º–µ–Ω–µ–Ω–∏–µ (web). –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            session.add(
                AuditLog(
                    action=AuditLogAction.balance_credited,
                    user_tg_id=user_id,
                    admin_tg_id=actor_tg,
                    details=f"–ú–∞—Å—Å–æ–≤–æ–µ –∏–∑–º–µ–Ω–µ–Ω–∏–µ –±–∞–ª–∞–Ω—Å–∞: {amount} USD. –ü—Ä–∏—á–∏–Ω–∞: {reason}",
                )
            )
            processed += 1
        
        await session.commit()
        return {"success": True, "processed": processed}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/admin/web/settings", response_class=HTMLResponse)
async def admin_web_settings(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ –Ω–∞—Å—Ç—Ä–æ–µ–∫ —Å–∏—Å—Ç–µ–º—ã"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    settings_result = await session.scalars(select(SystemSetting).order_by(SystemSetting.key))
    settings_list = settings_result.all()
    settings_dict = {s.key: s.value for s in settings_list}
    
    csrf_token = _get_csrf_token(request)
    return templates.TemplateResponse(
        "settings.html",
        {
            "request": request,
            "admin_user": admin_user,
            "settings": settings_dict,
            "csrf_token": csrf_token,
        },
    )


@app.post("/admin/web/settings")
async def admin_web_update_settings(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ –Ω–∞—Å—Ç—Ä–æ–µ–∫ —Å–∏—Å—Ç–µ–º—ã"""
    await _require_csrf(request)
    try:
        form_data = await request.form()
        actor_tg = admin_user.get("tg_id")
        
        for key, value in form_data.items():
            if key == "csrf_token":
                continue
            
            # –ù–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è –∫–ª—é—á–µ–π –Ω–∞—Å—Ç—Ä–æ–µ–∫
            if key == "trial_days":
                key = "trial_period_days"
            elif key == "auto_extend_subscription":
                key = "auto_renew_subscription"
            
            # –°–ø–µ—Ü–∏–∞–ª—å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –¥–ª—è –ø–æ–ª–∏—Ç–∏–∫–∏ –∫–æ–Ω—Ñ–∏–¥–µ–Ω—Ü–∏–∞–ª—å–Ω–æ—Å—Ç–∏
            if key == "privacy_policy_content":
                # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞—Ç—É –ø–æ—Å–ª–µ–¥–Ω–µ–≥–æ –∏–∑–º–µ–Ω–µ–Ω–∏—è –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –ø–æ–ª–∏—Ç–∏–∫–∏
                from datetime import timezone
                now_utc = datetime.now(timezone.utc)
                # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –¥–∞—Ç—É –≤ —á–∏—Ç–∞–µ–º—ã–π —Ñ–æ—Ä–º–∞—Ç (–ú–°–ö = UTC+3)
                from zoneinfo import ZoneInfo
                msk_tz = ZoneInfo("Europe/Moscow")
                now_msk = now_utc.astimezone(msk_tz)
                date_str = now_msk.strftime("%d.%m.%Y %H:%M –ú–°–ö")
                
                # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø–æ–ª–∏—Ç–∏–∫—É
                setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == key))
                if setting:
                    setting.value = str(value)
                    setting.updated_by_tg_id = actor_tg
                    setting.updated_at = now_utc
                else:
                    setting = SystemSetting(
                        key=key,
                        value=str(value),
                        updated_by_tg_id=actor_tg,
                        updated_at=now_utc,
                    )
                    session.add(setting)
                
                # –û–±–Ω–æ–≤–ª—è–µ–º –¥–∞—Ç—É –ø–æ—Å–ª–µ–¥–Ω–µ–≥–æ –∏–∑–º–µ–Ω–µ–Ω–∏—è
                date_setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == "privacy_policy_updated_at"))
                if date_setting:
                    date_setting.value = date_str
                    date_setting.updated_by_tg_id = actor_tg
                    date_setting.updated_at = now_utc
                else:
                    date_setting = SystemSetting(
                        key="privacy_policy_updated_at",
                        value=date_str,
                        updated_by_tg_id=actor_tg,
                        updated_at=now_utc,
                    )
                    session.add(date_setting)
                
                session.add(
                    AuditLog(
                        action=AuditLogAction.admin_action,
                        admin_tg_id=actor_tg,
                        details=f"–û–±–Ω–æ–≤–ª–µ–Ω–∞ –ø–æ–ª–∏—Ç–∏–∫–∞ –∫–æ–Ω—Ñ–∏–¥–µ–Ω—Ü–∏–∞–ª—å–Ω–æ—Å—Ç–∏",
                    )
                )
                continue
            
            # –°–ø–µ—Ü–∏–∞–ª—å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –¥–ª—è —Å—É–º–º –≤ RUB - –∫–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º –≤ –∫–æ–ø–µ–π–∫–∏
            # –í —Ñ–æ—Ä–º–µ –∏—Å–ø–æ–ª—å–∑—É—é—Ç—Å—è –∫–ª—é—á–∏ —Å _cents, –Ω–æ –∑–Ω–∞—á–µ–Ω–∏—è –≤–≤–æ–¥—è—Ç—Å—è –≤ RUB
            if key in ["referral_reward_referrer_cents", "referral_reward_referred_cents", "min_topup_amount_cents", "max_topup_amount_cents"]:
                try:
                    amount_rub = float(value)
                    value = str(int(amount_rub * 100))
                except (ValueError, TypeError):
                    pass
            
            # –î–ª—è max_topup_amount_cents - –µ—Å–ª–∏ –ø—É—Å—Ç–æ, —É–¥–∞–ª—è–µ–º –Ω–∞—Å—Ç—Ä–æ–π–∫—É
            if key == "max_topup_amount_cents" and (not value or value.strip() == ""):
                setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == key))
                if setting:
                    await session.delete(setting)
                continue
            
            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º –ø—É—Å—Ç—ã–µ –∑–Ω–∞—á–µ–Ω–∏—è –¥–ª—è —Ç–µ–∫—Å—Ç–æ–≤—ã—Ö –ø–æ–ª–µ–π (–∫—Ä–æ–º–µ –ø–æ–ª–∏—Ç–∏–∫–∏)
            if not value or (isinstance(value, str) and not value.strip()):
                continue
            
            setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == key))
            if setting:
                setting.value = str(value)
                setting.updated_by_tg_id = actor_tg
                setting.updated_at = datetime.utcnow()
            else:
                setting = SystemSetting(
                    key=key,
                    value=str(value),
                    updated_by_tg_id=actor_tg,
                    updated_at=datetime.utcnow(),
                )
                session.add(setting)
            
            session.add(
                AuditLog(
                    action=AuditLogAction.admin_action,
                    admin_tg_id=actor_tg,
                    details=f"–ò–∑–º–µ–Ω–µ–Ω–∞ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞: {key} = {value}",
                )
            )
        
        await session.commit()
        return RedirectResponse(url="/admin/web/settings?success=1", status_code=303)
    except Exception as e:
        return RedirectResponse(url=f"/admin/web/settings?error={str(e)}", status_code=303)


@app.get("/settings/bot")
async def get_bot_settings(
    session: AsyncSession = Depends(get_session),
) -> dict:
    """–ü–æ–ª—É—á–∏—Ç—å –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ –±–æ—Ç–∞ (–ø—É–±–ª–∏—á–Ω—ã–π endpoint)"""
    settings_result = await session.scalars(
        select(SystemSetting).where(
            SystemSetting.key.in_([
                "bot_welcome_message",
                "bot_help_message",
                "min_topup_amount_cents",
                "max_topup_amount_cents",
            ])
        )
    )
    settings_list = settings_result.all()
    settings_dict = {s.key: s.value for s in settings_list}
    
    # –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ–º —Å—É–º–º—ã –∏–∑ –∫–æ–ø–µ–µ–∫ –≤ —Ä—É–±–ª–∏ –¥–ª—è —É–¥–æ–±—Å—Ç–≤–∞
    result = {}
    if "bot_welcome_message" in settings_dict:
        result["welcome_message"] = settings_dict["bot_welcome_message"]
    if "bot_help_message" in settings_dict:
        result["help_message"] = settings_dict["bot_help_message"]
    if "min_topup_amount_cents" in settings_dict:
        try:
            result["min_topup_amount_rub"] = float(settings_dict["min_topup_amount_cents"]) / 100
        except (ValueError, TypeError):
            result["min_topup_amount_rub"] = 1.0
    if "max_topup_amount_cents" in settings_dict:
        try:
            result["max_topup_amount_rub"] = float(settings_dict["max_topup_amount_cents"]) / 100
        except (ValueError, TypeError):
            pass
    
    return result


@app.get("/admin/web/subscription-plans", response_class=HTMLResponse)
async def admin_web_subscription_plans(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ç–∞—Ä–∏—Ñ–∞–º–∏ –ø–æ–¥–ø–∏—Å–∫–∏"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    await _ensure_default_plans(session)
    
    plans_result = await session.scalars(
        select(SubscriptionPlan)
        .order_by(SubscriptionPlan.display_order, SubscriptionPlan.days)
    )
    plans = plans_result.all()
    
    csrf_token = _get_csrf_token(request)
    return templates.TemplateResponse(
        "subscription_plans.html",
        {
            "request": request,
            "admin_user": admin_user,
            "plans": plans,
            "csrf_token": csrf_token,
        },
    )


@app.post("/admin/web/subscription-plans/update")
async def admin_web_update_subscription_plan(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–û–±–Ω–æ–≤–ª–µ–Ω–∏–µ —Ç–∞—Ä–∏—Ñ–∞ –ø–æ–¥–ø–∏—Å–∫–∏"""
    await _require_csrf(request)
    try:
        form_data = await request.form()
        plan_id = int(form_data.get("plan_id"))
        name = form_data.get("name", "").strip()
        description = form_data.get("description", "").strip()
        price_rub = float(form_data.get("price_rub", 0))
        is_active = form_data.get("is_active") == "on"
        display_order = int(form_data.get("display_order", 0))
        
        plan = await session.scalar(select(SubscriptionPlan).where(SubscriptionPlan.id == plan_id))
        if not plan:
            return RedirectResponse(url="/admin/web/subscription-plans?error=plan_not_found", status_code=303)
        
        old_price = plan.price_cents
        plan.name = name
        plan.description = description
        plan.price_cents = int(price_rub * 100)
        plan.is_active = is_active
        plan.display_order = display_order
        plan.updated_at = datetime.now(timezone.utc)
        
        session.add(
            AuditLog(
                action=AuditLogAction.admin_action,
                admin_tg_id=admin_user.get("tg_id"),
                details=f"–û–±–Ω–æ–≤–ª–µ–Ω —Ç–∞—Ä–∏—Ñ –ø–æ–¥–ø–∏—Å–∫–∏: {plan.name} (ID: {plan_id}). –¶–µ–Ω–∞: {old_price / 100:.2f} RUB -> {price_rub:.2f} RUB",
            )
        )
        
        await session.commit()
        return RedirectResponse(url="/admin/web/subscription-plans?success=updated", status_code=303)
    except Exception as e:
        import logging
        logging.error(f"Error updating subscription plan: {e}", exc_info=True)
        return RedirectResponse(url=f"/admin/web/subscription-plans?error={str(e)}", status_code=303)


@app.post("/admin/web/subscription-plans/create")
async def admin_web_create_subscription_plan(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ —Ç–∞—Ä–∏—Ñ–∞ –ø–æ–¥–ø–∏—Å–∫–∏"""
    import logging
    logging.info(f"Creating subscription plan: path={request.url.path}, method={request.method}")
    
    await _require_csrf(request)
    logging.info("CSRF validation passed")
    
    try:
        form_data = await request.form()
        logging.info(f"Form data received: {dict(form_data)}")
        
        days = int(form_data.get("days", 0))
        name = str(form_data.get("name", "")).strip()
        description = str(form_data.get("description", "")).strip() or None
        price_rub = float(form_data.get("price_rub", 0))
        is_active = form_data.get("is_active") == "on"
        display_order = int(form_data.get("display_order", 0))
        
        if days < 1 or days > 3650:  # –ú–∞–∫—Å–∏–º—É–º 10 –ª–µ—Ç
            return RedirectResponse(url="/admin/web/subscription-plans?error=invalid_days", status_code=303)
        
        if not name:
            return RedirectResponse(url="/admin/web/subscription-plans?error=name_required", status_code=303)
        
        if price_rub <= 0:
            return RedirectResponse(url="/admin/web/subscription-plans?error=invalid_price", status_code=303)
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ —É–∂–µ —Ç–∞—Ä–∏—Ñ–∞ —Å —Ç–∞–∫–∏–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ–º –¥–Ω–µ–π
        existing_plan = await session.scalar(
            select(SubscriptionPlan).where(SubscriptionPlan.days == days)
        )
        if existing_plan:
            return RedirectResponse(url="/admin/web/subscription-plans?error=plan_with_days_exists", status_code=303)
        
        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ
        new_plan = SubscriptionPlan(
            days=days,
            name=name,
            description=description,
            price_cents=int(price_rub * 100),
            is_active=is_active,
            display_order=display_order,
        )
        session.add(new_plan)
        
        session.add(
            AuditLog(
                action=AuditLogAction.admin_action,
                admin_tg_id=admin_user.get("tg_id"),
                details=f"–°–æ–∑–¥–∞–Ω –Ω–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ –ø–æ–¥–ø–∏—Å–∫–∏: {name} ({days} –¥–Ω–µ–π). –¶–µ–Ω–∞: {price_rub:.2f} RUB",
            )
        )
        
        await session.commit()
        await session.refresh(new_plan)  # –û–±–Ω–æ–≤–ª—è–µ–º –æ–±—ä–µ–∫—Ç –ø–æ—Å–ª–µ commit, —á—Ç–æ–±—ã –ø–æ–ª—É—á–∏—Ç—å ID
        
        logging.info(f"–°–æ–∑–¥–∞–Ω –Ω–æ–≤—ã–π —Ç–∞—Ä–∏—Ñ –ø–æ–¥–ø–∏—Å–∫–∏: ID={new_plan.id}, name={name}, days={days}, price={price_rub:.2f} RUB")
        logging.info(f"Redirecting to /admin/web/subscription-plans?success=created")
        
        return RedirectResponse(url="/admin/web/subscription-plans?success=created", status_code=303)
    except HTTPException as e:
        # –ï—Å–ª–∏ —ç—Ç–æ HTTPException (–Ω–∞–ø—Ä–∏–º–µ—Ä, CSRF –æ—à–∏–±–∫–∞), –ø—Ä–æ–±—Ä–∞—Å—ã–≤–∞–µ–º –¥–∞–ª—å—à–µ
        logging.error(f"HTTPException in create subscription plan: {e.status_code} - {e.detail}")
        raise
    except Exception as e:
        logging.error(f"Error creating subscription plan: {e}", exc_info=True)
        return RedirectResponse(url=f"/admin/web/subscription-plans?error={str(e)}", status_code=303)


@app.get("/admin/web/promo-codes", response_class=HTMLResponse)
async def admin_web_promo_codes(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è –ø—Ä–æ–º–æ–∫–æ–¥–∞–º–∏"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    promo_codes_result = await session.scalars(
        select(PromoCode).order_by(PromoCode.created_at.desc())
    )
    promo_codes = promo_codes_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    promo_codes_data = []
    for p in promo_codes:
        promo_codes_data.append({
            "id": p.id,
            "code": p.code,
            "discount_percent": p.discount_percent,
            "discount_amount_cents": p.discount_amount_cents,
            "max_uses": p.max_uses,
            "used_count": p.used_count,
            "is_active": p.is_active,
            "valid_from": fmt(p.valid_from) if p.valid_from else "‚Äî",
            "valid_until": fmt(p.valid_until) if p.valid_until else "‚Äî",
            "created_at": fmt(p.created_at),
        })
    
    csrf_token = _get_csrf_token(request)
    return templates.TemplateResponse(
        "promo_codes.html",
        {
            "request": request,
            "admin_user": admin_user,
            "promo_codes": promo_codes_data,
            "csrf_token": csrf_token,
        },
    )


@app.post("/admin/web/promo-codes/create")
async def admin_web_create_promo_code(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–≥–æ –ø—Ä–æ–º–æ–∫–æ–¥–∞"""
    await _require_csrf(request)
    try:
        form = await request.form()
        
        code = str(form.get("code", "")).strip().upper()
        discount_type = str(form.get("discount_type", ""))
        discount_value = str(form.get("discount_value", ""))
        max_uses = form.get("max_uses")
        valid_from = form.get("valid_from")
        valid_until = form.get("valid_until")
        description = str(form.get("description", "")).strip()
        
        if not code:
            return RedirectResponse(url="/admin/web/promo-codes?error=code_required", status_code=303)
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å –∫–æ–¥–∞
        existing = await session.scalar(select(PromoCode).where(PromoCode.code == code))
        if existing:
            return RedirectResponse(url="/admin/web/promo-codes?error=code_exists", status_code=303)
        
        discount_percent = None
        discount_amount_cents = None
        
        if discount_type == "percent":
            try:
                discount_percent = int(float(discount_value))
                if discount_percent < 0 or discount_percent > 100:
                    return RedirectResponse(url="/admin/web/promo-codes?error=invalid_percent", status_code=303)
            except (ValueError, TypeError):
                return RedirectResponse(url="/admin/web/promo-codes?error=invalid_percent", status_code=303)
        elif discount_type == "amount":
            try:
                discount_amount_cents = int(float(discount_value) * 100)
                if discount_amount_cents <= 0:
                    return RedirectResponse(url="/admin/web/promo-codes?error=invalid_amount", status_code=303)
            except (ValueError, TypeError):
                return RedirectResponse(url="/admin/web/promo-codes?error=invalid_amount", status_code=303)
        else:
            return RedirectResponse(url="/admin/web/promo-codes?error=invalid_discount_type", status_code=303)
        
        max_uses_int = None
        if max_uses:
            try:
                max_uses_int = int(max_uses)
                if max_uses_int <= 0:
                    max_uses_int = None
            except (ValueError, TypeError):
                max_uses_int = None
        
        valid_from_dt = None
        if valid_from:
            try:
                from zoneinfo import ZoneInfo
                moscow_tz = ZoneInfo("Europe/Moscow")
                valid_from_dt = datetime.strptime(valid_from, "%Y-%m-%dT%H:%M")
                valid_from_dt = moscow_tz.localize(valid_from_dt).astimezone(ZoneInfo("UTC"))
            except Exception:
                pass
        
        valid_until_dt = None
        if valid_until:
            try:
                from zoneinfo import ZoneInfo
                moscow_tz = ZoneInfo("Europe/Moscow")
                valid_until_dt = datetime.strptime(valid_until, "%Y-%m-%dT%H:%M")
                valid_until_dt = moscow_tz.localize(valid_until_dt).astimezone(ZoneInfo("UTC"))
            except Exception:
                pass
        
        promo = PromoCode(
            code=code,
            discount_percent=discount_percent,
            discount_amount_cents=discount_amount_cents,
            max_uses=max_uses_int,
            used_count=0,
            is_active=True,
            valid_from=valid_from_dt,
            valid_until=valid_until_dt,
            created_by_tg_id=admin_user.get("tg_id"),
            description=description or None,
        )
        session.add(promo)
        
        session.add(
            AuditLog(
                action=AuditLogAction.admin_action,
                admin_tg_id=admin_user.get("tg_id"),
                details=f"–°–æ–∑–¥–∞–Ω –ø—Ä–æ–º–æ–∫–æ–¥ {code}",
            )
        )
        
        await session.commit()
        return RedirectResponse(url="/admin/web/promo-codes?success=created", status_code=303)
    except Exception as e:
        import traceback
        error_msg = str(e)[:100]  # –û–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º –¥–ª–∏–Ω—É —Å–æ–æ–±—â–µ–Ω–∏—è –æ–± –æ—à–∏–±–∫–µ
        return RedirectResponse(url=f"/admin/web/promo-codes?error={error_msg}", status_code=303)


@app.get("/admin/web/api/promo-codes/{promo_id}")
async def admin_api_promo_code_detail(
    promo_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ü–æ–ª—É—á–∏—Ç—å –¥–µ—Ç–∞–ª—å–Ω—É—é –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø—Ä–æ–º–æ–∫–æ–¥–µ"""
    promo = await session.scalar(select(PromoCode).where(PromoCode.id == promo_id))
    if not promo:
        raise HTTPException(status_code=404, detail="promo_code_not_found")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–ø–∏—Å–æ–∫ –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–π
    usages_result = await session.scalars(
        select(PromoCodeUsage)
        .where(PromoCodeUsage.promo_code_id == promo_id)
        .order_by(PromoCodeUsage.used_at.desc())
        .limit(50)
    )
    usages = usages_result.all()
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è—Ö
    user_ids = [u.user_id for u in usages]
    users_map = {}
    if user_ids:
        users_result = await session.scalars(select(User).where(User.id.in_(user_ids)))
        for u in users_result.all():
            users_map[u.id] = u
    
    # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ —Å–æ–∑–¥–∞—Ç–µ–ª–µ
    creator = None
    if promo.created_by_tg_id:
        creator = await session.scalar(select(User).where(User.tg_id == promo.created_by_tg_id))
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    usages_data = []
    for u in usages:
        user = users_map.get(u.user_id)
        usages_data.append({
            "user_tg_id": user.tg_id if user else 0,
            "username": user.username if user else "‚Äî",
            "first_name": user.first_name if user else "‚Äî",
            "discount_amount": u.discount_amount_cents / 100,
            "used_at": fmt(u.used_at),
        })
    
    return {
        "id": promo.id,
        "code": promo.code,
        "discount_percent": promo.discount_percent,
        "discount_amount_cents": promo.discount_amount_cents,
        "max_uses": promo.max_uses,
        "used_count": promo.used_count,
        "is_active": promo.is_active,
        "valid_from": fmt(promo.valid_from) if promo.valid_from else None,
        "valid_until": fmt(promo.valid_until) if promo.valid_until else None,
        "created_at": fmt(promo.created_at),
        "created_by": {
            "tg_id": creator.tg_id if creator else None,
            "username": creator.username if creator else None,
            "first_name": creator.first_name if creator else None,
        } if creator else None,
        "description": promo.description,
        "usages": usages_data,
    }


@app.post("/admin/web/promo-codes/{promo_id}/toggle")
async def admin_web_toggle_promo_code(
    promo_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–ê–∫—Ç–∏–≤–∞—Ü–∏—è/–¥–µ–∞–∫—Ç–∏–≤–∞—Ü–∏—è –ø—Ä–æ–º–æ–∫–æ–¥–∞"""
    await _require_csrf(request)
    promo = await session.scalar(select(PromoCode).where(PromoCode.id == promo_id))
    if not promo:
        return RedirectResponse(url="/admin/web/promo-codes?error=not_found", status_code=303)
    
    promo.is_active = not promo.is_active
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–ü—Ä–æ–º–æ–∫–æ–¥ {promo.code} {'–∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω' if promo.is_active else '–¥–µ–∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω'}",
        )
    )
    await session.commit()
    return RedirectResponse(url="/admin/web/promo-codes?success=toggled", status_code=303)


@app.get("/admin/web/backups", response_class=HTMLResponse)
async def admin_web_backups(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Ä–µ–∑–µ—Ä–≤–Ω—ã–º–∏ –∫–æ–ø–∏—è–º–∏"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    # –ü–æ–ª—É—á–∞–µ–º —Ç–æ–ª—å–∫–æ –∑–∞–≤–µ—Ä—à–µ–Ω–Ω—ã–µ –±—ç–∫–∞–ø—ã (failed –∏ in_progress —Ç–æ–∂–µ –ø–æ–∫–∞–∑—ã–≤–∞–µ–º, –Ω–æ –æ–Ω–∏ –Ω–µ —É–¥–∞–ª—è—é—Ç—Å—è)
    backups_result = await session.scalars(
        select(Backup)
        .where(Backup.backup_type == "database")  # –§–∏–ª—å—Ç—Ä—É–µ–º —Ç–æ–ª—å–∫–æ database –±—ç–∫–∞–ø—ã
        .order_by(Backup.created_at.desc())
        .limit(50)
    )
    backups = backups_result.all()
    
    try:
        from zoneinfo import ZoneInfo
        moscow_tz = ZoneInfo("Europe/Moscow")
        fmt = lambda dt: dt.astimezone(moscow_tz).strftime("%d.%m.%Y %H:%M") if dt else "‚Äî"
    except Exception:
        fmt = lambda dt: dt.isoformat()[:16] if dt else "‚Äî"
    
    backups_data = []
    for backup in backups:
        file_size_mb = backup.file_size_bytes / (1024 * 1024) if backup.file_size_bytes > 0 else 0
        
        backups_data.append({
            "id": backup.id,
            "backup_type": backup.backup_type,
            "file_path": backup.file_path,
            "file_size_mb": file_size_mb,
            "status": backup.status,
            "error_message": backup.error_message,
            "created_at": fmt(backup.created_at),
            "created_by_tg_id": backup.created_by_tg_id,
        })
    
    return templates.TemplateResponse("backups.html", {
        "request": request,
        "admin_user": admin_user,
        "backups": backups_data,
        "csrf_token": _get_csrf_token(request),
    })


@app.post("/admin/web/backups/create")
async def admin_web_create_backup(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°–æ–∑–¥–∞–Ω–∏–µ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏ –≤—Ä—É—á–Ω—É—é"""
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º CSRF —Ç–æ–∫–µ–Ω –∏–∑ –∑–∞–≥–æ–ª–æ–≤–∫–∞ –∏–ª–∏ —Ñ–æ—Ä–º—ã
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    if not provided:
        # –ü—Ä–æ–±—É–µ–º –ø–æ–ª—É—á–∏—Ç—å –∏–∑ —Ñ–æ—Ä–º—ã
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        return RedirectResponse(url="/admin/web/backups?error=csrf_forbidden", status_code=303)
    
    # –°–æ–∑–¥–∞–µ–º –±—ç–∫–∞–ø –≤ —Ñ–æ–Ω–µ
    asyncio.create_task(_create_database_backup(created_by_tg_id=admin_user.get("tg_id")))
    
    # –õ–æ–≥–∏—Ä—É–µ–º –¥–µ–π—Å—Ç–≤–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.backup_action,
            admin_tg_id=admin_user.get("tg_id"),
            details="–ó–∞–ø—É—â–µ–Ω–æ —Å–æ–∑–¥–∞–Ω–∏–µ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏",
        )
    )
    await session.commit()
    
    return RedirectResponse(url="/admin/web/backups?success=backup_started", status_code=303)


@app.get("/admin/web/backups/{backup_id}/download")
async def admin_web_download_backup(
    backup_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°–∫–∞—á–∞—Ç—å —Ä–µ–∑–µ—Ä–≤–Ω—É—é –∫–æ–ø–∏—é"""
    backup = await session.scalar(select(Backup).where(Backup.id == backup_id))
    if not backup:
        raise HTTPException(status_code=404, detail="backup_not_found")
    
    if backup.status != "completed":
        raise HTTPException(status_code=400, detail="backup_not_ready")
    
    from pathlib import Path
    backup_path = Path(backup.file_path)
    
    if not backup_path.exists():
        raise HTTPException(status_code=404, detail="backup_file_not_found")
    
    # –õ–æ–≥–∏—Ä—É–µ–º —Å–∫–∞—á–∏–≤–∞–Ω–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.backup_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–°–∫–∞—á–∞–Ω –±—ç–∫–∞–ø #{backup_id}",
        )
    )
    await session.commit()
    
    def file_generator():
        with open(backup_path, "rb") as f:
            while True:
                chunk = f.read(8192)
                if not chunk:
                    break
                yield chunk
    
    return StreamingResponse(
        file_generator(),
        media_type="application/octet-stream",
        headers={
            "Content-Disposition": f'attachment; filename="{backup_path.name}"',
        },
    )


@app.post("/admin/web/backups/{backup_id}/delete")
async def admin_web_delete_backup(
    backup_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–£–¥–∞–ª–µ–Ω–∏–µ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏"""
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º CSRF —Ç–æ–∫–µ–Ω
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    if not provided:
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        return JSONResponse({"success": False, "error": "csrf_forbidden"}, status_code=403)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –±—ç–∫–∞–ø —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
    backup = await session.scalar(select(Backup).where(Backup.id == backup_id))
    if not backup:
        return JSONResponse({"success": False, "error": "backup_not_found"}, status_code=404)
    
    # –£–¥–∞–ª—è–µ–º —Ñ–∞–π–ª –±—ç–∫–∞–ø–∞, –µ—Å–ª–∏ –æ–Ω —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –∏ —ç—Ç–æ –≤–∞–ª–∏–¥–Ω—ã–π —Ñ–∞–π–ª
    from pathlib import Path
    if backup.file_path and backup.file_path.strip():
        backup_path = Path(backup.file_path)
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –ø—É—Ç—å –≤–∞–ª–∏–¥–Ω—ã–π, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –∏ —ç—Ç–æ —Ñ–∞–π–ª (–Ω–µ –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏—è)
        if backup_path.exists() and backup_path.is_file():
            try:
                backup_path.unlink()
                import logging
                logging.info(f"Deleted backup file: {backup_path}")
            except Exception as e:
                import logging
                logging.warning(f"Could not delete backup file {backup_path}: {e}")
        elif backup_path.exists() and backup_path.is_dir():
            import logging
            logging.warning(f"Backup path is a directory, not a file: {backup_path}")
        elif not backup_path.exists():
            import logging
            logging.info(f"Backup file does not exist (may have been deleted already): {backup_path}")
    
    # –£–¥–∞–ª—è–µ–º –∑–∞–ø–∏—Å—å –∏–∑ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö
    await session.delete(backup)
    
    # –õ–æ–≥–∏—Ä—É–µ–º —É–¥–∞–ª–µ–Ω–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.backup_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–£–¥–∞–ª–µ–Ω –±—ç–∫–∞–ø #{backup_id}",
        )
    )
    await session.commit()
    
    return JSONResponse({"success": True, "message": "Backup deleted successfully"})


async def _restore_database_backup(backup_id: int, restored_by_tg_id: int | None = None) -> dict:
    """–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö –∏–∑ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏"""
    import os
    import subprocess
    from pathlib import Path
    
    try:
        # –ü–æ–ª—É—á–∞–µ–º –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—é –æ –±—ç–∫–∞–ø–µ –î–û –∑–∞–∫—Ä—ã—Ç–∏—è —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–π
        async with SessionLocal() as session:
            backup = await session.scalar(select(Backup).where(Backup.id == backup_id))
            if not backup:
                return {"success": False, "error": "backup_not_found"}
            
            if backup.status != "completed":
                return {"success": False, "error": "backup_not_ready"}
            
            backup_path = Path(backup.file_path)
            if not backup_path.exists():
                return {"success": False, "error": "backup_file_not_found"}
        
        # –ó–∞–∫—Ä—ã–≤–∞–µ–º –≤—Å–µ –∞–∫—Ç–∏–≤–Ω—ã–µ —Å–æ–µ–¥–∏–Ω–µ–Ω–∏—è –ø–µ—Ä–µ–¥ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ–º
        # –≠—Ç–æ –≤–∞–∂–Ω–æ, —á—Ç–æ–±—ã –∏–∑–±–µ–∂–∞—Ç—å –∫–æ–Ω—Ñ–ª–∏–∫—Ç–æ–≤ —Å —Ç–∏–ø–∞–º–∏ –¥–∞–Ω–Ω—ã—Ö
        await engine.dispose()
        
        # –ù–µ–±–æ–ª—å—à–∞—è –∑–∞–¥–µ—Ä–∂–∫–∞ –¥–ª—è –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è –≤—Å–µ—Ö –æ–ø–µ—Ä–∞—Ü–∏–π
        await asyncio.sleep(1)
        
        # –ü–æ–ª—É—á–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ –ë–î
        db_url = settings.db_url
        if "postgresql" not in db_url:
            # –ü–µ—Ä–µ—Å–æ–∑–¥–∞–µ–º engine –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏
            await recreate_engine()
            return {"success": False, "error": "only_postgresql_supported"}
        
        # –ò–∑–≤–ª–µ–∫–∞–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –∏–∑ URL
        import re
        match = re.match(r"postgresql\+asyncpg://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)", db_url)
        if not match:
            # –ü–µ—Ä–µ—Å–æ–∑–¥–∞–µ–º engine –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏
            await recreate_engine()
            return {"success": False, "error": "invalid_db_url"}
        
        db_user, db_password, db_host, db_port, db_name = match.groups()
        
        env = os.environ.copy()
        env["PGPASSWORD"] = db_password
        
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º pg_restore –¥–ª—è –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è –∏–∑ custom format
        # --clean —É–¥–∞–ª–∏—Ç –∏ –ø–µ—Ä–µ—Å–æ–∑–¥–∞—Å—Ç –æ–±—ä–µ–∫—Ç—ã —Å—Ö–µ–º—ã, –∑–∞—Ç–µ–º –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–∏—Ç –¥–∞–Ω–Ω—ã–µ –≤ –ø—Ä–∞–≤–∏–ª—å–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ
        import logging
        
        # –û—Ç–∫–ª—é—á–∞–µ–º –ø—Ä–æ–≤–µ—Ä–∫—É –≤–Ω–µ—à–Ω–∏—Ö –∫–ª—é—á–µ–π –ø–µ—Ä–µ–¥ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ–º
        # –≠—Ç–æ –ø–æ–∑–≤–æ–ª–∏—Ç –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –¥–∞–Ω–Ω—ã–µ –≤ –ª—é–±–æ–º –ø–æ—Ä—è–¥–∫–µ
        try:
            disable_fk_cmd = [
                "psql",
                "-h", db_host,
                "-p", db_port,
                "-U", db_user,
                "-d", db_name,
                "-c", "SET session_replication_role = 'replica';"
            ]
            subprocess.run(disable_fk_cmd, env=env, capture_output=True, timeout=10)
        except Exception as e:
            logging.warning(f"Could not disable FK checks: {e}")
        
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º pg_restore —Å --clean –¥–ª—è –ø–æ–ª–Ω–æ–≥–æ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è
        # --clean —É–¥–∞–ª–∏—Ç –∏ –ø–µ—Ä–µ—Å–æ–∑–¥–∞—Å—Ç —Ç–∞–±–ª–∏—Ü—ã, –∑–∞—Ç–µ–º –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–∏—Ç –¥–∞–Ω–Ω—ã–µ –≤ –ø—Ä–∞–≤–∏–ª—å–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ
        cmd = [
            "pg_restore",
            "-h", db_host,
            "-p", db_port,
            "-U", db_user,
            "-d", db_name,
            "--clean",  # –û—á–∏—Å—Ç–∏—Ç—å –æ–±—ä–µ–∫—Ç—ã –ø–µ—Ä–µ–¥ —Å–æ–∑–¥–∞–Ω–∏–µ–º (—É–¥–∞–ª–∏—Ç —Ç–∞–±–ª–∏—Ü—ã –∏ –ø–µ—Ä–µ—Å–æ–∑–¥–∞—Å—Ç)
            "--if-exists",  # –ù–µ –≤—ã–¥–∞–≤–∞—Ç—å –æ—à–∏–±–∫—É –µ—Å–ª–∏ –æ–±—ä–µ–∫—Ç –Ω–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç
            "--no-owner",  # –ù–µ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—Ç—å –≤–ª–∞–¥–µ–ª—å—Ü–∞ –æ–±—ä–µ–∫—Ç–æ–≤
            "--no-privileges",  # –ù–µ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—Ç—å –ø—Ä–∏–≤–∏–ª–µ–≥–∏–∏
            "--verbose",  # –ü–æ–¥—Ä–æ–±–Ω—ã–π –≤—ã–≤–æ–¥ –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
            str(backup_path),
        ]
        
        result = subprocess.run(
            cmd,
            env=env,
            capture_output=True,
            text=True,
            timeout=600  # 10 –º–∏–Ω—É—Ç —Ç–∞–π–º–∞—É—Ç
        )
        
        # –í–∫–ª—é—á–∞–µ–º –ø—Ä–æ–≤–µ—Ä–∫—É –≤–Ω–µ—à–Ω–∏—Ö –∫–ª—é—á–µ–π –æ–±—Ä–∞—Ç–Ω–æ
        try:
            enable_fk_cmd = [
                "psql",
                "-h", db_host,
                "-p", db_port,
                "-U", db_user,
                "-d", db_name,
                "-c", "SET session_replication_role = 'origin';"
            ]
            subprocess.run(enable_fk_cmd, env=env, capture_output=True, timeout=10)
        except Exception as e:
            logging.warning(f"Could not enable FK checks: {e}")
        
        if result.returncode != 0:
            error_msg = result.stderr[:500] if result.stderr else result.stdout[:500] if result.stdout else "Unknown error"
            # –ü–µ—Ä–µ—Å–æ–∑–¥–∞–µ–º engine –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏
            await recreate_engine()
            return {"success": False, "error": f"pg_restore failed: {error_msg}"}
        
        # –ü–µ—Ä–µ—Å–æ–∑–¥–∞–µ–º engine –∏ SessionLocal –ø–æ—Å–ª–µ –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏—è
        # –≠—Ç–æ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏ –≤–∞–∂–Ω–æ –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –º–µ—Ç–∞–¥–∞–Ω–Ω—ã—Ö —Ç–∏–ø–æ–≤ –¥–∞–Ω–Ω—ã—Ö
        await recreate_engine()
        
        # –ù–µ–±–æ–ª—å—à–∞—è –∑–∞–¥–µ—Ä–∂–∫–∞ –¥–ª—è —Å—Ç–∞–±–∏–ª–∏–∑–∞—Ü–∏–∏ —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–π
        await asyncio.sleep(0.5)
        
        # –õ–æ–≥–∏—Ä—É–µ–º –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ —Å –Ω–æ–≤—ã–º —Å–æ–µ–¥–∏–Ω–µ–Ω–∏–µ–º
        # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ø—Ä—è–º–æ–π –¥–æ—Å—Ç—É–ø –∫ –º–æ–¥—É–ª—é –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –æ–±–Ω–æ–≤–ª–µ–Ω–Ω–æ–≥–æ SessionLocal
        import core.db.session as session_module
        async with session_module.SessionLocal() as session:
            session.add(
                AuditLog(
                    action=AuditLogAction.backup_action,
                    admin_tg_id=restored_by_tg_id,
                    details=f"–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∞ –±–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –∏–∑ –±—ç–∫–∞–ø–∞ #{backup_id}",
                )
            )
            await session.commit()
        
        return {"success": True, "message": "Database restored successfully. Please refresh the page."}
        
    except Exception as e:
        import logging
        logging.error(f"Error restoring backup: {e}")
        
        # –ü–µ—Ä–µ—Å–æ–∑–¥–∞–µ–º engine –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏
        try:
            await recreate_engine()
        except Exception as restore_error:
            logging.error(f"Error recreating engine: {restore_error}")
        
        return {"success": False, "error": str(e)[:500]}


@app.post("/admin/web/backups/{backup_id}/restore")
async def admin_web_restore_backup(
    backup_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö –∏–∑ —Ä–µ–∑–µ—Ä–≤–Ω–æ–π –∫–æ–ø–∏–∏"""
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º CSRF —Ç–æ–∫–µ–Ω
    expected = request.session.get("csrf_token")
    provided = request.headers.get("X-CSRF-Token")
    if not provided:
        try:
            form = await request.form()
            provided = form.get("csrf_token")
        except Exception:
            pass
    
    if not expected or not provided or provided != expected:
        raise HTTPException(status_code=403, detail="csrf_forbidden")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ –±—ç–∫–∞–ø —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –∏ –≥–æ—Ç–æ–≤
    backup = await session.scalar(select(Backup).where(Backup.id == backup_id))
    if not backup:
        raise HTTPException(status_code=404, detail="backup_not_found")
    
    if backup.status != "completed":
        raise HTTPException(status_code=400, detail="backup_not_ready")
    
    # –ó–∞–ø—É—Å–∫–∞–µ–º –≤–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ –≤ —Ñ–æ–Ω–µ (—ç—Ç–æ –¥–æ–ª–≥–∞—è –æ–ø–µ—Ä–∞—Ü–∏—è)
    asyncio.create_task(_restore_database_backup(backup_id, restored_by_tg_id=admin_user.get("tg_id")))
    
    return JSONResponse({"success": True, "message": "–í–æ—Å—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω–∏–µ –∑–∞–ø—É—â–µ–Ω–æ. –≠—Ç–æ –º–æ–∂–µ—Ç –∑–∞–Ω—è—Ç—å –Ω–µ—Å–∫–æ–ª—å–∫–æ –º–∏–Ω—É—Ç."})


async def _generate_vpn_config_for_user_server(user_id: int, server_id: int, session: AsyncSession, expires_at: datetime):
    """–ì–µ–Ω–µ—Ä–∏—Ä—É–µ—Ç VPN –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –Ω–∞ —É–∫–∞–∑–∞–Ω–Ω–æ–º —Å–µ—Ä–≤–µ—Ä–µ"""
    import json
    # –ü–æ–ª—É—á–∞–µ–º –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–π —Å–µ—Ä–≤–µ—Ä
    server = await session.scalar(
        select(Server)
        .where(Server.id == server_id)
        .where(Server.is_enabled == True)
        .where(
            (Server.x3ui_api_url.isnot(None)) | (Server.xray_uuid.isnot(None))
        )
    )
    
    if not server:
        raise ValueError(f"–°–µ—Ä–≤–µ—Ä {server_id} –Ω–µ –Ω–∞–π–¥–µ–Ω –∏–ª–∏ –Ω–µ –∞–∫—Ç–∏–≤–µ–Ω")
    
    user = await session.get(User, user_id)
    if not user:
        return
    
    # –°–æ–∑–¥–∞–µ–º –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ —É–∂–µ –∞–∫—Ç–∏–≤–Ω–æ–≥–æ –∫–æ–Ω—Ñ–∏–≥–∞ –¥–ª—è —ç—Ç–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
    existing = await session.scalar(
        select(VpnCredential)
        .where(VpnCredential.user_id == user_id)
        .where(VpnCredential.server_id == server.id)
        .where(VpnCredential.active == True)
    )
    
    config_text = None
    user_uuid = None
    
    # –ï—Å–ª–∏ —Å–µ—Ä–≤–µ—Ä –∏—Å–ø–æ–ª—å–∑—É–µ—Ç API 3x-UI - —Å–æ–∑–¥–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏
    if server.x3ui_api_url and server.x3ui_username and server.x3ui_password:
        from core.x3ui_api import X3UIAPI
        from core.xray import generate_uuid
        
        config_text = None
        user_uuid = None
        x3ui = None
        
        try:
            x3ui = X3UIAPI(
                api_url=server.x3ui_api_url,
                username=server.x3ui_username,
                password=server.x3ui_password,
            )
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º ID –∏–Ω–±–∞—É–Ω–¥–∞ (–ø—Ä–æ—Å—Ç–æ–π –ø–æ–¥—Ö–æ–¥ - —Ç—Ä–µ–±—É–µ–º —É–∫–∞–∑–∞–Ω–∏—è ID)
            inbound_id = server.x3ui_inbound_id
            
            if not inbound_id:
                # –ï—Å–ª–∏ ID –Ω–µ —É–∫–∞–∑–∞–Ω, –ø—Ä–æ–±—É–µ–º –Ω–∞–π—Ç–∏ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ (fallback)
                logger.warning(f"Inbound ID –Ω–µ —É–∫–∞–∑–∞–Ω –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}, –ø—ã—Ç–∞–µ–º—Å—è –Ω–∞–π—Ç–∏ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏")
                found_inbound = await x3ui.find_first_vless_inbound()
                
                if found_inbound:
                    inbound_id = found_inbound.get("id")
                    logger.info(f"–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –Ω–∞–π–¥–µ–Ω VLESS Inbound ID {inbound_id} –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
                    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –Ω–∞–π–¥–µ–Ω–Ω—ã–π ID –≤ –±–∞–∑—É –¥–∞–Ω–Ω—ã—Ö
                    server.x3ui_inbound_id = inbound_id
                    await session.commit()
                else:
                    # –õ–æ–≥–∏—Ä—É–µ–º –≤—Å–µ –¥–æ—Å—Ç—É–ø–Ω—ã–µ Inbounds –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
                    all_inbounds = await x3ui.list_inbounds()
                    if all_inbounds:
                        inbound_list = []
                        for inb in all_inbounds:
                            inbound_list.append({
                                "id": inb.get("id"),
                                "port": inb.get("port"),
                                "protocol": inb.get("protocol"),
                                "remark": inb.get("remark", ""),
                                "enable": inb.get("enable", False)
                            })
                        logger.error(f"–î–æ—Å—Ç—É–ø–Ω—ã–µ Inbounds –≤ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {inbound_list}")
                        inbound_info = ", ".join([f"ID:{inb['id']} ({inb['protocol']}, –ø–æ—Ä—Ç:{inb['port']})" for inb in inbound_list])
                        raise ValueError(
                            f"Inbound ID –Ω–µ —É–∫–∞–∑–∞–Ω –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} –∏ –Ω–µ –Ω–∞–π–¥–µ–Ω VLESS Inbound –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏. "
                            f"–î–æ—Å—Ç—É–ø–Ω—ã–µ Inbounds: {inbound_info}. "
                            f"–£–∫–∞–∂–∏—Ç–µ –ø—Ä–∞–≤–∏–ª—å–Ω—ã–π Inbound ID –≤ –Ω–∞—Å—Ç—Ä–æ–π–∫–∞—Ö —Å–µ—Ä–≤–µ—Ä–∞."
                        )
                    else:
                        logger.error(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ Inbounds –∏–∑ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
                        raise ValueError(
                            f"Inbound ID –Ω–µ —É–∫–∞–∑–∞–Ω –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} –∏ –Ω–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ Inbounds. "
                            f"–ü—Ä–æ–≤–µ—Ä—å—Ç–µ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ API 3x-UI (URL: {server.x3ui_api_url}, username: {server.x3ui_username})."
                        )
            
            # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º —É–Ω–∏–∫–∞–ª—å–Ω—ã–π email –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ —Å tg_id
            client_email = f"tg_{user.tg_id}_server_{server.id}@fiorevpn"
            
            # –í–ê–ñ–ù–û: –£–¥–∞–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–≥–æ –∫–ª–∏–µ–Ω—Ç–∞ –ø–µ—Ä–µ–¥ —Å–æ–∑–¥–∞–Ω–∏–µ–º –Ω–æ–≤–æ–≥–æ (–¥–ª—è regenerate –∏ duplicate email)
            # –ò–≥–Ω–æ—Ä–∏—Ä—É–µ–º –æ—à–∏–±–∫–∏, —Ç–∞–∫ –∫–∞–∫ –∫–ª–∏–µ–Ω—Ç –º–æ–∂–µ—Ç –Ω–µ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞—Ç—å (–Ω–∞–ø—Ä–∏–º–µ—Ä, –ø—Ä–∏ –ø–µ—Ä–≤–æ–π –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏)
            try:
                deleted = await x3ui.delete_client(inbound_id, client_email)
                if deleted:
                    logger.info(f"–£–¥–∞–ª–µ–Ω —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π –∫–ª–∏–µ–Ω—Ç {client_email} –∏–∑ Inbound {inbound_id}")
                else:
                    logger.debug(f"–ö–ª–∏–µ–Ω—Ç {client_email} –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ Inbound {inbound_id} (—ç—Ç–æ –Ω–æ—Ä–º–∞–ª—å–Ω–æ –¥–ª—è –Ω–æ–≤–æ–π –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏)")
            except Exception as del_err:
                # –ù–µ –∫—Ä–∏—Ç–∏—á–Ω–∞—è –æ—à–∏–±–∫–∞ - –∫–ª–∏–µ–Ω—Ç –º–æ–∂–µ—Ç –Ω–µ —Å—É—â–µ—Å—Ç–≤–æ–≤–∞—Ç—å
                logger.debug(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –∫–ª–∏–µ–Ω—Ç–∞ {client_email} (–≤–æ–∑–º–æ–∂–Ω–æ, –µ–≥–æ –Ω–µ—Ç): {del_err}")
            
            # –ü–æ–ª—É—á–∞–µ–º –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ –ª–∏–º–∏—Ç–æ–≤ –∏–∑ SystemSetting
            limit_ip_setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == "vpn_limit_ip"))
            limit_traffic_setting = await session.scalar(select(SystemSetting).where(SystemSetting.key == "vpn_limit_traffic_gb"))
            
            limit_ip = 1  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é 1 IP
            if limit_ip_setting:
                try:
                    limit_ip = int(limit_ip_setting.value)
                except (ValueError, TypeError):
                    limit_ip = 1
            
            total_gb = 0  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é –±–µ–∑ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π
            if limit_traffic_setting:
                try:
                    total_gb = int(float(limit_traffic_setting.value))
                except (ValueError, TypeError):
                    total_gb = 0
            
            # –°–æ–∑–¥–∞–µ–º –∫–ª–∏–µ–Ω—Ç–∞ –≤ 3x-UI
            expire_timestamp = int(expires_at.timestamp() * 1000) if expires_at else 0  # 3x-UI –∏—Å–ø–æ–ª—å–∑—É–µ—Ç –º–∏–ª–ª–∏—Å–µ–∫—É–Ω–¥—ã
            logger.info(
                f"–°–æ–∑–¥–∞–Ω–∏–µ –∫–ª–∏–µ–Ω—Ç–∞ —á–µ—Ä–µ–∑ 3x-UI API –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user_id} (tg_id: {user.tg_id}) –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ {server.name} (ID: {server.id}): "
                f"API URL={server.x3ui_api_url}, Inbound ID={inbound_id}, email={client_email}"
            )
            
            try:
                client_data = await x3ui.add_client(
                    inbound_id=inbound_id,
                    email=client_email,
                    uuid=None,  # –ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∞—è –≥–µ–Ω–µ—Ä–∞—Ü–∏—è
                    flow=server.xray_flow or "",
                    expire=expire_timestamp,
                    limit_ip=limit_ip,
                    total_gb=total_gb,  # –õ–∏–º–∏—Ç —Ç—Ä–∞—Ñ–∏–∫–∞ –∏–∑ –Ω–∞—Å—Ç—Ä–æ–µ–∫ (0 = –±–µ–∑ –æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–π)
                )
            except ConnectionError as e:
                # –°–ø–µ—Ü–∏–∞–ª—å–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –æ—à–∏–±–æ–∫ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è
                error_msg = str(e)
                if "localhost" in server.x3ui_api_url or "127.0.0.1" in server.x3ui_api_url:
                    error_msg = f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ 3x-UI —á–µ—Ä–µ–∑ SSH-—Ç—É–Ω–Ω–µ–ª—å. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ, —á—Ç–æ SSH-—Ç—É–Ω–Ω–µ–ª—å –∑–∞–ø—É—â–µ–Ω –∏ –ø–æ—Ä—Ç –¥–æ—Å—Ç—É–ø–µ–Ω. –û—à–∏–±–∫–∞: {e}"
                logger.error(f"–û—à–∏–±–∫–∞ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {error_msg}")
                raise ValueError(error_msg) from e
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫–ª–∏–µ–Ω—Ç–∞ –≤ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {e}")
                raise
            
            if client_data:
                # –ü–æ–ª—É—á–∞–µ–º UUID —Å–æ–∑–¥–∞–Ω–Ω–æ–≥–æ –∫–ª–∏–µ–Ω—Ç–∞
                user_uuid = client_data.get("id") or client_data.get("uuid")
                
                if not user_uuid:
                    logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å UUID –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ {client_email} –∏–∑ –æ—Ç–≤–µ—Ç–∞ API 3x-UI")
                    raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å UUID –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ {server.name}")
                
                # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –∫–æ–Ω—Ñ–∏–≥ –Ω–∞–ø—Ä—è–º—É—é –∏–∑ –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤ —Å–µ—Ä–≤–µ—Ä–∞ (–∫–∞–∫ –≤ –ø—Ä–∏–º–µ—Ä–µ ChatGPT)
                from core.xray import generate_vless_config
                config_text = generate_vless_config(
                    user_uuid=user_uuid,
                    server_host=server.host,
                    server_port=server.xray_port or 443,
                    server_uuid=user_uuid,  # –ò—Å–ø–æ–ª—å–∑—É–µ–º UUID –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
                    server_flow=server.xray_flow,
                    server_network=server.xray_network or "tcp",
                    server_security=server.xray_security or "tls",
                    server_sni=server.xray_sni,
                    server_reality_public_key=server.xray_reality_public_key,
                    server_reality_short_id=server.xray_reality_short_id,
                    server_path=server.xray_path,
                    server_host_header=server.xray_host,
                    remark=f"{server.name}",
                )
                
                if not config_text:
                    logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ {client_email} –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ {server.name}")
                    raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è –∫–ª–∏–µ–Ω—Ç–∞ –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ {server.name}")
            else:
                # –ö–ª–∏–µ–Ω—Ç –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω (–∏–Ω–±–∞—É–Ω–¥ –Ω–µ –Ω–∞–π–¥–µ–Ω –∏–ª–∏ –¥—Ä—É–≥–∞—è –æ—à–∏–±–∫–∞)
                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –∫–ª–∏–µ–Ω—Ç–∞ —á–µ—Ä–µ–∑ API 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} (ID: {server.id}, Inbound ID: {server.x3ui_inbound_id})")
                # –ï—Å–ª–∏ –µ—Å—Ç—å UUID, –∏—Å–ø–æ–ª—å–∑—É–µ–º fallback
                if server.xray_uuid:
                    logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º fallback –Ω–∞ UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} (–∫–ª–∏–µ–Ω—Ç –Ω–µ —Å–æ–∑–¥–∞–Ω)")
                    raise ValueError("INBOUND_NOT_FOUND_FALLBACK_TO_UUID")
                raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –∫–ª–∏–µ–Ω—Ç–∞ –Ω–∞ —Å–µ—Ä–≤–µ—Ä–µ {server.name}")
        except ValueError as e:
            error_msg = str(e)
            # –ï—Å–ª–∏ —ç—Ç–æ —Å–ø–µ—Ü–∏–∞–ª—å–Ω—ã–π —Å–∏–≥–Ω–∞–ª –¥–ª—è fallback –Ω–∞ UUID
            if error_msg == "INBOUND_NOT_FOUND_FALLBACK_TO_UUID" and server.xray_uuid:
                logger.info(f"–ü–µ—Ä–µ—Ö–æ–¥–∏–º –Ω–∞ fallback —Å UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
                # –ù–µ —É—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º config_text –∏ user_uuid, —á—Ç–æ–±—ã –∫–æ–¥ –ø–µ—Ä–µ—à–µ–ª –∫ elif server.xray_uuid
                config_text = None
                user_uuid = None
            elif "Inbound –Ω–µ –Ω–∞–π–¥–µ–Ω" in error_msg and server.xray_uuid:
                # –ï—Å–ª–∏ Inbound –Ω–µ –Ω–∞–π–¥–µ–Ω, –Ω–æ –µ—Å—Ç—å UUID, –∏—Å–ø–æ–ª—å–∑—É–µ–º fallback
                logger.info(f"–ü–µ—Ä–µ—Ö–æ–¥–∏–º –Ω–∞ fallback —Å UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} (Inbound –Ω–µ –Ω–∞–π–¥–µ–Ω)")
                config_text = None
                user_uuid = None
            else:
                logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫–ª–∏–µ–Ω—Ç–∞ —á–µ—Ä–µ–∑ API 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {e}")
                raise
        except ConnectionError as e:
            # –û—à–∏–±–∫–∞ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è (–Ω–∞–ø—Ä–∏–º–µ—Ä, SSH-—Ç—É–Ω–Ω–µ–ª—å –Ω–µ —Ä–∞–±–æ—Ç–∞–µ—Ç)
            error_msg = str(e)
            logger.error(f"–û—à–∏–±–∫–∞ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è –∫ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {error_msg}")
            # –ï—Å–ª–∏ –µ—Å—Ç—å UUID, –∏—Å–ø–æ–ª—å–∑—É–µ–º fallback
            if server.xray_uuid:
                logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º fallback –Ω–∞ UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏—è")
                config_text = None
                user_uuid = None
            else:
                # –ü—Ä–æ–±—Ä–∞—Å—ã–≤–∞–µ–º –æ—à–∏–±–∫—É —Å –ø–æ–Ω—è—Ç–Ω—ã–º —Å–æ–æ–±—â–µ–Ω–∏–µ–º
                raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}. {error_msg}")
        except Exception as e:
            import httpx
            # –ï—Å–ª–∏ —ç—Ç–æ HTTP –æ—à–∏–±–∫–∞ –æ—Ç 3x-UI API, –ø—Ä–æ–±—Ä–∞—Å—ã–≤–∞–µ–º –µ—ë –¥–∞–ª—å—à–µ –¥–ª—è –ø—Ä–∞–≤–∏–ª—å–Ω–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–∏
            if isinstance(e, (httpx.HTTPStatusError, httpx.RequestError)):
                logger.warning(f"HTTP –æ—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫–ª–∏–µ–Ω—Ç–∞ —á–µ—Ä–µ–∑ API 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {e}")
                raise
            # –ï—Å–ª–∏ –µ—Å—Ç—å UUID, –ø—ã—Ç–∞–µ–º—Å—è –∏—Å–ø–æ–ª—å–∑–æ–≤–∞—Ç—å fallback
            if server.xray_uuid:
                logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º fallback –Ω–∞ UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name} –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏ API")
                config_text = None
                user_uuid = None
            else:
                logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –∫–ª–∏–µ–Ω—Ç–∞ —á–µ—Ä–µ–∑ API 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}: {e}")
                raise
        finally:
            # –ó–∞–∫—Ä—ã–≤–∞–µ–º —Å–µ—Å—Å–∏—é 3x-UI API
            if x3ui:
                try:
                    await x3ui.close()
                except:
                    pass
        
        # –ï—Å–ª–∏ –∫–æ–Ω—Ñ–∏–≥ –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω —á–µ—Ä–µ–∑ API, –Ω–æ –µ—Å—Ç—å UUID, –∏—Å–ø–æ–ª—å–∑—É–µ–º fallback
        if not config_text or not user_uuid:
            if server.xray_uuid:
                logger.info(f"–ò—Å–ø–æ–ª—å–∑—É–µ–º fallback –Ω–∞ UUID –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
                # –ü–µ—Ä–µ—Ö–æ–¥–∏–º –∫ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ —á–µ—Ä–µ–∑ UUID
                pass
            else:
                raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
    
    # –ï—Å–ª–∏ API 3x-UI –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω –∏–ª–∏ –Ω–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å —á–µ—Ä–µ–∑ API, –∏—Å–ø–æ–ª—å–∑—É–µ–º —Å—Ç–∞—Ä—ã–π —Å–ø–æ—Å–æ–± —Å UUID
    if (not config_text or not user_uuid) and server.xray_uuid:
        user_uuid = server.xray_uuid
        config_text = generate_vless_config(
            user_uuid=user_uuid,
            server_host=server.host,
            server_port=server.xray_port or 443,
            server_uuid=server.xray_uuid,
            server_flow=server.xray_flow,
            server_network=server.xray_network or "tcp",
            server_security=server.xray_security or "tls",
            server_sni=server.xray_sni,
            server_reality_public_key=server.xray_reality_public_key,
            server_reality_short_id=server.xray_reality_short_id,
            server_path=server.xray_path,
            server_host_header=server.xray_host,
            remark=f"{server.name}",
        )
    elif not config_text or not user_uuid:
        # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º —Å–µ—Ä–≤–µ—Ä—ã –±–µ–∑ –Ω–∞—Å—Ç—Ä–æ–µ–∫
        raise ValueError(f"–°–µ—Ä–≤–µ—Ä {server.name} –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω (–Ω–µ—Ç API 3x-UI –∏ UUID)")
        
    if not config_text:
        raise ValueError(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –∫–æ–Ω—Ñ–∏–≥ –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server.name}")
    
    if existing:
        # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π –∫–æ–Ω—Ñ–∏–≥
        existing.expires_at = expires_at
        existing.config_text = config_text
        if user_uuid:
            existing.user_uuid = user_uuid
    else:
        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π –∫–æ–Ω—Ñ–∏–≥
        credential = VpnCredential(
            user_id=user_id,
            server_id=server.id,
            user_uuid=user_uuid,
            config_text=config_text,
            active=True,
            expires_at=expires_at,
        )
        session.add(credential)
    
    await session.commit()


# API endpoints –¥–ª—è —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Å–µ—Ä–≤–µ—Ä–∞–º–∏
@app.get("/admin/web/servers", response_class=HTMLResponse)
async def admin_web_servers(
    request: Request,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """–°—Ç—Ä–∞–Ω–∏—Ü–∞ —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è —Å–µ—Ä–≤–µ—Ä–∞–º–∏"""
    if not templates:
        return HTMLResponse(content="<h1>–®–∞–±–ª–æ–Ω—ã –Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω—ã</h1>", status_code=500)
    
    servers_result = await session.scalars(
        select(Server)
        .order_by(Server.created_at.desc())
    )
    servers = servers_result.all()
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–µ —Å—Ç–∞—Ç—É—Å—ã —Å–µ—Ä–≤–µ—Ä–æ–≤
    servers_with_status = []
    for server in servers:
        last_status = await session.scalar(
            select(ServerStatus)
            .where(ServerStatus.server_id == server.id)
            .order_by(ServerStatus.checked_at.desc())
        )
        servers_with_status.append({
            "server": server,
            "status": last_status,
        })
    
    csrf_token = _get_csrf_token(request)
    return templates.TemplateResponse(
        "servers.html",
        {
            "request": request,
            "admin_user": admin_user,
            "servers": servers_with_status,
            "csrf_token": csrf_token,
        },
    )


@app.get("/admin/web/api/servers")
async def admin_api_servers(
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ —Å–µ—Ä–≤–µ—Ä–æ–≤"""
    servers_result = await session.scalars(
        select(Server)
        .order_by(Server.created_at.desc())
    )
    servers = servers_result.all()
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–µ —Å—Ç–∞—Ç—É—Å—ã
    servers_list = []
    for server in servers:
        last_status = await session.scalar(
            select(ServerStatus)
            .where(ServerStatus.server_id == server.id)
            .order_by(ServerStatus.checked_at.desc())
        )
        server_dict = {
            "id": server.id,
            "name": server.name,
            "host": server.host,
            "location": server.location,
            "is_enabled": server.is_enabled,
            "capacity": server.capacity,
            "created_at": server.created_at.isoformat(),
            "xray_port": server.xray_port,
            "xray_uuid": server.xray_uuid,
            "xray_flow": server.xray_flow,
            "xray_network": server.xray_network,
            "xray_security": server.xray_security,
            "xray_sni": server.xray_sni,
            "xray_reality_public_key": server.xray_reality_public_key,
            "xray_reality_short_id": server.xray_reality_short_id,
            "xray_path": server.xray_path,
            "xray_host": server.xray_host,
            "x3ui_api_url": server.x3ui_api_url,
            "x3ui_username": server.x3ui_username,
            "x3ui_password": server.x3ui_password,
            "x3ui_inbound_id": server.x3ui_inbound_id,
        }
        if last_status:
            # –í—Ä–µ–º—è —É–∂–µ –≤ UTC, –Ω–∞ –∫–ª–∏–µ–Ω—Ç–µ –¥–æ–±–∞–≤–∏–º +3 —á–∞—Å–∞ —á–µ—Ä–µ–∑ JavaScript
            server_dict["status"] = {
                "is_online": last_status.is_online,
                "response_time_ms": last_status.response_time_ms,
                "connection_speed_mbps": float(last_status.connection_speed_mbps) if last_status.connection_speed_mbps else None,
                "checked_at": last_status.checked_at.isoformat() if last_status.checked_at else None,
            }
        servers_list.append(server_dict)
    
    return {"servers": servers_list}


@app.get("/admin/web/api/servers/{server_id}/inbounds")
async def admin_api_get_server_inbounds(
    server_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ Inbounds –∏–∑ 3x-UI –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞"""
    server = await session.get(Server, server_id)
    if not server:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    if not server.x3ui_api_url or not server.x3ui_username or not server.x3ui_password:
        raise HTTPException(status_code=400, detail="3x_ui_api_not_configured")
    
    x3ui = None
    try:
        from core.x3ui_api import X3UIAPI
        x3ui = X3UIAPI(
            api_url=server.x3ui_api_url,
            username=server.x3ui_username,
            password=server.x3ui_password,
        )
        
        inbounds = await x3ui.list_inbounds()
        if not inbounds:
            return {"inbounds": [], "error": "–ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ Inbounds. –ü—Ä–æ–≤–µ—Ä—å—Ç–µ –Ω–∞—Å—Ç—Ä–æ–π–∫–∏ API 3x-UI."}
        
        # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º —Å–ø–∏—Å–æ–∫ Inbounds
        inbounds_list = []
        for inbound in inbounds:
            inbounds_list.append({
                "id": inbound.get("id"),
                "port": inbound.get("port"),
                "protocol": inbound.get("protocol", "").lower(),
                "remark": inbound.get("remark", ""),
                "enable": inbound.get("enable", False),
                "is_vless": inbound.get("protocol", "").lower() == "vless",
            })
        
        return {"inbounds": inbounds_list}
    except Exception as e:
        import logging
        logging.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ —Å–ø–∏—Å–∫–∞ Inbounds –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {server_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–ª—É—á–µ–Ω–∏–∏ —Å–ø–∏—Å–∫–∞ Inbounds: {str(e)}")
    finally:
        if x3ui:
            try:
                await x3ui.close()
            except:
                pass


@app.post("/admin/web/api/servers")
async def admin_api_create_server(
    payload: ServerCreateIn,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –°–æ–∑–¥–∞—Ç—å —Å–µ—Ä–≤–µ—Ä"""
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å –∏–º–µ–Ω–∏
    existing = await session.scalar(select(Server).where(Server.name == payload.name))
    if existing:
        raise HTTPException(status_code=400, detail="server_name_exists")
    
    server = Server(
        name=payload.name,
        host=payload.host,
        location=payload.location,
        is_enabled=payload.is_enabled,
        capacity=payload.capacity,
        xray_port=payload.xray_port,
        xray_uuid=payload.xray_uuid,
        xray_flow=payload.xray_flow,
        xray_network=payload.xray_network,
        xray_security=payload.xray_security,
        xray_sni=payload.xray_sni,
        xray_reality_public_key=payload.xray_reality_public_key,
        xray_reality_short_id=payload.xray_reality_short_id,
        xray_path=payload.xray_path,
        xray_host=payload.xray_host,
        x3ui_api_url=payload.x3ui_api_url,
        x3ui_username=payload.x3ui_username,
        x3ui_password=payload.x3ui_password,
        x3ui_inbound_id=payload.x3ui_inbound_id,
    )
    session.add(server)
    await session.commit()
    await session.refresh(server)
    
    # –õ–æ–≥–∏—Ä—É–µ–º —Å–æ–∑–¥–∞–Ω–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–°–æ–∑–¥–∞–Ω —Å–µ—Ä–≤–µ—Ä: {server.name} ({server.host})",
        )
    )
    await session.commit()
    
    return {"id": server.id, "name": server.name}


@app.put("/admin/web/api/servers/{server_id}")
async def admin_api_update_server(
    server_id: int,
    payload: ServerUpdateIn,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –û–±–Ω–æ–≤–∏—Ç—å —Å–µ—Ä–≤–µ—Ä"""
    server = await session.get(Server, server_id)
    if not server:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    # –û–±–Ω–æ–≤–ª—è–µ–º –ø–æ–ª—è
    if payload.name is not None:
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —É–Ω–∏–∫–∞–ª—å–Ω–æ—Å—Ç—å –∏–º–µ–Ω–∏
        existing = await session.scalar(select(Server).where(Server.name == payload.name).where(Server.id != server_id))
        if existing:
            raise HTTPException(status_code=400, detail="server_name_exists")
        server.name = payload.name
    
    if payload.host is not None:
        server.host = payload.host
    if payload.location is not None:
        server.location = payload.location
    if payload.is_enabled is not None:
        server.is_enabled = payload.is_enabled
    if payload.capacity is not None:
        server.capacity = payload.capacity
    if payload.xray_port is not None:
        server.xray_port = payload.xray_port
    if payload.xray_uuid is not None:
        server.xray_uuid = payload.xray_uuid
    if payload.xray_flow is not None:
        server.xray_flow = payload.xray_flow
    if payload.xray_network is not None:
        server.xray_network = payload.xray_network
    if payload.xray_security is not None:
        server.xray_security = payload.xray_security
    if payload.xray_sni is not None:
        server.xray_sni = payload.xray_sni
    if payload.xray_reality_public_key is not None:
        server.xray_reality_public_key = payload.xray_reality_public_key
    if payload.xray_reality_short_id is not None:
        server.xray_reality_short_id = payload.xray_reality_short_id
    if payload.xray_path is not None:
        server.xray_path = payload.xray_path
    if payload.xray_host is not None:
        server.xray_host = payload.xray_host
    if payload.x3ui_api_url is not None:
        server.x3ui_api_url = payload.x3ui_api_url
    if payload.x3ui_username is not None:
        server.x3ui_username = payload.x3ui_username
    if payload.x3ui_password is not None:
        server.x3ui_password = payload.x3ui_password
    if payload.x3ui_inbound_id is not None:
        server.x3ui_inbound_id = payload.x3ui_inbound_id
    
    await session.commit()
    
    # –õ–æ–≥–∏—Ä—É–µ–º –æ–±–Ω–æ–≤–ª–µ–Ω–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–û–±–Ω–æ–≤–ª–µ–Ω —Å–µ—Ä–≤–µ—Ä: {server.name} (ID: {server_id})",
        )
    )
    await session.commit()
    
    return {"id": server.id, "name": server.name}


@app.delete("/admin/web/api/servers/{server_id}")
async def admin_api_delete_server(
    server_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –£–¥–∞–ª–∏—Ç—å —Å–µ—Ä–≤–µ—Ä"""
    server = await session.get(Server, server_id)
    if not server:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    server_name = server.name
    await session.delete(server)
    await session.commit()
    
    # –õ–æ–≥–∏—Ä—É–µ–º —É–¥–∞–ª–µ–Ω–∏–µ
    session.add(
        AuditLog(
            action=AuditLogAction.admin_action,
            admin_tg_id=admin_user.get("tg_id"),
            details=f"–£–¥–∞–ª–µ–Ω —Å–µ—Ä–≤–µ—Ä: {server_name} (ID: {server_id})",
        )
    )
    await session.commit()
    
    return {"success": True}


@app.post("/admin/web/api/servers/{server_id}/check")
async def admin_api_check_server(
    server_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
):
    """API: –ü—Ä–æ–≤–µ—Ä–∏—Ç—å —Å–æ—Å—Ç–æ—è–Ω–∏–µ —Å–µ—Ä–≤–µ—Ä–∞ –≤—Ä—É—á–Ω—É—é"""
    try:
        server = await session.get(Server, server_id)
        if not server:
            raise HTTPException(status_code=404, detail="server_not_found")
        
        logger.info(f"–†—É—á–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ —Å–µ—Ä–≤–µ—Ä–∞ {server.name} (ID: {server_id})")
        
        # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å–æ—Å—Ç–æ—è–Ω–∏–µ (–∏—Å–ø–æ–ª—å–∑—É–µ–º —Ç–æ—Ç –∂–µ –º–µ—Ç–æ–¥, —á—Ç–æ –∏ –∞–≤—Ç–æ–ø—Ä–æ–≤–µ—Ä–∫–∞)
        status_result = await _check_server_status(server)
        
        logger.info(f"–†–µ–∑—É–ª—å—Ç–∞—Ç —Ä—É—á–Ω–æ–π –ø—Ä–æ–≤–µ—Ä–∫–∏ {server.name}: online={status_result['is_online']}, time={status_result.get('response_time_ms')}ms")
        
        # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Å—Ç–∞—Ç—É—Å
        status = ServerStatus(
            server_id=server.id,
            is_online=status_result["is_online"],
            response_time_ms=status_result["response_time_ms"],
            error_message=status_result["error_message"],
        )
        session.add(status)
        await session.commit()
        
        # –í—Ä–µ–º—è —É–∂–µ –≤ UTC, –Ω–∞ –∫–ª–∏–µ–Ω—Ç–µ –¥–æ–±–∞–≤–∏–º +3 —á–∞—Å–∞ —á–µ—Ä–µ–∑ JavaScript
        return {
            "server_id": server.id,
            "server_name": server.name,
            "status": {
                "is_online": status_result["is_online"],
                "response_time_ms": status_result["response_time_ms"],
                "error_message": status_result["error_message"],
                "checked_at": status.checked_at.isoformat() if status.checked_at else None,
            }
        }
    except Exception as e:
        logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Ä—É—á–Ω–æ–π –ø—Ä–æ–≤–µ—Ä–∫–µ —Å–µ—Ä–≤–µ—Ä–∞ {server_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"–û—à–∏–±–∫–∞ –ø—Ä–æ–≤–µ—Ä–∫–∏: {str(e)}")


@app.get("/admin/web/api/servers/{server_id}/history")
async def admin_api_server_history(
    server_id: int,
    session: AsyncSession = Depends(get_session),
    admin_user: dict = Depends(_require_web_admin),
    limit: int = 100,
):
    """API: –ü–æ–ª—É—á–∏—Ç—å –∏—Å—Ç–æ—Ä–∏—é —Å—Ç–∞—Ç—É—Å–æ–≤ —Å–µ—Ä–≤–µ—Ä–∞ –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–∞"""
    server = await session.get(Server, server_id)
    if not server:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–µ N –∑–∞–ø–∏—Å–µ–π —Å—Ç–∞—Ç—É—Å–æ–≤
    statuses_result = await session.scalars(
        select(ServerStatus)
        .where(ServerStatus.server_id == server.id)
        .order_by(ServerStatus.checked_at.desc())
        .limit(limit)
    )
    statuses = statuses_result.all()
    
    # –§–æ—Ä–º–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è –≥—Ä–∞—Ñ–∏–∫–∞ (–≤ –æ–±—Ä–∞—Ç–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ - –æ—Ç —Å—Ç–∞—Ä—ã—Ö –∫ –Ω–æ–≤—ã–º)
    history = []
    for status in reversed(statuses):
        history.append({
            "checked_at": status.checked_at.isoformat() if status.checked_at else None,
            "is_online": status.is_online,
            "response_time_ms": status.response_time_ms,
            "error_message": status.error_message,
        })
    
    return {
        "server_id": server.id,
        "server_name": server.name,
        "history": history,
    }


@app.get("/servers/available")
async def get_available_servers(
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ –¥–æ—Å—Ç—É–ø–Ω—ã—Ö —Å–µ—Ä–≤–µ—Ä–æ–≤ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–π"""
    servers_result = await session.scalars(
        select(Server)
        .where(Server.is_enabled == True)
        .where(
            (Server.x3ui_api_url.isnot(None)) | (Server.xray_uuid.isnot(None))
        )  # –°–µ—Ä–≤–µ—Ä –¥–æ–ª–∂–µ–Ω –∏–º–µ—Ç—å –ª–∏–±–æ API 3x-UI, –ª–∏–±–æ UUID
        .order_by(Server.name)
    )
    servers = servers_result.all()
    
    # –ü–æ–ª—É—á–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω–∏–µ —Å—Ç–∞—Ç—É—Å—ã
    servers_list = []
    for server in servers:
        last_status = await session.scalar(
            select(ServerStatus)
            .where(ServerStatus.server_id == server.id)
            .order_by(ServerStatus.checked_at.desc())
        )
        server_dict = {
            "id": server.id,
            "name": server.name,
            "host": server.host,
            "location": server.location,
            "capacity": server.capacity,
        }
        if last_status:
            server_dict["status"] = {
                "is_online": last_status.is_online,
                "response_time_ms": last_status.response_time_ms,
            }
        else:
            server_dict["status"] = {
                "is_online": False,
                "response_time_ms": None,
            }
        servers_list.append(server_dict)
    
    return {"servers": servers_list}


@app.post("/users/{tg_id}/select-server")
async def select_server_for_user(
    tg_id: int,
    payload: dict,
    session: AsyncSession = Depends(get_session),
):
    """–£—Å—Ç–∞–Ω–æ–≤–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Å–µ—Ä–≤–µ—Ä –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    server_id = payload.get("server_id")
    if not server_id:
        raise HTTPException(status_code=400, detail="server_id_required")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —á—Ç–æ —Å–µ—Ä–≤–µ—Ä —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –∏ –∞–∫—Ç–∏–≤–µ–Ω
    server = await session.scalar(
        select(Server)
        .where(Server.id == server_id)
        .where(Server.is_enabled == True)
    )
    if not server:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    old_server_id = user.selected_server_id
    
    # –ï—Å–ª–∏ –º–µ–Ω—è–µ–º —Å–µ—Ä–≤–µ—Ä ‚Äî —É–¥–∞–ª—è–µ–º –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
    # –í–ê–ñ–ù–û: –≠—Ç–æ –Ω–µ –¥–æ–ª–∂–Ω–æ –±–ª–æ–∫–∏—Ä–æ–≤–∞—Ç—å —Å–º–µ–Ω—É —Å–µ—Ä–≤–µ—Ä–∞, –ø–æ—ç—Ç–æ–º—É –≤—Å–µ –æ—à–∏–±–∫–∏ –ª–æ–≥–∏—Ä—É—é—Ç—Å—è, –Ω–æ –Ω–µ –ø—Ä–µ—Ä—ã–≤–∞—é—Ç –ø—Ä–æ—Ü–µ—Å—Å
    if old_server_id and old_server_id != server_id:
        old_server = await session.get(Server, old_server_id)
        if old_server and old_server.x3ui_api_url and old_server.x3ui_username and old_server.x3ui_password:
            try:
                from core.x3ui_api import X3UIAPI
                import asyncio
                
                # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º —Ç–∞–π–º–∞—É—Ç –¥–ª—è —É–¥–∞–ª–µ–Ω–∏—è –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ (5 —Å–µ–∫—É–Ω–¥)
                async def delete_old_client():
                    x3ui = None
                    try:
                        x3ui = X3UIAPI(
                            api_url=old_server.x3ui_api_url,
                            username=old_server.x3ui_username,
                            password=old_server.x3ui_password,
                        )
                        client_email = f"tg_{user.tg_id}_server_{old_server.id}@fiorevpn"
                        inbound_id = old_server.x3ui_inbound_id
                        if inbound_id:
                            # –ü—ã—Ç–∞–µ–º—Å—è —É–¥–∞–ª–∏—Ç—å —Å —Ç–∞–π–º–∞—É—Ç–æ–º
                            deleted = await asyncio.wait_for(
                                x3ui.delete_client(inbound_id, client_email),
                                timeout=5.0
                            )
                            if deleted:
                                logger.info(f"–£–¥–∞–ª–µ–Ω –∫–ª–∏–µ–Ω—Ç {client_email} —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ {old_server.name}")
                            else:
                                logger.info(f"–ö–ª–∏–µ–Ω—Ç {client_email} –Ω–µ –Ω–∞–π–¥–µ–Ω –Ω–∞ —Å—Ç–∞—Ä–æ–º —Å–µ—Ä–≤–µ—Ä–µ {old_server.name} (–≤–æ–∑–º–æ–∂–Ω–æ, —É–∂–µ —É–¥–∞–ª–µ–Ω)")
                    except asyncio.TimeoutError:
                        logger.warning(f"–¢–∞–π–º–∞—É—Ç –ø—Ä–∏ —É–¥–∞–ª–µ–Ω–∏–∏ –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ {old_server.name}")
                    except Exception as e:
                        logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —É–¥–∞–ª–∏—Ç—å –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ {old_server.name}: {e}")
                    finally:
                        if x3ui:
                            try:
                                await x3ui.close()
                            except:
                                pass
                
                # –ó–∞–ø—É—Å–∫–∞–µ–º —É–¥–∞–ª–µ–Ω–∏–µ –≤ —Ñ–æ–Ω–µ, –Ω–µ –±–ª–æ–∫–∏—Ä—É—è —Å–º–µ–Ω—É —Å–µ—Ä–≤–µ—Ä–∞
                asyncio.create_task(delete_old_client())
                logger.info(f"–ó–∞–ø—É—â–µ–Ω–∞ —Ñ–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ —É–¥–∞–ª–µ–Ω–∏—è –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ {old_server.name}")
            except Exception as e:
                logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –∏–Ω–∏—Ü–∏–∞–ª–∏–∑–∞—Ü–∏–∏ —É–¥–∞–ª–µ–Ω–∏—è –∫–ª–∏–µ–Ω—Ç–∞ —Å–æ —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞ {old_server.name}: {e}")
        
        # –î–µ–∞–∫—Ç–∏–≤–∏—Ä—É–µ–º —Å—Ç–∞—Ä—ã–µ VPN credentials –¥–ª—è —Å—Ç–∞—Ä–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
        old_credentials = await session.scalars(
            select(VpnCredential)
            .where(VpnCredential.user_id == user.id)
            .where(VpnCredential.server_id == old_server_id)
            .where(VpnCredential.active == True)
        )
        for cred in old_credentials:
            cred.active = False
        logger.info(f"–î–µ–∞–∫—Ç–∏–≤–∏—Ä–æ–≤–∞–Ω—ã —Å—Ç–∞—Ä—ã–µ VPN credentials –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id} –¥–ª—è —Å–µ—Ä–≤–µ—Ä–∞ {old_server_id}")
    
    # –£—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞–µ–º –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Å–µ—Ä–≤–µ—Ä
    user.selected_server_id = server_id
    await session.commit()
    
    return {"success": True, "server_id": server_id, "server_name": server.name}


@app.get("/users/{tg_id}/vpn-key")
async def get_user_vpn_key(
    tg_id: int,
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å VPN –∫–ª—é—á –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    if not user.selected_server_id:
        return {"key": None, "server_name": None}
    
    # –ü–æ–ª—É—á–∞–µ–º –∞–∫—Ç–∏–≤–Ω—ã–π –∫–ª—é—á –¥–ª—è –≤—ã–±—Ä–∞–Ω–Ω–æ–≥–æ —Å–µ—Ä–≤–µ—Ä–∞
    credential = await session.scalar(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.server_id == user.selected_server_id)
        .where(VpnCredential.active == True)
        .order_by(VpnCredential.created_at.desc())
    )
    
    server = await session.get(Server, user.selected_server_id)
    server_name = server.name if server else None
    
    if credential and credential.config_text:
        return {"key": credential.config_text, "server_name": server_name}
    
    return {"key": None, "server_name": server_name}


@app.post("/users/{tg_id}/vpn-key/generate")
async def generate_user_vpn_key(
    tg_id: int,
    request: Request,
    session: AsyncSession = Depends(get_session),
):
    """–°–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å VPN –∫–ª—é—á –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    # –ü–æ–ª—É—á–∞–µ–º payload –∏–∑ –∑–∞–ø—Ä–æ—Å–∞
    try:
        payload = await request.json() if request.headers.get("content-type") == "application/json" else {}
    except:
        payload = {}
    
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    if not user.selected_server_id:
        raise HTTPException(status_code=400, detail="server_not_selected")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∞–∫—Ç–∏–≤–Ω—É—é –ø–æ–¥–ø–∏—Å–∫—É
    if not user.has_active_subscription or not user.subscription_ends_at:
        raise HTTPException(status_code=403, detail="no_active_subscription")
    
    from datetime import datetime, timezone
    if user.subscription_ends_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=403, detail="subscription_expired")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–µ—Ä–≤–µ—Ä
    server = await session.get(Server, user.selected_server_id)
    if not server or not server.is_enabled:
        raise HTTPException(status_code=404, detail="server_not_found")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —É–∂–µ –∞–∫—Ç–∏–≤–Ω—ã–π –∫–ª—é—á
    existing_active = await session.scalar(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.server_id == user.selected_server_id)
        .where(VpnCredential.active == True)
        .order_by(VpnCredential.created_at.desc())
    )
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–∞—Ä–∞–º–µ—Ç—Ä regenerate (–¥–ª—è "–°–º–µ–Ω–∏—Ç—å –∫–ª—é—á")
    regenerate = payload.get("regenerate", False) if payload else False
    
    # –ï—Å–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å —É–∂–µ –∏–º–µ–µ—Ç –∞–∫—Ç–∏–≤–Ω—ã–π –∫–ª—é—á –∏ –Ω–µ –∑–∞–ø—Ä–æ—à–µ–Ω–∞ —Ä–µ–≥–µ–Ω–µ—Ä–∞—Ü–∏—è, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º 400
    if existing_active and existing_active.config_text and not regenerate:
        raise HTTPException(status_code=400, detail="user_already_has_key")
    
    # –ï—Å–ª–∏ –∑–∞–ø—Ä–æ—à–µ–Ω–∞ —Ä–µ–≥–µ–Ω–µ—Ä–∞—Ü–∏—è, –¥–µ–∞–∫—Ç–∏–≤–∏—Ä—É–µ–º —Å—Ç–∞—Ä—ã–π –∫–ª—é—á
    if existing_active and regenerate:
        existing_active.active = False
        await session.commit()
    
    # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –Ω–æ–≤—ã–π –∫–ª—é—á
    try:
        import httpx
        await _generate_vpn_config_for_user_server(user.id, user.selected_server_id, session, user.subscription_ends_at)
    except ValueError as e:
        error_msg = str(e)
        # 503 - 3x-UI –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω
        if "–Ω–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å —Å–ø–∏—Å–æ–∫ Inbounds" in error_msg or "–Ω–µ —É–¥–∞–ª–æ—Å—å –ø–æ–ª—É—á–∏—Ç—å" in error_msg.lower() or "–Ω–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è" in error_msg.lower():
            raise HTTPException(status_code=503, detail=f"3x_ui_unavailable: {error_msg}")
        # 400 - –æ—à–∏–±–∫–∞ –∫–æ–Ω—Ñ–∏–≥—É—Ä–∞—Ü–∏–∏ —Å–µ—Ä–≤–µ—Ä–∞
        if "Inbound –Ω–µ –Ω–∞–π–¥–µ–Ω" in error_msg or "–Ω–µ –Ω–∞—Å—Ç—Ä–æ–µ–Ω" in error_msg:
            raise HTTPException(status_code=400, detail=f"server_configuration_error: {error_msg}")
        # 500 - —Ä–µ–∞–ª—å–Ω—ã–π –±–∞–≥
        import logging
        logging.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –∫–ª—é—á–∞ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"key_generation_failed: {error_msg}")
    except httpx.HTTPStatusError as e:
        # 503 - 3x-UI –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω (HTTP –æ—à–∏–±–∫–∏ –æ—Ç API)
        if e.response.status_code in (404, 503, 502, 504):
            raise HTTPException(status_code=503, detail=f"3x_ui_unavailable: HTTP {e.response.status_code}")
        # 500 - –¥—Ä—É–≥–∏–µ HTTP –æ—à–∏–±–∫–∏
        import logging
        logging.error(f"HTTP –æ—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –∫–ª—é—á–∞ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="key_generation_failed")
    except httpx.RequestError as e:
        # 503 - 3x-UI –Ω–µ–¥–æ—Å—Ç—É–ø–µ–Ω (—Å–µ—Ç–µ–≤—ã–µ –æ—à–∏–±–∫–∏)
        import logging
        logging.warning(f"–°–µ—Ç–µ–≤–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –ø–æ–¥–∫–ª—é—á–µ–Ω–∏–∏ –∫ 3x-UI –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: {e}")
        raise HTTPException(status_code=503, detail="3x_ui_unavailable: –ù–µ —É–¥–∞–ª–æ—Å—å –ø–æ–¥–∫–ª—é—á–∏—Ç—å—Å—è –∫ 3x-UI")
    except Exception as e:
        # 500 - —Ç–æ–ª—å–∫–æ —Ä–µ–∞–ª—å–Ω—ã–µ –±–∞–≥–∏
        import logging
        logging.error(f"–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ –∫–ª—é—á–∞ –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="key_generation_failed")
    
    # –ü–æ–ª—É—á–∞–µ–º —Å–æ–∑–¥–∞–Ω–Ω—ã–π –∫–ª—é—á
    credential = await session.scalar(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.server_id == user.selected_server_id)
        .where(VpnCredential.active == True)
        .order_by(VpnCredential.created_at.desc())
    )
    
    if not credential or not credential.config_text:
        import logging
        logging.error(f"–ö–ª—é—á –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω –¥–ª—è –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è {user.tg_id}, —Ö–æ—Ç—è –æ—à–∏–±–æ–∫ –Ω–µ –±—ã–ª–æ")
        raise HTTPException(status_code=500, detail="key_generation_failed")
    
    return {"key": credential.config_text, "server_name": server.name}


@app.get("/users/{tg_id}/vpn-configs")
async def get_user_vpn_configs(
    tg_id: int,
    session: AsyncSession = Depends(get_session),
):
    """–ü–æ–ª—É—á–∏—Ç—å VPN –∫–æ–Ω—Ñ–∏–≥–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è"""
    user = await session.scalar(select(User).where(User.tg_id == tg_id))
    if not user:
        raise HTTPException(status_code=404, detail="user_not_found")
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –∞–∫—Ç–∏–≤–Ω–∞—è –ø–æ–¥–ø–∏—Å–∫–∞
    if not user.has_active_subscription:
        raise HTTPException(status_code=403, detail="no_active_subscription")
    
    # –ü–æ–ª—É—á–∞–µ–º –∞–∫—Ç–∏–≤–Ω—ã–µ –∫–æ–Ω—Ñ–∏–≥–∏
    credentials = await session.scalars(
        select(VpnCredential)
        .where(VpnCredential.user_id == user.id)
        .where(VpnCredential.active == True)
        .options(selectinload(VpnCredential.server))
    )
    configs_list = credentials.all()
    
    configs = []
    for cred in configs_list:
        configs.append({
            "id": cred.id,
            "server_name": cred.server.name if cred.server else None,
            "config": cred.config_text,
            "expires_at": cred.expires_at.isoformat() if cred.expires_at else None,
        })
    
    return {"configs": configs}


# Catch-all —Ä–æ—É—Ç –¥–ª—è –≤—Å–µ—Ö –Ω–µ–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã—Ö –ø—É—Ç–µ–π (–¥–æ–ª–∂–µ–Ω –±—ã—Ç—å –ø–æ—Å–ª–µ–¥–Ω–∏–º)
@app.get("/{path:path}")
async def catch_all(request: Request, path: str):
    """–ü–µ—Ä–µ—Ö–≤–∞—Ç—ã–≤–∞–µ—Ç –≤—Å–µ –Ω–µ–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–µ GET –∑–∞–ø—Ä–æ—Å—ã –∏ –ø–æ–∫–∞–∑—ã–≤–∞–µ—Ç 404"""
    url_path = request.url.path
    
    # –î–ª—è API endpoints –≤–æ–∑–≤—Ä–∞—â–∞–µ–º JSON (–∏—Å–∫–ª—é—á–∞–µ–º /admin/web –ø—É—Ç–∏)
    is_api_path = (
        url_path.startswith("/api/") or 
        url_path.startswith("/subscriptions/") or 
        url_path.startswith("/payments/") or 
        (url_path.startswith("/users/") and not url_path.startswith("/admin/web")) or 
        url_path.startswith("/promo-codes/") or
        (url_path.startswith("/tickets/") and not url_path.startswith("/admin/web")) or
        url_path.startswith("/health") or
        url_path.startswith("/support/")
    )
    
    if is_api_path:
        raise HTTPException(status_code=404, detail="Not found")
    
    # –î–ª—è –≤—Å–µ—Ö –æ—Å—Ç–∞–ª—å–Ω—ã—Ö –ø—É—Ç–µ–π (–≤–∫–ª—é—á–∞—è /admin/web/*) –ø–æ–∫–∞–∑—ã–≤–∞–µ–º –∫—Ä–∞—Å–∏–≤—É—é 404
    if templates:
        return templates.TemplateResponse(
            "404.html",
            {"request": request},
            status_code=404
        )
    # –ï—Å–ª–∏ —à–∞–±–ª–æ–Ω—ã –Ω–µ –∑–∞–≥—Ä—É–∂–µ–Ω—ã, –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –ø—Ä–æ—Å—Ç–æ–π —Ç–µ–∫—Å—Ç
    return HTMLResponse(
        content="<h1>404 - –°—Ç—Ä–∞–Ω–∏—Ü–∞ –Ω–µ –Ω–∞–π–¥–µ–Ω–∞</h1><p><a href='/admin/login'>–ü–µ—Ä–µ–π—Ç–∏ –≤ –∞–¥–º–∏–Ω-–ø–∞–Ω–µ–ª—å</a></p>",
        status_code=404
    )

